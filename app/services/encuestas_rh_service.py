# app/services/encuestas_rh_service.py
"""Logica de negocio del modulo Encuestas RH (Level Up).

Responsabilidades:
  - CRUD de encuesta/preguntas/opciones (solo en estado borrador).
  - Ciclo de vida: publicar (materializa audiencia + notifica), cerrar
    (manual o automatico via `procesar_cierres_vencidos`, sin reapertura).
  - Resolucion de audiencia por area/turno/rol (preview + materializacion).
  - Responder con soporte de anonimato (desasocia `empleado_id` del "sobre"
    de respuestas cuando la encuesta es anonima).
  - Plantillas predefinidas -> nueva encuesta en borrador.

El commit lo realiza la dependencia `get_db` al cierre del request; aqui solo
se usa flush() (via el repositorio), como el resto de services del proyecto.
"""

from __future__ import annotations

import random
import re
import unicodedata
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Optional
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import (
    ConflictError,
    DomainValidationError,
    ForbiddenError,
    NotFoundError,
)
from app.models.empleados import Empleado
from app.models.encuestas_rh import (
    Encuesta,
    EncuestaOpcion,
    EncuestaParticipante,
    EncuestaPregunta,
    EncuestaRespuesta,
    EncuestaRespuestaGrupo,
    EncuestaRespuestaOpcion,
)
from app.repositories.encuestas_rh_repository import EncuestasRhRepository
from app.schemas.encuestas_rh import (
    AudienciaAreaConteo,
    AudienciaFiltros,
    AudienciaPreview,
    AudienciaTurnoConteo,
    DistribucionLikert,
    EncuestaCreate,
    EncuestaResponse,
    EncuestaUpdate,
    MiEncuestaItem,
    OpcionConteo,
    OpcionResponse,
    ParticipanteItem,
    PlantillaResponse,
    PreguntaCreate,
    PreguntaResponse,
    PreguntaUpdate,
    PublicarRequest,
    RecordatoriosResultado,
    ResponderRequest,
    ResultadoPregunta,
    ResultadosGlobal,
    ResultadosSegmentos,
    SegmentoCelda,
    TextosResponse,
)
from app.services.notificacion_service import NotificacionService

MIS_ENCUESTAS_TARGET_URL = "#/talento/mis-encuestas"

# Campos editables de la encuesta una vez publicada (el resto es inmutable).
_CAMPOS_EDITABLES_PUBLICADA = {"titulo", "descripcion", "fecha_cierre_programada"}

# Dimensiones validas para /resultados/segmentos (parametro `dimension`).
DIMENSIONES_SEGMENTO = ("area", "turno", "clasificacion")

SIN_DATO = "Sin dato"


class EncuestasRhService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = EncuestasRhRepository(db)
        self.notificaciones = NotificacionService(db)

    # ══════════════════════════════════════════════════════════════════════
    # CRUD encuesta (borrador)
    # ══════════════════════════════════════════════════════════════════════
    async def crear_encuesta(self, data: EncuestaCreate) -> EncuestaResponse:
        if data.creado_por_id is None:
            raise DomainValidationError("creado_por_id es obligatorio para crear una encuesta")

        encuesta = Encuesta(
            titulo=data.titulo,
            descripcion=data.descripcion,
            tipo=data.tipo,
            es_anonima=data.es_anonima,
            umbral_minimo_respuestas=data.umbral_minimo_respuestas,
            recordatorio_cada_dias=data.recordatorio_cada_dias,
            creado_por_id=data.creado_por_id,
            estado="borrador",
        )
        self.db.add(encuesta)
        await self.db.flush()

        for pregunta_data in data.preguntas:
            self._crear_pregunta_obj(encuesta.id, pregunta_data)
        await self.db.flush()

        return await self.obtener_encuesta(encuesta.id)

    async def obtener_encuesta(self, encuesta_id: int) -> EncuestaResponse:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        return self._encuesta_to_response(encuesta)

    async def listar_encuestas(self, estado: Optional[str] = None) -> list[EncuestaResponse]:
        encuestas = await self.repo.list_encuestas(estado=estado)
        return [self._encuesta_to_response(e) for e in encuestas]

    async def actualizar_encuesta(
        self, encuesta_id: int, data: EncuestaUpdate
    ) -> EncuestaResponse:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado == "cerrada":
            raise ConflictError("No se puede editar una encuesta cerrada")

        payload = data.model_dump(exclude_unset=True)

        if encuesta.estado == "publicada":
            campos_no_editables = set(payload) - _CAMPOS_EDITABLES_PUBLICADA
            if campos_no_editables:
                raise DomainValidationError(
                    "Una encuesta publicada solo permite editar titulo, "
                    "descripcion y extender fecha_cierre_programada "
                    f"(campos no permitidos: {', '.join(sorted(campos_no_editables))})"
                )
            if "fecha_cierre_programada" in payload:
                nueva_fecha = payload["fecha_cierre_programada"]
                actual = encuesta.fecha_cierre_programada
                if actual is not None and nueva_fecha == actual:
                    # No-op: el cliente puede reenviar la fecha actual sin
                    # intencion de cambiarla (p. ej. al editar solo el titulo).
                    payload.pop("fecha_cierre_programada")
                elif nueva_fecha is None or (actual is not None and nueva_fecha <= actual):
                    raise DomainValidationError(
                        "La nueva fecha_cierre_programada debe ser posterior a la actual"
                    )

        for key, value in payload.items():
            setattr(encuesta, key, value)
        await self.db.flush()
        return await self.obtener_encuesta(encuesta.id)

    async def eliminar_encuesta(self, encuesta_id: int) -> None:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado != "borrador":
            raise ConflictError("Solo se pueden borrar encuestas en borrador")
        await self.db.delete(encuesta)
        await self.db.flush()

    # ── Preguntas / opciones (solo borrador) ───────────────────────────────
    def _crear_pregunta_obj(self, encuesta_id: int, data: PreguntaCreate) -> EncuestaPregunta:
        pregunta = EncuestaPregunta(
            encuesta_id=encuesta_id,
            orden=data.orden,
            tipo=data.tipo,
            texto=data.texto,
            requerida=data.requerida,
            seleccion_multiple=data.seleccion_multiple,
            # `opciones=[]` explicito: marca la coleccion como "ya cargada" en
            # el identity map. Sin esto, `agregar_pregunta` (que NO recarga
            # via el repo con selectinload) dispara un lazy-load real al
            # serializar `pregunta.opciones` tras el flush si la pregunta no
            # trae opciones (likert/texto) -> MissingGreenlet en async.
            opciones=[],
        )
        self.db.add(pregunta)
        for opcion_data in data.opciones:
            pregunta.opciones.append(
                EncuestaOpcion(texto=opcion_data.texto, orden=opcion_data.orden)
            )
        return pregunta

    async def _get_encuesta_borrador(self, encuesta_id: int) -> Encuesta:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado != "borrador":
            raise ConflictError(
                "Las preguntas y opciones solo se pueden modificar en encuestas en borrador"
            )
        return encuesta

    async def agregar_pregunta(
        self, encuesta_id: int, data: PreguntaCreate
    ) -> PreguntaResponse:
        await self._get_encuesta_borrador(encuesta_id)
        pregunta = self._crear_pregunta_obj(encuesta_id, data)
        await self.db.flush()
        return self._pregunta_to_response(pregunta)

    async def actualizar_pregunta(
        self, encuesta_id: int, pregunta_id: int, data: PreguntaUpdate
    ) -> PreguntaResponse:
        encuesta = await self._get_encuesta_borrador(encuesta_id)
        pregunta = next((p for p in encuesta.preguntas if p.id == pregunta_id), None)
        if not pregunta:
            raise NotFoundError("Pregunta", pregunta_id)

        payload = data.model_dump(exclude_unset=True, exclude={"opciones"})
        for key, value in payload.items():
            setattr(pregunta, key, value)

        if data.opciones is not None:
            # Se muta la coleccion ORM (no `db.delete`/`db.add` sueltos) para que
            # el cascade "delete-orphan" dispare los DELETE/INSERT y, sobre todo,
            # para que `pregunta.opciones` quede sincronizada en el identity map
            # (si no, una lectura posterior en la misma sesion devuelve la
            # coleccion vieja cacheada en vez del estado recien escrito).
            for opcion in list(pregunta.opciones):
                pregunta.opciones.remove(opcion)
            for opcion_data in data.opciones:
                pregunta.opciones.append(
                    EncuestaOpcion(texto=opcion_data.texto, orden=opcion_data.orden)
                )
        await self.db.flush()

        return self._pregunta_to_response(pregunta)

    async def eliminar_pregunta(self, encuesta_id: int, pregunta_id: int) -> None:
        encuesta = await self._get_encuesta_borrador(encuesta_id)
        pregunta = next((p for p in encuesta.preguntas if p.id == pregunta_id), None)
        if not pregunta:
            raise NotFoundError("Pregunta", pregunta_id)
        # Mutar la coleccion del padre (no `db.delete` suelto) para que el
        # cascade "delete-orphan" dispare el DELETE y `encuesta.preguntas`
        # quede sincronizada en el identity map (ver nota en actualizar_pregunta).
        encuesta.preguntas.remove(pregunta)
        await self.db.flush()

    async def reordenar_preguntas(
        self, encuesta_id: int, pregunta_ids: list[int]
    ) -> list[PreguntaResponse]:
        """Reasigna `orden` segun la posicion de cada id en `pregunta_ids`.

        La lista debe cubrir exactamente el conjunto de preguntas existentes
        de la encuesta (ni de mas ni de menos) para evitar huecos/ambiguedad
        de orden.
        """
        encuesta = await self._get_encuesta_borrador(encuesta_id)
        preguntas_por_id = {p.id: p for p in encuesta.preguntas}
        if set(pregunta_ids) != set(preguntas_por_id.keys()):
            raise DomainValidationError(
                "pregunta_ids debe incluir exactamente todas las preguntas de la encuesta"
            )
        for orden, pregunta_id in enumerate(pregunta_ids, start=1):
            preguntas_por_id[pregunta_id].orden = orden
        await self.db.flush()
        return [
            self._pregunta_to_response(p)
            for p in sorted(encuesta.preguntas, key=lambda p: (p.orden, p.id))
        ]

    # ══════════════════════════════════════════════════════════════════════
    # Audiencia
    # ══════════════════════════════════════════════════════════════════════
    async def _resolver_audiencia(self, filtros: AudienciaFiltros) -> list[Empleado]:
        candidatos = await self.repo.list_empleados_activos()

        areas = set(filtros.areas) if filtros.areas else None
        turnos = {t.strip().upper() for t in filtros.turnos if t.strip()} if filtros.turnos else None
        roles = {r.strip().lower() for r in filtros.roles if r.strip()} if filtros.roles else None

        resultado: list[Empleado] = []
        for empleado in candidatos:
            if areas is not None and empleado.area_id not in areas:
                continue
            if roles is not None:
                rol = empleado.rol
                rol_nombre = rol.nombre.strip().lower() if rol else None
                if rol_nombre not in roles:
                    continue
            if turnos is not None:
                turno_valor = self._turno_normalizado(empleado)
                if turno_valor is None or turno_valor not in turnos:
                    continue
            resultado.append(empleado)
        return resultado

    @staticmethod
    def _turno_normalizado(empleado: Empleado) -> Optional[str]:
        te = empleado.turno_empleado
        if not te or not te.turno:
            return None
        valor = te.turno.strip().upper()
        return valor or None

    async def preview_audiencia(self, filtros: AudienciaFiltros) -> AudienciaPreview:
        empleados = await self._resolver_audiencia(filtros)

        por_area: dict[tuple[Optional[int], Optional[str]], int] = {}
        por_turno: dict[Optional[str], int] = {}
        for empleado in empleados:
            area_key = (
                empleado.area_id,
                empleado.area.descripcion if empleado.area else None,
            )
            por_area[area_key] = por_area.get(area_key, 0) + 1

            turno_valor = self._turno_normalizado(empleado)
            por_turno[turno_valor] = por_turno.get(turno_valor, 0) + 1

        return AudienciaPreview(
            total=len(empleados),
            por_area=[
                AudienciaAreaConteo(area_id=area_id, area_nombre=area_nombre, total=total)
                for (area_id, area_nombre), total in por_area.items()
            ],
            por_turno=[
                AudienciaTurnoConteo(turno=turno, total=total)
                for turno, total in por_turno.items()
            ],
        )

    async def _materializar_participantes(
        self, encuesta_id: int, empleados: list[Empleado]
    ) -> list[Empleado]:
        """Crea EncuestaParticipante en pendiente por cada empleado que aun no
        lo es (idempotente). Devuelve los empleados recien materializados."""
        existentes = await self.repo.list_empleado_ids_participantes(encuesta_id)
        nuevos: list[Empleado] = []
        for empleado in empleados:
            if empleado.empleado_id in existentes:
                continue
            self.db.add(
                EncuestaParticipante(
                    encuesta_id=encuesta_id,
                    empleado_id=empleado.empleado_id,
                    estado="pendiente",
                    # Momento de la convocatoria inicial: base para la cadencia
                    # de recordatorios cuando aun no hay ultimo_recordatorio_at
                    # (ver EncuestasRhService.procesar_recordatorios).
                    notificado_at=datetime.now(timezone.utc),
                )
            )
            nuevos.append(empleado)
        await self.db.flush()
        return nuevos

    # ══════════════════════════════════════════════════════════════════════
    # Ciclo de vida
    # ══════════════════════════════════════════════════════════════════════
    async def publicar_encuesta(
        self, encuesta_id: int, data: PublicarRequest
    ) -> EncuestaResponse:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado != "borrador":
            raise ConflictError("Solo las encuestas en borrador se pueden publicar")

        if not encuesta.preguntas:
            raise DomainValidationError("La encuesta debe tener al menos una pregunta")
        for pregunta in encuesta.preguntas:
            if pregunta.tipo == "opcion_multiple" and len(pregunta.opciones) < 2:
                raise DomainValidationError(
                    f"La pregunta de opcion multiple '{pregunta.texto}' "
                    "debe tener al menos 2 opciones"
                )

        if data.fecha_cierre_programada <= date.today():
            raise DomainValidationError("fecha_cierre_programada debe ser una fecha futura")

        empleados = await self._resolver_audiencia(data.filtros)
        if not empleados:
            raise DomainValidationError("La audiencia resuelta con esos filtros esta vacia")

        encuesta.audiencia_criterios = data.filtros.model_dump()
        encuesta.fecha_publicacion = datetime.now(timezone.utc)
        encuesta.fecha_cierre_programada = data.fecha_cierre_programada
        encuesta.estado = "publicada"
        await self.db.flush()

        nuevos = await self._materializar_participantes(encuesta.id, empleados)

        for empleado in nuevos:
            await self.notificaciones.enviar(
                destinatario_id=empleado.empleado_id,
                asunto=f"Nueva encuesta: {encuesta.titulo}",
                cuerpo=(
                    f"Se te ha invitado a responder la encuesta '{encuesta.titulo}'. "
                    "Tu participacion es importante."
                ),
                canal="in_app",
                target_url=MIS_ENCUESTAS_TARGET_URL,
                metadata={"encuesta_id": encuesta.id},
            )

        return await self.obtener_encuesta(encuesta.id)

    async def cerrar_encuesta(self, encuesta_id: int) -> EncuestaResponse:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado != "publicada":
            raise ConflictError("Solo las encuestas publicadas se pueden cerrar")
        encuesta.estado = "cerrada"
        encuesta.fecha_cierre_real = datetime.now(timezone.utc)
        await self.db.flush()
        return await self.obtener_encuesta(encuesta.id)

    async def procesar_cierres_vencidos(self) -> int:
        """Cierra automaticamente las encuestas publicadas cuya fecha de
        cierre programada ya paso. Pensado para invocarse desde un botón
        manual / CLI (no hay job de scheduler para este modulo)."""
        hoy = date.today()
        vencidas = await self.repo.list_publicadas_vencidas(hoy)
        for encuesta in vencidas:
            encuesta.estado = "cerrada"
            encuesta.fecha_cierre_real = datetime.now(timezone.utc)
        await self.db.flush()
        return len(vencidas)

    # ══════════════════════════════════════════════════════════════════════
    # Recordatorios automaticos (invocado por APScheduler, ver
    # `_encuestas_rh_recordatorios_job` en app/main.py) + endpoint manual
    # ══════════════════════════════════════════════════════════════════════
    @staticmethod
    def _dt_utc(value: Optional[datetime]) -> Optional[datetime]:
        """Normaliza un datetime leido de BD a aware UTC.

        Postgres/asyncpg devuelve datetimes aware para columnas timestamptz;
        SQLite (tests) los devuelve naive. Todos los timestamps de este modulo
        se escriben con `datetime.now(timezone.utc)`, asi que asumir UTC en el
        caso naive es seguro."""
        if value is None:
            return None
        return value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)

    async def _notificar_recordatorio(
        self, encuesta: Encuesta, participante: EncuestaParticipante
    ) -> None:
        await self.notificaciones.enviar(
            destinatario_id=participante.empleado_id,
            asunto=f"Recordatorio: encuesta pendiente '{encuesta.titulo}'",
            cuerpo=(
                f"Aun no has respondido la encuesta '{encuesta.titulo}'. "
                "Tu participacion es importante."
            ),
            canal="in_app",
            target_url=MIS_ENCUESTAS_TARGET_URL,
            metadata={"encuesta_id": encuesta.id},
        )
        participante.ultimo_recordatorio_at = datetime.now(timezone.utc)
        participante.recordatorios_enviados += 1

    async def procesar_recordatorios(self) -> RecordatoriosResultado:
        """Cierra encuestas publicadas vencidas (reusa `procesar_cierres_vencidos`,
        misma regla, sin duplicarla) y notifica a los participantes `pendiente`
        de las encuestas que siguen publicadas, respetando la cadencia
        `recordatorio_cada_dias` de cada encuesta:
          - `ultimo_recordatorio_at` no nulo: se remite si pasaron >= N dias
            desde el ultimo recordatorio.
          - `ultimo_recordatorio_at` nulo: se remite si pasaron >= N dias desde
            `notificado_at` (la convocatoria inicial al publicar).
          - Si ninguno de los dos esta poblado, no se notifica (no hay
            referencia temporal)."""
        encuestas_cerradas = await self.procesar_cierres_vencidos()

        ahora = datetime.now(timezone.utc)
        recordatorios_enviados = 0
        for encuesta in await self.repo.list_encuestas(estado="publicada"):
            cadencia = timedelta(days=encuesta.recordatorio_cada_dias)
            for participante in await self.repo.list_participantes(encuesta.id):
                if participante.estado != "pendiente":
                    continue
                referencia = self._dt_utc(participante.ultimo_recordatorio_at) or self._dt_utc(
                    participante.notificado_at
                )
                if referencia is None or ahora - referencia < cadencia:
                    continue
                await self._notificar_recordatorio(encuesta, participante)
                recordatorios_enviados += 1

        await self.db.flush()
        return RecordatoriosResultado(
            encuestas_cerradas=encuestas_cerradas,
            recordatorios_enviados=recordatorios_enviados,
        )

    async def forzar_recordatorios(self, encuesta_id: int) -> int:
        """Endpoint manual de gestion: fuerza un recordatorio a TODOS los
        participantes `pendiente` de la encuesta, sin respetar la cadencia
        `recordatorio_cada_dias`. Requiere que la encuesta este publicada.
        Devuelve el numero de recordatorios enviados."""
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado != "publicada":
            raise ConflictError(
                "Solo se pueden enviar recordatorios de encuestas publicadas"
            )

        enviados = 0
        for participante in await self.repo.list_participantes(encuesta_id):
            if participante.estado != "pendiente":
                continue
            await self._notificar_recordatorio(encuesta, participante)
            enviados += 1

        await self.db.flush()
        return enviados

    # ══════════════════════════════════════════════════════════════════════
    # Responder
    # ══════════════════════════════════════════════════════════════════════
    async def responder(
        self, encuesta_id: int, empleado_id: int, payload: ResponderRequest
    ) -> None:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)

        participante = await self.repo.get_participante(encuesta_id, empleado_id)
        if not participante:
            raise ForbiddenError("No eres participante de esta encuesta")

        if encuesta.estado != "publicada":
            raise ConflictError("La encuesta no esta publicada")
        if participante.estado == "respondida":
            raise ConflictError("Ya respondiste esta encuesta")

        preguntas_por_id = {p.id: p for p in encuesta.preguntas}
        respondidas_ids: set[int] = set()

        for item in payload.respuestas:
            pregunta = preguntas_por_id.get(item.pregunta_id)
            if not pregunta:
                raise DomainValidationError(
                    f"La pregunta {item.pregunta_id} no pertenece a esta encuesta"
                )

            if pregunta.tipo == "likert":
                if item.valor_likert is None:
                    continue
                if not (1 <= item.valor_likert <= 5):
                    raise DomainValidationError(
                        f"valor_likert fuera de rango (1..5) en pregunta {pregunta.id}"
                    )
                respondidas_ids.add(pregunta.id)
            elif pregunta.tipo == "texto":
                if item.texto is None or not item.texto.strip():
                    continue
                respondidas_ids.add(pregunta.id)
            elif pregunta.tipo == "opcion_multiple":
                if not item.opcion_ids:
                    continue
                opciones_validas = {o.id for o in pregunta.opciones}
                for opcion_id in item.opcion_ids:
                    if opcion_id not in opciones_validas:
                        raise DomainValidationError(
                            f"La opcion {opcion_id} no pertenece a la pregunta {pregunta.id}"
                        )
                if not pregunta.seleccion_multiple and len(item.opcion_ids) != 1:
                    raise DomainValidationError(
                        f"La pregunta {pregunta.id} acepta exactamente 1 opcion"
                    )
                respondidas_ids.add(pregunta.id)

        for pregunta in encuesta.preguntas:
            if pregunta.requerida and pregunta.id not in respondidas_ids:
                raise DomainValidationError(
                    f"Falta responder la pregunta requerida: {pregunta.texto}"
                )

        empleado = participante.empleado
        es_anonima = encuesta.es_anonima
        grupo_id = uuid4()
        grupo = EncuestaRespuestaGrupo(
            id=grupo_id,
            encuesta_id=encuesta_id,
            empleado_id=None if es_anonima else empleado_id,
            segmento_area=empleado.area.descripcion if empleado and empleado.area else None,
            segmento_turno=self._turno_normalizado(empleado) if empleado else None,
            segmento_clasificacion=(
                empleado.clasificacion.descripcion
                if empleado and empleado.clasificacion
                else None
            ),
            fecha_dia=date.today(),
            created_at=None if es_anonima else datetime.now(timezone.utc),
        )
        self.db.add(grupo)

        respuestas_creadas: list[tuple[EncuestaRespuesta, list[int]]] = []
        for item in payload.respuestas:
            pregunta = preguntas_por_id.get(item.pregunta_id)
            if not pregunta or pregunta.id not in respondidas_ids:
                continue
            respuesta = EncuestaRespuesta(
                grupo_id=grupo_id,
                pregunta_id=pregunta.id,
                valor_likert=item.valor_likert if pregunta.tipo == "likert" else None,
                texto=item.texto if pregunta.tipo == "texto" else None,
            )
            self.db.add(respuesta)
            respuestas_creadas.append((respuesta, item.opcion_ids or []))

        await self.db.flush()

        for respuesta, opcion_ids in respuestas_creadas:
            for opcion_id in opcion_ids:
                self.db.add(
                    EncuestaRespuestaOpcion(respuesta_id=respuesta.id, opcion_id=opcion_id)
                )

        participante.estado = "respondida"
        participante.fecha_respuesta = datetime.now(timezone.utc)
        await self.db.flush()

    # ══════════════════════════════════════════════════════════════════════
    # Mis encuestas
    # ══════════════════════════════════════════════════════════════════════
    async def listar_mis_encuestas(self, empleado_id: int) -> list[MiEncuestaItem]:
        participaciones = await self.repo.list_participaciones_empleado(empleado_id)
        return [
            MiEncuestaItem(
                encuesta_id=p.encuesta_id,
                titulo=p.encuesta.titulo,
                tipo=p.encuesta.tipo,
                estado=p.encuesta.estado,
                participante_estado=p.estado,
                fecha_respuesta=p.fecha_respuesta,
                fecha_cierre_programada=p.encuesta.fecha_cierre_programada,
                es_anonima=p.encuesta.es_anonima,
            )
            for p in participaciones
        ]

    async def obtener_para_responder(
        self, encuesta_id: int, empleado_id: int
    ) -> EncuestaResponse:
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        participante = await self.repo.get_participante(encuesta_id, empleado_id)
        if not participante:
            raise ForbiddenError("No eres participante de esta encuesta")
        if encuesta.estado != "publicada":
            raise ConflictError("La encuesta no esta publicada")
        return self._encuesta_to_response(encuesta)

    async def listar_participantes(self, encuesta_id: int) -> list[ParticipanteItem]:
        encuesta = await self.repo.get(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        participantes = await self.repo.list_participantes(encuesta_id)
        return [
            ParticipanteItem(
                empleado_id=p.empleado_id,
                empleado_nombre=p.empleado.nombre if p.empleado else None,
                estado=p.estado,
                fecha_respuesta=p.fecha_respuesta,
            )
            for p in participantes
        ]

    # ══════════════════════════════════════════════════════════════════════
    # Plantillas
    # ══════════════════════════════════════════════════════════════════════
    async def listar_plantillas(self) -> list[PlantillaResponse]:
        plantillas = await self.repo.list_plantillas()
        return [PlantillaResponse.model_validate(p) for p in plantillas]

    async def crear_encuesta_desde_plantilla(
        self, plantilla_id: int, creado_por_id: Optional[int], es_anonima: bool = True
    ) -> EncuestaResponse:
        if creado_por_id is None:
            raise DomainValidationError("creado_por_id es obligatorio para crear una encuesta")

        plantilla = await self.repo.get_plantilla(plantilla_id)
        if not plantilla:
            raise NotFoundError("Plantilla", plantilla_id)

        encuesta = Encuesta(
            titulo=plantilla.nombre,
            descripcion=plantilla.descripcion,
            tipo=plantilla.tipo or "otra",
            es_anonima=es_anonima,
            estado="borrador",
            creado_por_id=creado_por_id,
        )
        self.db.add(encuesta)
        await self.db.flush()

        for pregunta_def in plantilla.definicion:
            pregunta = EncuestaPregunta(
                encuesta_id=encuesta.id,
                orden=pregunta_def.get("orden", 1),
                tipo=pregunta_def["tipo"],
                texto=pregunta_def["texto"],
                requerida=pregunta_def.get("requerida", True),
                seleccion_multiple=pregunta_def.get("seleccion_multiple", False),
            )
            self.db.add(pregunta)
            await self.db.flush()
            for orden, texto_opcion in enumerate(pregunta_def.get("opciones") or [], start=1):
                self.db.add(
                    EncuestaOpcion(pregunta_id=pregunta.id, texto=texto_opcion, orden=orden)
                )
        await self.db.flush()

        return await self.obtener_encuesta(encuesta.id)

    # ══════════════════════════════════════════════════════════════════════
    # Resultados / analitica (Tarea 4)
    # ══════════════════════════════════════════════════════════════════════
    async def _get_encuesta_con_resultados(self, encuesta_id: int) -> Encuesta:
        """Encuesta publicada o cerrada (borrador -> ConflictError, mismo
        patron que el resto del ciclo de vida: conflicto de estado -> 409)."""
        encuesta = await self.repo.get_detalle(encuesta_id)
        if not encuesta:
            raise NotFoundError("Encuesta", encuesta_id)
        if encuesta.estado == "borrador":
            raise ConflictError(
                "Una encuesta en borrador aun no tiene resultados (publicala primero)"
            )
        return encuesta

    async def _resultado_pregunta_global(self, pregunta: EncuestaPregunta) -> ResultadoPregunta:
        if pregunta.tipo == "likert":
            promedio, n = await self.repo.likert_stats_global(pregunta.id)
            distribucion_dict = await self.repo.likert_distribucion_global(pregunta.id)
            return ResultadoPregunta(
                pregunta_id=pregunta.id,
                tipo=pregunta.tipo,
                texto=pregunta.texto,
                n=n,
                promedio=round(promedio, 2) if promedio is not None else None,
                distribucion=[
                    DistribucionLikert(valor=v, conteo=distribucion_dict.get(v, 0))
                    for v in range(1, 6)
                ],
            )
        if pregunta.tipo == "opcion_multiple":
            n = await self.repo.respuesta_count_global(pregunta.id)
            conteos = await self.repo.opcion_conteos_global(pregunta.id)
            return ResultadoPregunta(
                pregunta_id=pregunta.id,
                tipo=pregunta.tipo,
                texto=pregunta.texto,
                n=n,
                opciones=[
                    OpcionConteo(opcion_id=o.id, texto=o.texto, conteo=conteos.get(o.id, 0))
                    for o in sorted(pregunta.opciones, key=lambda o: (o.orden or 0, o.id))
                ],
            )
        # texto: solo el conteo de respuestas no vacias (el contenido va en
        # /resultados/textos, sujeto ademas a shuffle).
        n = await self.repo.respuesta_count_global(pregunta.id)
        return ResultadoPregunta(pregunta_id=pregunta.id, tipo=pregunta.tipo, texto=pregunta.texto, n=n)

    async def obtener_resultados_globales(self, encuesta_id: int) -> ResultadosGlobal:
        encuesta = await self._get_encuesta_con_resultados(encuesta_id)

        n_global = await self.repo.count_grupos_respuesta(encuesta_id)
        total_participantes = await self.repo.count_participantes(encuesta_id)
        respondidos = await self.repo.count_participantes_respondidos(encuesta_id)
        tasa_respuesta = (
            round(respondidos / total_participantes * 100, 1) if total_participantes else 0.0
        )

        # Regla min-N: solo las encuestas anonimas ocultan el global (en
        # nominales el vinculo con el empleado ya existe por diseño).
        oculto_global = encuesta.es_anonima and n_global < encuesta.umbral_minimo_respuestas

        preguntas: list[ResultadoPregunta] = []
        if not oculto_global:
            for pregunta in sorted(encuesta.preguntas, key=lambda p: (p.orden, p.id)):
                preguntas.append(await self._resultado_pregunta_global(pregunta))

        return ResultadosGlobal(
            encuesta_id=encuesta.id,
            titulo=encuesta.titulo,
            es_anonima=encuesta.es_anonima,
            estado=encuesta.estado,
            umbral_minimo_respuestas=encuesta.umbral_minimo_respuestas,
            n=n_global,
            total_participantes=total_participantes,
            tasa_respuesta=tasa_respuesta,
            oculto_global=oculto_global,
            preguntas=preguntas,
        )

    async def obtener_resultados_segmentos(
        self, encuesta_id: int, dimension: str
    ) -> ResultadosSegmentos:
        if dimension not in DIMENSIONES_SEGMENTO:
            raise DomainValidationError(
                f"dimension invalida: {dimension!r} (validas: {', '.join(DIMENSIONES_SEGMENTO)})"
            )
        encuesta = await self._get_encuesta_con_resultados(encuesta_id)

        conteos = await self.repo.count_grupos_por_segmento(encuesta_id, dimension)

        # Metricas por pregunta precalculadas UNA vez para todos los
        # segmentos (evita N consultas por celda).
        metricas: dict[int, dict] = {}
        for pregunta in encuesta.preguntas:
            if pregunta.tipo == "likert":
                metricas[pregunta.id] = {
                    "stats": await self.repo.likert_stats_por_segmento(
                        encuesta_id, pregunta.id, dimension
                    ),
                    "dist": await self.repo.likert_distribucion_por_segmento(
                        encuesta_id, pregunta.id, dimension
                    ),
                }
            elif pregunta.tipo == "opcion_multiple":
                metricas[pregunta.id] = {
                    "n": await self.repo.respuesta_count_por_segmento(
                        encuesta_id, pregunta.id, dimension
                    ),
                    "opciones": await self.repo.opcion_conteos_por_segmento(
                        encuesta_id, pregunta.id, dimension
                    ),
                }
            else:
                metricas[pregunta.id] = {
                    "n": await self.repo.respuesta_count_por_segmento(
                        encuesta_id, pregunta.id, dimension
                    ),
                }

        celdas: list[SegmentoCelda] = []
        for valor_crudo, n in sorted(
            conteos.items(), key=lambda kv: (kv[0] is None, kv[0] or "")
        ):
            nombre = valor_crudo if valor_crudo is not None else SIN_DATO
            if n < encuesta.umbral_minimo_respuestas:
                celdas.append(SegmentoCelda(segmento=nombre, n=n, oculto=True))
                continue

            preguntas: list[ResultadoPregunta] = []
            for pregunta in sorted(encuesta.preguntas, key=lambda p: (p.orden, p.id)):
                m = metricas[pregunta.id]
                if pregunta.tipo == "likert":
                    promedio, cnt = m["stats"].get(valor_crudo, (None, 0))
                    dist_dict = m["dist"].get(valor_crudo, {})
                    preguntas.append(
                        ResultadoPregunta(
                            pregunta_id=pregunta.id,
                            tipo=pregunta.tipo,
                            texto=pregunta.texto,
                            n=cnt,
                            promedio=round(promedio, 2) if promedio is not None else None,
                            distribucion=[
                                DistribucionLikert(valor=v, conteo=dist_dict.get(v, 0))
                                for v in range(1, 6)
                            ],
                        )
                    )
                elif pregunta.tipo == "opcion_multiple":
                    cnt = m["n"].get(valor_crudo, 0)
                    conteos_opciones = m["opciones"].get(valor_crudo, {})
                    preguntas.append(
                        ResultadoPregunta(
                            pregunta_id=pregunta.id,
                            tipo=pregunta.tipo,
                            texto=pregunta.texto,
                            n=cnt,
                            opciones=[
                                OpcionConteo(
                                    opcion_id=o.id,
                                    texto=o.texto,
                                    conteo=conteos_opciones.get(o.id, 0),
                                )
                                for o in sorted(pregunta.opciones, key=lambda o: (o.orden or 0, o.id))
                            ],
                        )
                    )
                else:
                    cnt = m["n"].get(valor_crudo, 0)
                    preguntas.append(
                        ResultadoPregunta(
                            pregunta_id=pregunta.id, tipo=pregunta.tipo, texto=pregunta.texto, n=cnt
                        )
                    )
            celdas.append(SegmentoCelda(segmento=nombre, n=n, oculto=False, preguntas=preguntas))

        return ResultadosSegmentos(
            encuesta_id=encuesta.id,
            dimension=dimension,
            umbral_minimo_respuestas=encuesta.umbral_minimo_respuestas,
            celdas=celdas,
        )

    async def obtener_textos(self, encuesta_id: int, pregunta_id: int) -> TextosResponse:
        encuesta = await self._get_encuesta_con_resultados(encuesta_id)
        pregunta = next((p for p in encuesta.preguntas if p.id == pregunta_id), None)
        if not pregunta:
            raise NotFoundError("Pregunta", pregunta_id)
        if pregunta.tipo != "texto":
            raise DomainValidationError(f"La pregunta {pregunta_id} no es de tipo texto")

        n_global = await self.repo.count_grupos_respuesta(encuesta_id)
        oculto = encuesta.es_anonima and n_global < encuesta.umbral_minimo_respuestas

        textos: list[str] = []
        if not oculto:
            textos = list(await self.repo.list_textos(pregunta_id))
            random.shuffle(textos)

        return TextosResponse(
            encuesta_id=encuesta_id,
            pregunta_id=pregunta_id,
            n=n_global,
            umbral_minimo_respuestas=encuesta.umbral_minimo_respuestas,
            oculto=oculto,
            textos=textos,
        )

    @staticmethod
    def _slug(titulo: str) -> str:
        normalizado = unicodedata.normalize("NFKD", titulo).encode("ascii", "ignore").decode()
        normalizado = re.sub(r"[^a-zA-Z0-9]+", "_", normalizado).strip("_").lower()
        return normalizado or "encuesta"

    async def exportar_resultados_excel(self, encuesta_id: int) -> tuple[BytesIO, str]:
        encuesta = await self._get_encuesta_con_resultados(encuesta_id)
        resumen = await self.obtener_resultados_globales(encuesta_id)
        segmentos = {
            dimension: await self.obtener_resultados_segmentos(encuesta_id, dimension)
            for dimension in DIMENSIONES_SEGMENTO
        }
        textos_por_pregunta: dict[int, TextosResponse] = {}
        for pregunta in encuesta.preguntas:
            if pregunta.tipo == "texto":
                textos_por_pregunta[pregunta.id] = await self.obtener_textos(
                    encuesta_id, pregunta.id
                )

        output = self._resultados_excel(encuesta, resumen, segmentos, textos_por_pregunta)
        filename = f"resultados_encuesta_{self._slug(encuesta.titulo)}.xlsx"
        return output, filename

    @staticmethod
    def _resultados_excel(
        encuesta: Encuesta,
        resumen: ResultadosGlobal,
        segmentos: dict[str, ResultadosSegmentos],
        textos_por_pregunta: dict[int, TextosResponse],
    ) -> BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()

        # ── Resumen ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resumen"
        ws.cell(row=1, column=1, value=f"Resultados — {encuesta.titulo}").font = Font(
            bold=True, size=14
        )
        filas = [
            ("Tipo", encuesta.tipo),
            ("Anonima", "Si" if resumen.es_anonima else "No"),
            ("Estado", resumen.estado),
            ("Umbral minimo de respuestas", resumen.umbral_minimo_respuestas),
            ("Total de respuestas (n)", resumen.n),
            ("Total de participantes", resumen.total_participantes),
            ("Tasa de respuesta (%)", resumen.tasa_respuesta),
            ("Resultados globales ocultos (min-N)", "Si" if resumen.oculto_global else "No"),
        ]
        for i, (etiqueta, valor) in enumerate(filas, start=3):
            ws.cell(row=i, column=1, value=etiqueta).font = Font(bold=True)
            ws.cell(row=i, column=2, value=valor)

        # ── Preguntas (global) ──────────────────────────────────────────
        ws_preguntas = wb.create_sheet("Preguntas")
        headers = ["Pregunta", "Tipo", "n", "Promedio", "Distribucion 1-5", "Opciones (conteo)"]
        for col, h in enumerate(headers, 1):
            ws_preguntas.cell(row=1, column=col, value=h).font = Font(bold=True)
        if resumen.oculto_global:
            ws_preguntas.cell(
                row=2, column=1,
                value=f"Oculto: n={resumen.n} < umbral={resumen.umbral_minimo_respuestas}",
            )
        else:
            for i, p in enumerate(resumen.preguntas, start=2):
                ws_preguntas.cell(row=i, column=1, value=p.texto)
                ws_preguntas.cell(row=i, column=2, value=p.tipo)
                ws_preguntas.cell(row=i, column=3, value=p.n)
                ws_preguntas.cell(row=i, column=4, value=p.promedio)
                ws_preguntas.cell(
                    row=i, column=5,
                    value=", ".join(f"{d.valor}:{d.conteo}" for d in p.distribucion) or None,
                )
                ws_preguntas.cell(
                    row=i, column=6,
                    value=", ".join(f"{o.texto}:{o.conteo}" for o in p.opciones) or None,
                )

        # ── Segmentos (3 dimensiones, min-N aplicado) ────────────────────
        ws_seg = wb.create_sheet("Segmentos")
        headers = ["Dimension", "Segmento", "n", "Oculto", "Pregunta", "Metrica"]
        for col, h in enumerate(headers, 1):
            ws_seg.cell(row=1, column=col, value=h).font = Font(bold=True)
        r = 2
        for dimension, resultado in segmentos.items():
            for celda in resultado.celdas:
                if celda.oculto:
                    ws_seg.cell(row=r, column=1, value=dimension)
                    ws_seg.cell(row=r, column=2, value=celda.segmento)
                    ws_seg.cell(row=r, column=3, value=celda.n)
                    ws_seg.cell(row=r, column=4, value="Si")
                    ws_seg.cell(row=r, column=5, value=f"Oculto (n < {resumen.umbral_minimo_respuestas})")
                    r += 1
                    continue
                if not celda.preguntas:
                    ws_seg.cell(row=r, column=1, value=dimension)
                    ws_seg.cell(row=r, column=2, value=celda.segmento)
                    ws_seg.cell(row=r, column=3, value=celda.n)
                    ws_seg.cell(row=r, column=4, value="No")
                    r += 1
                    continue
                for p in celda.preguntas:
                    ws_seg.cell(row=r, column=1, value=dimension)
                    ws_seg.cell(row=r, column=2, value=celda.segmento)
                    ws_seg.cell(row=r, column=3, value=celda.n)
                    ws_seg.cell(row=r, column=4, value="No")
                    ws_seg.cell(row=r, column=5, value=p.texto)
                    if p.tipo == "likert":
                        metrica = f"promedio={p.promedio}"
                    elif p.tipo == "opcion_multiple":
                        metrica = ", ".join(f"{o.texto}:{o.conteo}" for o in p.opciones)
                    else:
                        metrica = f"n={p.n}"
                    ws_seg.cell(row=r, column=6, value=metrica)
                    r += 1

        # ── Textos ────────────────────────────────────────────────────────
        ws_textos = wb.create_sheet("Textos")
        headers = ["Pregunta", "Oculto", "Texto"]
        for col, h in enumerate(headers, 1):
            ws_textos.cell(row=1, column=col, value=h).font = Font(bold=True)
        r = 2
        preguntas_texto = [p for p in encuesta.preguntas if p.tipo == "texto"]
        for pregunta in sorted(preguntas_texto, key=lambda p: (p.orden, p.id)):
            respuesta_textos = textos_por_pregunta.get(pregunta.id)
            if respuesta_textos is None:
                continue
            if respuesta_textos.oculto or not respuesta_textos.textos:
                ws_textos.cell(row=r, column=1, value=pregunta.texto)
                ws_textos.cell(row=r, column=2, value="Si" if respuesta_textos.oculto else "No")
                ws_textos.cell(row=r, column=3, value=None)
                r += 1
                continue
            for texto in respuesta_textos.textos:
                ws_textos.cell(row=r, column=1, value=pregunta.texto)
                ws_textos.cell(row=r, column=2, value="No")
                ws_textos.cell(row=r, column=3, value=texto)
                r += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════════════════
    # Helpers de conversion
    # ══════════════════════════════════════════════════════════════════════
    @staticmethod
    def _pregunta_to_response(pregunta: EncuestaPregunta) -> PreguntaResponse:
        return PreguntaResponse(
            id=pregunta.id,
            orden=pregunta.orden,
            tipo=pregunta.tipo,
            texto=pregunta.texto,
            requerida=pregunta.requerida,
            seleccion_multiple=pregunta.seleccion_multiple,
            opciones=[
                OpcionResponse(id=o.id, texto=o.texto, orden=o.orden)
                for o in sorted(pregunta.opciones, key=lambda o: (o.orden or 0, o.id))
            ],
        )

    @classmethod
    def _encuesta_to_response(cls, encuesta: Encuesta) -> EncuestaResponse:
        return EncuestaResponse(
            id=encuesta.id,
            titulo=encuesta.titulo,
            descripcion=encuesta.descripcion,
            tipo=encuesta.tipo,
            es_anonima=encuesta.es_anonima,
            umbral_minimo_respuestas=encuesta.umbral_minimo_respuestas,
            estado=encuesta.estado,
            fecha_publicacion=encuesta.fecha_publicacion,
            fecha_cierre_programada=encuesta.fecha_cierre_programada,
            fecha_cierre_real=encuesta.fecha_cierre_real,
            audiencia_criterios=encuesta.audiencia_criterios,
            recordatorio_cada_dias=encuesta.recordatorio_cada_dias,
            creado_por_id=encuesta.creado_por_id,
            created_at=encuesta.created_at,
            preguntas=[
                cls._pregunta_to_response(p)
                for p in sorted(encuesta.preguntas, key=lambda p: (p.orden, p.id))
            ],
        )
