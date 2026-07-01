# app/services/evaluacion360_service.py
"""
Logica de negocio del modulo Evaluacion 360 (Level Up) — Fase 1.

Responsabilidades:
  - Configuracion global, escalas Likert y banco de preguntas por competencia.
  - CRUD de campanas (competencias, tipos de evaluador, participantes).
  - Sugerencia automatica de evaluadores desde la estructura organizacional
    (`empleados.lider_id`, subordinados, pares del mismo subarea/area).
  - Activacion de campana -> generacion de hojas de evaluacion + notificaciones.
  - Responder evaluaciones (borrador/enviar) con validaciones.
  - Calculo de resultados (promedios por competencia/tipo, auto vs externo,
    brechas vs nivel esperado normalizado a la escala Likert de la campana).
  - Dashboard ejecutivo.

Reutiliza el catalogo de competencias (`levelup_competencias`) y no duplica datos.
El commit lo realiza la dependencia `get_db` al cierre del request; aqui solo se
usa flush() (via el repositorio).
"""

from __future__ import annotations

import logging
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Optional, Sequence

from fastapi import BackgroundTasks
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import (
    ConflictError,
    DomainValidationError,
    ForbiddenError,
    NotFoundError,
)
from app.models.empleados import Empleado
from app.models.evaluacion360 import (
    EVAL360_TIPOS_EVALUADOR,
    Eval360Campana,
    Eval360CampanaCompetencia,
    Eval360CampanaEvaluadorTipo,
    Eval360Comentario,
    Eval360Config,
    Eval360Escala,
    Eval360Evaluacion,
    Eval360Participante,
    Eval360Plantilla,
    Eval360PlantillaCompetencia,
    Eval360PlantillaEvaluadorTipo,
    Eval360Pregunta,
    Eval360Respuesta,
    Eval360Resultado,
)
from app.models.talento import Competencia, PerfilFunciones
from app.repositories.evaluacion360_repository import Evaluacion360Repository
from app.schemas.evaluacion360 import (
    CampanaAvance,
    CampanaCompetenciaResponse,
    CampanaCreate,
    CampanaDetalleResponse,
    CampanaEvaluadorTipoResponse,
    CampanaListResponse,
    CampanaResponse,
    CampanaCompetenciaIn,
    CampanaConfigIn,
    CampanaEvaluadorTipoIn,
    CampanaUpdate,
    ComentarioIn,
    ComentarioReporte,
    CompetenciaEvaluacion,
    ConfigResponse,
    ConfigUpdate,
    DashboardKpis,
    DashboardResponse,
    DashboardSeriePunto,
    EvolucionPunto,
    EscalaCreate,
    EscalaResponse,
    EscalaUpdate,
    EvaluacionDetalleResponse,
    EvaluacionRespuestasIn,
    MiEvaluacionResumen,
    ParticipanteResponse,
    PreguntaCreate,
    PreguntaEvaluacion,
    PreguntaResponse,
    PreguntaUpdate,
    PlantillaCreate,
    PlantillaResponse,
    PlantillaUpdate,
    RecordatoriosResultado,
    ReporteIndividualResponse,
    ResultadoCompetencia,
    ResultadoParticipanteResponse,
    SugerenciaEvaluadorResponse,
)
from app.services.notificacion_service import NotificacionService
from app.utils.audit_logger import audit_background

logger = logging.getLogger(__name__)

AUDIT_MODULE = "EVALUACION_360"
# Tope de pares auto-sugeridos por participante para evitar explosion de hojas.
MAX_PARES_SUGERIDOS = 8
# Tipos que NO se auto-sugieren (requieren alta manual).
TIPOS_MANUALES = ("cliente_interno", "cliente_externo")


def _f(value) -> Optional[float]:
    """Convierte Decimal/None a float para las respuestas."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _fmt(value) -> str:
    """Formatea un número (o None) para celdas de PDF."""
    if value is None:
        return "—"
    return f"{float(value):.2f}"


class Evaluacion360Service:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = Evaluacion360Repository(db)
        self.notificaciones = NotificacionService(db)

    # ══════════════════════════════════════════════════════════════════════════
    # Configuracion + escalas
    # ══════════════════════════════════════════════════════════════════════════
    async def get_or_create_config(self) -> Eval360Config:
        config = await self.repo.get_config()
        if config:
            return config
        # Lazy init: crea escala 1-5 por defecto y config base.
        escalas = await self.repo.list_escalas()
        if escalas:
            escala = escalas[0]
        else:
            escala = Eval360Escala(
                nombre="Estandar 1-5",
                valor_min=1,
                valor_max=5,
                etiquetas={
                    "1": "Nunca",
                    "2": "Rara vez",
                    "3": "A veces",
                    "4": "Frecuentemente",
                    "5": "Siempre",
                },
            )
            self.db.add(escala)
            await self.db.flush()
        config = Eval360Config(
            escala_id=escala.id,
            pesos_evaluadores={
                "jefe": 40,
                "autoevaluacion": 10,
                "par": 20,
                "subordinado": 20,
                "cliente_interno": 5,
                "cliente_externo": 5,
            },
        )
        self.db.add(config)
        await self.db.flush()
        return config

    async def get_config(self) -> ConfigResponse:
        config = await self.get_or_create_config()
        return ConfigResponse.model_validate(config)

    async def update_config(
        self, data: ConfigUpdate, current_user: Empleado, background_tasks: BackgroundTasks
    ) -> ConfigResponse:
        config = await self.get_or_create_config()
        payload = data.model_dump(exclude_unset=True)
        for key, value in payload.items():
            setattr(config, key, value)
        config.updated_by = current_user.empleado_id
        await self.db.flush()
        audit_background(
            background_tasks, self.db, "CONFIG_UPDATE", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=config.id,
            datos_despues=payload,
        )
        return ConfigResponse.model_validate(config)

    async def list_escalas(self) -> list[EscalaResponse]:
        escalas = await self.repo.list_escalas()
        return [EscalaResponse.model_validate(e) for e in escalas]

    async def create_escala(
        self, data: EscalaCreate, current_user: Empleado
    ) -> EscalaResponse:
        escala = Eval360Escala(
            nombre=data.nombre,
            valor_min=data.valor_min,
            valor_max=data.valor_max,
            etiquetas=data.etiquetas,
            created_by=current_user.empleado_id,
        )
        self.db.add(escala)
        await self.db.flush()
        return EscalaResponse.model_validate(escala)

    async def update_escala(self, escala_id: int, data: EscalaUpdate) -> EscalaResponse:
        escala = await self.repo.get_escala(escala_id)
        if not escala:
            raise NotFoundError("Escala no encontrada")
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(escala, key, value)
        await self.db.flush()
        return EscalaResponse.model_validate(escala)

    async def delete_escala(self, escala_id: int) -> None:
        escala = await self.repo.get_escala(escala_id)
        if not escala:
            raise NotFoundError("Escala no encontrada")
        escala.activo = False
        await self.db.flush()

    # ══════════════════════════════════════════════════════════════════════════
    # Banco de preguntas
    # ══════════════════════════════════════════════════════════════════════════
    async def list_preguntas(
        self, competencia_id: Optional[int] = None
    ) -> list[PreguntaResponse]:
        preguntas = await self.repo.list_preguntas(competencia_id=competencia_id)
        return [PreguntaResponse.model_validate(p) for p in preguntas]

    async def create_pregunta(
        self, data: PreguntaCreate, current_user: Empleado
    ) -> PreguntaResponse:
        comp = await self.db.get(Competencia, data.competencia_id)
        if not comp:
            raise NotFoundError("Competencia no encontrada")
        pregunta = Eval360Pregunta(
            competencia_id=data.competencia_id,
            texto=data.texto,
            orden=data.orden,
            created_by=current_user.empleado_id,
        )
        self.db.add(pregunta)
        await self.db.flush()
        return PreguntaResponse.model_validate(pregunta)

    async def update_pregunta(
        self, pregunta_id: int, data: PreguntaUpdate
    ) -> PreguntaResponse:
        pregunta = await self.repo.get_pregunta(pregunta_id)
        if not pregunta:
            raise NotFoundError("Pregunta no encontrada")
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(pregunta, key, value)
        await self.db.flush()
        return PreguntaResponse.model_validate(pregunta)

    async def delete_pregunta(self, pregunta_id: int) -> None:
        pregunta = await self.repo.get_pregunta(pregunta_id)
        if not pregunta:
            raise NotFoundError("Pregunta no encontrada")
        pregunta.activo = False
        await self.db.flush()

    # ══════════════════════════════════════════════════════════════════════════
    # Campanas
    # ══════════════════════════════════════════════════════════════════════════
    async def list_campanas(
        self,
        page: int = 1,
        page_size: int = 10,
        estado: Optional[str] = None,
        search: Optional[str] = None,
    ) -> CampanaListResponse:
        filters = [Eval360Campana.activo.is_(True)]
        if estado:
            filters.append(Eval360Campana.estado == estado)
        if search:
            filters.append(Eval360Campana.nombre.ilike(f"%{search}%"))
        campanas, total = await self.repo.list_campanas(filters, page, page_size)
        items = [await self._campana_to_response(c) for c in campanas]
        return CampanaListResponse(
            items=items, total=total, page=page, page_size=page_size
        )

    async def get_campana(self, campana_id: int) -> CampanaDetalleResponse:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        base = await self._campana_to_response(campana)
        comp_nombres = await self._competencia_nombres(
            [c.competencia_id for c in campana.competencias]
        )
        competencias = [
            CampanaCompetenciaResponse(
                competencia_id=c.competencia_id,
                competencia_nombre=comp_nombres.get(c.competencia_id),
                peso=_f(c.peso) or 0.0,
                num_preguntas=c.num_preguntas,
                nivel_esperado=c.nivel_esperado,
                obligatoria=c.obligatoria,
                orden=c.orden,
            )
            for c in sorted(campana.competencias, key=lambda x: (x.orden or 0, x.id))
        ]
        evaluador_tipos = [
            CampanaEvaluadorTipoResponse(
                tipo=t.tipo, peso=_f(t.peso) or 0.0, activo=t.activo
            )
            for t in campana.evaluador_tipos
        ]
        return CampanaDetalleResponse(
            **base.model_dump(),
            competencias=competencias,
            evaluador_tipos=evaluador_tipos,
        )

    async def create_campana(
        self, data: CampanaCreate, current_user: Empleado, background_tasks: BackgroundTasks
    ) -> CampanaDetalleResponse:
        self._validar_fechas(data.fecha_inicio, data.fecha_cierre)

        # Aplicar plantilla: si se indica plantilla_id y no se enviaron
        # competencias/evaluadores/escala/config, se copian de la plantilla.
        competencias = data.competencias
        evaluador_tipos = data.evaluador_tipos
        escala_id = data.escala_id
        config = data.config
        if data.plantilla_id:
            plantilla = await self.repo.get_plantilla(data.plantilla_id)
            if not plantilla:
                raise NotFoundError("Plantilla no encontrada")
            if not competencias:
                competencias = [
                    CampanaCompetenciaIn(
                        competencia_id=c.competencia_id, peso=_f(c.peso) or 0.0,
                        num_preguntas=c.num_preguntas, nivel_esperado=c.nivel_esperado,
                        obligatoria=c.obligatoria, orden=c.orden,
                    )
                    for c in plantilla.competencias
                ]
            if not evaluador_tipos:
                evaluador_tipos = [
                    CampanaEvaluadorTipoIn(tipo=t.tipo, peso=_f(t.peso) or 0.0, activo=t.activo)
                    for t in plantilla.evaluador_tipos
                ]
            if escala_id is None:
                escala_id = plantilla.escala_id
            if config is None and plantilla.config:
                config = CampanaConfigIn(**plantilla.config)

        campana = Eval360Campana(
            nombre=data.nombre,
            descripcion=data.descripcion,
            objetivo=data.objetivo,
            fecha_inicio=data.fecha_inicio,
            fecha_cierre=data.fecha_cierre,
            tipo=data.tipo or "evaluacion_360",
            escala_id=escala_id,
            plantilla_id=data.plantilla_id,
            estado="borrador",
            config=config.model_dump(mode="json") if config else None,
            created_by=current_user.empleado_id,
        )
        self.db.add(campana)
        await self.db.flush()

        await self._sync_competencias(campana, competencias)
        await self._sync_evaluador_tipos(campana, evaluador_tipos)
        await self._sync_participantes(campana, data.empleado_ids)
        await self.db.flush()

        audit_background(
            background_tasks, self.db, "CAMPANA_CREATE", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=campana.id,
            datos_despues={"nombre": campana.nombre},
        )
        return await self.get_campana(campana.id)

    async def update_campana(
        self, campana_id: int, data: CampanaUpdate, current_user: Empleado
    ) -> CampanaDetalleResponse:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        if campana.estado not in ("borrador",):
            raise ConflictError(
                "Solo se pueden editar campanas en estado borrador"
            )
        payload = data.model_dump(exclude_unset=True)
        for field in ("nombre", "descripcion", "objetivo", "fecha_inicio",
                      "fecha_cierre", "escala_id"):
            if field in payload:
                setattr(campana, field, payload[field])
        if "config" in payload and data.config is not None:
            campana.config = data.config.model_dump(mode="json")
        self._validar_fechas(campana.fecha_inicio, campana.fecha_cierre)
        if data.competencias is not None:
            await self._sync_competencias(campana, data.competencias, replace=True)
        if data.evaluador_tipos is not None:
            await self._sync_evaluador_tipos(campana, data.evaluador_tipos, replace=True)
        if data.empleado_ids is not None:
            await self._sync_participantes(campana, data.empleado_ids, replace=True)
        campana.updated_by = current_user.empleado_id
        await self.db.flush()
        return await self.get_campana(campana.id)

    async def delete_campana(self, campana_id: int, current_user: Empleado) -> None:
        campana = await self.repo.get_campana(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        campana.activo = False
        campana.updated_by = current_user.empleado_id
        await self.db.flush()

    async def duplicar_campana(
        self, campana_id: int, current_user: Empleado
    ) -> CampanaDetalleResponse:
        origen = await self.repo.get_campana_detalle(campana_id)
        if not origen or not origen.activo:
            raise NotFoundError("Campana no encontrada")
        nueva = Eval360Campana(
            nombre=f"{origen.nombre} (copia)",
            descripcion=origen.descripcion,
            objetivo=origen.objetivo,
            tipo=origen.tipo,
            escala_id=origen.escala_id,
            plantilla_id=origen.plantilla_id,
            estado="borrador",
            config=origen.config,
            created_by=current_user.empleado_id,
        )
        self.db.add(nueva)
        await self.db.flush()
        for c in origen.competencias:
            self.db.add(Eval360CampanaCompetencia(
                campana_id=nueva.id, competencia_id=c.competencia_id, peso=c.peso,
                num_preguntas=c.num_preguntas, nivel_esperado=c.nivel_esperado,
                obligatoria=c.obligatoria, orden=c.orden,
            ))
        for t in origen.evaluador_tipos:
            self.db.add(Eval360CampanaEvaluadorTipo(
                campana_id=nueva.id, tipo=t.tipo, peso=t.peso, activo=t.activo,
            ))
        await self.db.flush()
        return await self.get_campana(nueva.id)

    async def activar_campana(
        self, campana_id: int, current_user: Empleado, background_tasks: BackgroundTasks
    ) -> CampanaDetalleResponse:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        if campana.estado not in ("borrador",):
            raise ConflictError("La campana ya fue activada")
        if not campana.participantes:
            raise DomainValidationError("La campana no tiene participantes")

        activos = [t for t in campana.evaluador_tipos if t.activo]
        if not activos:
            raise DomainValidationError("Debe activar al menos un tipo de evaluador")
        total_peso = round(sum(_f(t.peso) or 0.0 for t in activos), 2)
        if abs(total_peso - 100.0) > 0.01:
            raise DomainValidationError(
                f"Los pesos de evaluadores activos deben sumar 100% (suman {total_peso})"
            )

        generadas = await self._generar_evaluaciones(campana)
        campana.estado = "activa"
        campana.updated_by = current_user.empleado_id
        await self.db.flush()

        await self._notificar_evaluadores(campana, generadas)
        audit_background(
            background_tasks, self.db, "CAMPANA_ACTIVAR", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=campana.id,
            datos_despues={"evaluaciones_generadas": len(generadas)},
        )
        return await self.get_campana(campana.id)

    async def cerrar_campana(
        self, campana_id: int, current_user: Empleado, background_tasks: BackgroundTasks
    ) -> CampanaDetalleResponse:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        if campana.estado in ("cerrada", "cancelada"):
            raise ConflictError("La campana ya esta cerrada o cancelada")
        await self._calcular_resultados_campana(campana)
        campana.estado = "cerrada"
        campana.updated_by = current_user.empleado_id
        await self.db.flush()

        # Notificar a los evaluados que su resultado está disponible.
        config = await self.get_or_create_config()
        for p in await self.repo.list_participantes(campana_id):
            asunto, cuerpo = self._texto_correo(
                config, "finalizada",
                asunto_def="Tu Evaluación 360° ha finalizado",
                cuerpo_def=(
                    f"La campaña '{campana.nombre}' ha finalizado. "
                    f"Tus resultados ya están disponibles."
                ),
                campana=campana.nombre,
            )
            await self._enviar_a_empleado(
                p.empleado_id, asunto, cuerpo,
                metadata={"campana_id": campana.id, "evento": "finalizada"},
            )

        audit_background(
            background_tasks, self.db, "CAMPANA_CERRAR", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=campana.id,
        )
        return await self.get_campana(campana.id)

    async def cancelar_campana(
        self, campana_id: int, current_user: Empleado, background_tasks: BackgroundTasks
    ) -> CampanaDetalleResponse:
        campana = await self.repo.get_campana(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        if campana.estado in ("cerrada", "cancelada"):
            raise ConflictError("La campana ya esta cerrada o cancelada")
        campana.estado = "cancelada"
        campana.updated_by = current_user.empleado_id
        await self.db.flush()
        audit_background(
            background_tasks, self.db, "CAMPANA_CANCELAR", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=campana.id,
        )
        return await self.get_campana(campana.id)

    # ── Sub-recursos de campana ───────────────────────────────────────────────
    async def list_participantes(self, campana_id: int) -> list[ParticipanteResponse]:
        campana = await self.repo.get_campana(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        participantes = await self.repo.list_participantes(campana_id)
        out: list[ParticipanteResponse] = []
        for p in participantes:
            total = len(p.evaluaciones)
            completadas = sum(1 for e in p.evaluaciones if e.estado == "completada")
            out.append(ParticipanteResponse(
                id=p.id,
                empleado_id=p.empleado_id,
                empleado_nombre=p.empleado.nombre if p.empleado else None,
                puesto=self._puesto_nombre(p.empleado),
                area=self._area_nombre(p.empleado),
                estado=p.estado,
                evaluaciones_total=total,
                evaluaciones_completadas=completadas,
                avance=round(completadas / total * 100, 1) if total else 0.0,
            ))
        return out

    async def sugerir_evaluadores(
        self, campana_id: int
    ) -> list[SugerenciaEvaluadorResponse]:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        participantes = await self.repo.list_participantes(campana_id)
        tipos_activos = {t.tipo for t in campana.evaluador_tipos if t.activo}
        out: list[SugerenciaEvaluadorResponse] = []
        for p in participantes:
            if not p.empleado:
                continue
            for tipo, evaluador in await self._resolver_evaluadores(p.empleado, tipos_activos):
                out.append(SugerenciaEvaluadorResponse(
                    participante_id=p.id,
                    empleado_id=p.empleado_id,
                    empleado_nombre=p.empleado.nombre,
                    tipo_evaluador=tipo,
                    evaluador_empleado_id=evaluador.empleado_id if evaluador else None,
                    evaluador_nombre=evaluador.nombre if evaluador else None,
                ))
        return out

    # ══════════════════════════════════════════════════════════════════════════
    # Mis Evaluaciones (responder)
    # ══════════════════════════════════════════════════════════════════════════
    async def list_mis_evaluaciones(
        self, current_user: Empleado, estado: Optional[str] = None
    ) -> list[MiEvaluacionResumen]:
        evals = await self.repo.list_mis_evaluaciones(
            current_user.empleado_id, estado=estado
        )
        # Prefetch campana names
        campana_ids = {e.campana_id for e in evals}
        campanas = {
            c.id: c for c in await self._get_campanas_by_ids(campana_ids)
        }
        out: list[MiEvaluacionResumen] = []
        for e in evals:
            campana = campanas.get(e.campana_id)
            evaluado = (
                e.participante.empleado.nombre
                if e.participante and e.participante.empleado
                else None
            )
            n_preg = await self._num_preguntas_evaluacion(e)
            avance = round(len(e.respuestas) / n_preg * 100, 1) if n_preg else 0.0
            out.append(MiEvaluacionResumen(
                id=e.id,
                campana_id=e.campana_id,
                campana_nombre=campana.nombre if campana else None,
                evaluado_nombre=None if e.es_anonima else evaluado,
                tipo_evaluador=e.tipo_evaluador,
                estado=e.estado,
                fecha_asignacion=e.fecha_asignacion,
                fecha_limite=e.fecha_limite,
                avance=avance,
            ))
        return out

    async def get_evaluacion_detalle(
        self, evaluacion_id: int, current_user: Empleado
    ) -> EvaluacionDetalleResponse:
        evaluacion = await self.repo.get_evaluacion(evaluacion_id)
        if not evaluacion:
            raise NotFoundError("Evaluacion no encontrada")
        if evaluacion.evaluador_empleado_id != current_user.empleado_id:
            raise ForbiddenError("No puedes acceder a esta evaluacion")

        campana = await self.repo.get_campana_detalle(evaluacion.campana_id)
        escala = await self._escala_de_campana(campana)
        respuestas_por_pregunta = {r.pregunta_id: _f(r.valor) for r in evaluacion.respuestas}
        comentario_por_comp = {
            c.competencia_id: c.texto for c in evaluacion.comentarios
        }

        comp_ids = [c.competencia_id for c in campana.competencias]
        comp_nombres = await self._competencia_nombres(comp_ids)
        competencias: list[CompetenciaEvaluacion] = []
        for cc in sorted(campana.competencias, key=lambda x: (x.orden or 0, x.id)):
            preguntas = await self.repo.list_preguntas(
                competencia_id=cc.competencia_id, solo_activas=True
            )
            if cc.num_preguntas:
                preguntas = list(preguntas)[: cc.num_preguntas]
            competencias.append(CompetenciaEvaluacion(
                competencia_id=cc.competencia_id,
                competencia_nombre=comp_nombres.get(cc.competencia_id, "Competencia"),
                nivel_esperado=cc.nivel_esperado,
                preguntas=[
                    PreguntaEvaluacion(
                        pregunta_id=p.id,
                        texto=p.texto,
                        valor=respuestas_por_pregunta.get(p.id),
                    )
                    for p in preguntas
                ],
                comentario=comentario_por_comp.get(cc.competencia_id),
            ))

        cfg = (campana.config or {})
        return EvaluacionDetalleResponse(
            id=evaluacion.id,
            campana_id=evaluacion.campana_id,
            campana_nombre=campana.nombre,
            evaluado_nombre=(
                None if evaluacion.es_anonima
                else (evaluacion.participante.empleado.nombre
                      if evaluacion.participante and evaluacion.participante.empleado
                      else None)
            ),
            tipo_evaluador=evaluacion.tipo_evaluador,
            estado=evaluacion.estado,
            es_anonima=evaluacion.es_anonima,
            escala=EscalaResponse.model_validate(escala) if escala else None,
            comentarios_obligatorios=bool(cfg.get("comentarios_obligatorios", False)),
            fecha_limite=evaluacion.fecha_limite,
            competencias=competencias,
        )

    async def guardar_borrador(
        self, evaluacion_id: int, data: EvaluacionRespuestasIn, current_user: Empleado
    ) -> EvaluacionDetalleResponse:
        evaluacion = await self._get_evaluacion_editable(evaluacion_id, current_user)
        await self._guardar_respuestas(evaluacion, data)
        if evaluacion.estado == "pendiente":
            evaluacion.estado = "en_progreso"
        await self.db.flush()
        return await self.get_evaluacion_detalle(evaluacion_id, current_user)

    async def enviar_evaluacion(
        self,
        evaluacion_id: int,
        data: EvaluacionRespuestasIn,
        current_user: Empleado,
        background_tasks: BackgroundTasks,
    ) -> EvaluacionDetalleResponse:
        evaluacion = await self._get_evaluacion_editable(evaluacion_id, current_user)
        campana = await self.repo.get_campana_detalle(evaluacion.campana_id)
        await self._guardar_respuestas(evaluacion, data)

        # Validar completitud: todas las preguntas de las competencias respondidas.
        total_preguntas = await self._num_preguntas_evaluacion(evaluacion, campana)
        respuestas = await self.repo.get_respuestas_evaluacion(evaluacion.id)
        if len(respuestas) < total_preguntas:
            raise DomainValidationError(
                "Debes responder todas las preguntas antes de enviar"
            )
        cfg = campana.config or {}
        if cfg.get("comentarios_obligatorios"):
            if not data.comentarios and not evaluacion.comentarios:
                raise DomainValidationError(
                    "Esta evaluacion requiere al menos un comentario"
                )

        evaluacion.estado = "completada"
        evaluacion.fecha_completada = datetime.now(timezone.utc)
        await self._actualizar_estado_participante(evaluacion.participante_id)
        await self.db.flush()
        audit_background(
            background_tasks, self.db, "EVALUACION_ENVIAR", AUDIT_MODULE,
            usuario_id=current_user.empleado_id, entidad_id=evaluacion.id,
        )
        return await self.get_evaluacion_detalle(evaluacion_id, current_user)

    # ══════════════════════════════════════════════════════════════════════════
    # Resultados / reportes
    # ══════════════════════════════════════════════════════════════════════════
    async def get_resultados_campana(
        self, campana_id: int
    ) -> list[ResultadoParticipanteResponse]:
        campana = await self.repo.get_campana_detalle(campana_id)
        if not campana or not campana.activo:
            raise NotFoundError("Campana no encontrada")
        await self._calcular_resultados_campana(campana)
        await self.db.flush()
        participantes = await self.repo.list_participantes(campana_id)
        return [await self._resultado_participante(p.id) for p in participantes]

    async def get_resultado_participante(
        self, participante_id: int
    ) -> ResultadoParticipanteResponse:
        participante = await self.repo.get_participante(participante_id)
        if not participante:
            raise NotFoundError("Participante no encontrado")
        campana = await self.repo.get_campana_detalle(participante.campana_id)
        await self._calcular_resultados_participante(participante, campana)
        await self.db.flush()
        return await self._resultado_participante(participante_id)

    async def get_reporte_individual(
        self, participante_id: int
    ) -> ReporteIndividualResponse:
        """Reporte individual completo: promedios, brechas, comentarios y evolucion."""
        participante = await self.repo.get_participante(participante_id)
        if not participante:
            raise NotFoundError("Participante no encontrado")
        campana = await self.repo.get_campana_detalle(participante.campana_id)
        await self._calcular_resultados_participante(participante, campana)
        await self.db.flush()

        base = await self._resultado_participante(participante_id)

        # Promedios auto vs externo (media entre competencias).
        autos = [c.autoevaluacion for c in base.competencias if c.autoevaluacion is not None]
        externos: list[float] = []
        for c in base.competencias:
            if c.promedio_por_tipo:
                vals = [v for t, v in c.promedio_por_tipo.items() if t != "autoevaluacion"]
                if vals:
                    externos.append(sum(vals) / len(vals))
        promedio_auto = round(sum(autos) / len(autos), 2) if autos else None
        promedio_externo = round(sum(externos) / len(externos), 2) if externos else None

        # Comentarios agrupados por competencia/tipo.
        comp_nombres = await self._competencia_nombres(
            [c.competencia_id for c in campana.competencias]
        )
        comentarios: list[ComentarioReporte] = []
        for com, tipo in await self.repo.list_comentarios_participante(participante_id):
            comentarios.append(ComentarioReporte(
                tipo_evaluador=tipo,
                competencia_id=com.competencia_id,
                competencia_nombre=comp_nombres.get(com.competencia_id) if com.competencia_id else None,
                texto=com.texto,
                tipo=com.tipo,
            ))

        # Evolucion historica (calificacion global por campana del empleado).
        evolucion: list[EvolucionPunto] = []
        for part, camp, res in await self.repo.list_resultados_globales_empleado(
            participante.empleado_id
        ):
            evolucion.append(EvolucionPunto(
                campana_id=camp.id,
                campana_nombre=camp.nombre,
                fecha=camp.fecha_cierre,
                calificacion_general=_f(res.calificacion_general),
            ))

        return ReporteIndividualResponse(
            participante_id=participante_id,
            empleado_id=participante.empleado_id,
            empleado_nombre=participante.empleado.nombre if participante.empleado else None,
            puesto=self._puesto_nombre(participante.empleado),
            area=self._area_nombre(participante.empleado),
            campana_id=campana.id,
            campana_nombre=campana.nombre,
            calificacion_general=base.calificacion_general,
            promedio_autoevaluacion=promedio_auto,
            promedio_externo=promedio_externo,
            competencias=base.competencias,
            fortalezas=base.fortalezas,
            oportunidades=base.oportunidades,
            comentarios=comentarios,
            evolucion=evolucion,
        )

    # ══════════════════════════════════════════════════════════════════════════
    # Exportacion (PDF / Excel)
    # ══════════════════════════════════════════════════════════════════════════
    async def export_reporte_individual(self, participante_id: int, formato: str) -> BytesIO:
        rep = await self.get_reporte_individual(participante_id)
        if formato == "excel":
            return self._reporte_individual_excel(rep)
        return self._reporte_individual_pdf(rep)

    async def export_resultados_campana(self, campana_id: int, formato: str) -> BytesIO:
        resultados = await self.get_resultados_campana(campana_id)
        campana = await self.repo.get_campana(campana_id)
        nombre = campana.nombre if campana else "Campaña"
        if formato == "excel":
            return self._resultados_campana_excel(nombre, resultados)
        return self._resultados_campana_pdf(nombre, resultados)

    def _reporte_individual_excel(self, rep: ReporteIndividualResponse) -> BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte 360"
        ws.cell(row=1, column=1, value="Reporte Evaluación 360°").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value="Colaborador")
        ws.cell(row=2, column=2, value=rep.empleado_nombre or "—")
        ws.cell(row=3, column=1, value="Puesto")
        ws.cell(row=3, column=2, value=rep.puesto or "—")
        ws.cell(row=4, column=1, value="Campaña")
        ws.cell(row=4, column=2, value=rep.campana_nombre or "—")
        ws.cell(row=5, column=1, value="Calificación general")
        ws.cell(row=5, column=2, value=rep.calificacion_general)
        ws.cell(row=6, column=1, value="Autoevaluación / Externo")
        ws.cell(row=6, column=2, value=f"{rep.promedio_autoevaluacion} / {rep.promedio_externo}")

        headers = ["Competencia", "Promedio", "Autoevaluación", "Nivel esperado", "Brecha", "Estado"]
        r = 8
        for col, h in enumerate(headers, 1):
            ws.cell(row=r, column=col, value=h).font = Font(bold=True)
        for c in rep.competencias:
            r += 1
            ws.cell(row=r, column=1, value=c.competencia_nombre or "—")
            ws.cell(row=r, column=2, value=c.promedio_general)
            ws.cell(row=r, column=3, value=c.autoevaluacion)
            ws.cell(row=r, column=4, value=c.nivel_esperado)
            ws.cell(row=r, column=5, value=c.brecha)
            ws.cell(row=r, column=6, value=c.estado_brecha)

        r += 2
        ws.cell(row=r, column=1, value="Comentarios").font = Font(bold=True)
        for com in rep.comentarios:
            r += 1
            ws.cell(row=r, column=1, value=com.competencia_nombre or "General")
            ws.cell(row=r, column=2, value=com.tipo_evaluador or "")
            ws.cell(row=r, column=3, value=com.texto)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _reporte_individual_pdf(self, rep: ReporteIndividualResponse) -> BytesIO:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        el = []
        el.append(Paragraph("Reporte Evaluación 360°", styles["Title"]))
        el.append(Paragraph(f"Colaborador: {rep.empleado_nombre or '—'}", styles["Normal"]))
        el.append(Paragraph(f"Puesto: {rep.puesto or '—'}", styles["Normal"]))
        el.append(Paragraph(f"Campaña: {rep.campana_nombre or '—'}", styles["Normal"]))
        el.append(Paragraph(
            f"Calificación general: {rep.calificacion_general}  ·  "
            f"Auto: {rep.promedio_autoevaluacion}  ·  Externo: {rep.promedio_externo}",
            styles["Normal"],
        ))
        el.append(Spacer(1, 12))

        data = [["Competencia", "Prom.", "Auto", "Esperado", "Brecha", "Estado"]]
        for c in rep.competencias:
            data.append([
                (c.competencia_nombre or "—")[:32],
                _fmt(c.promedio_general), _fmt(c.autoevaluacion),
                _fmt(c.nivel_esperado), _fmt(c.brecha), c.estado_brecha or "—",
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A1628")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        el.append(table)

        if rep.comentarios:
            el.append(Spacer(1, 12))
            el.append(Paragraph("Comentarios", styles["Heading3"]))
            for com in rep.comentarios:
                etiqueta = com.competencia_nombre or "General"
                el.append(Paragraph(
                    f"<b>{etiqueta}</b> ({com.tipo_evaluador or 's/d'}): {com.texto}",
                    styles["Normal"],
                ))
        doc.build(el)
        output.seek(0)
        return output

    def _resultados_campana_excel(
        self, nombre: str, resultados: list[ResultadoParticipanteResponse]
    ) -> BytesIO:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.cell(row=1, column=1, value=f"Resultados — {nombre}").font = Font(bold=True, size=14)
        headers = ["Colaborador", "Puesto", "Calificación general", "Fortalezas", "Oportunidades"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = Font(bold=True)
        for i, r in enumerate(resultados, 4):
            ws.cell(row=i, column=1, value=r.empleado_nombre or "—")
            ws.cell(row=i, column=2, value=r.puesto or "—")
            ws.cell(row=i, column=3, value=r.calificacion_general)
            ws.cell(row=i, column=4, value=", ".join(r.fortalezas))
            ws.cell(row=i, column=5, value=", ".join(r.oportunidades))
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _resultados_campana_pdf(
        self, nombre: str, resultados: list[ResultadoParticipanteResponse]
    ) -> BytesIO:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        el = [Paragraph(f"Resultados 360° — {nombre}", styles["Title"]),
              Paragraph(f"Fecha: {date.today().isoformat()}", styles["Normal"]),
              Spacer(1, 12)]
        data = [["Colaborador", "Puesto", "Calif. general", "Fortalezas", "Oportunidades"]]
        for r in resultados:
            data.append([
                (r.empleado_nombre or "—")[:28],
                (r.puesto or "—")[:24],
                _fmt(r.calificacion_general),
                ", ".join(r.fortalezas)[:40],
                ", ".join(r.oportunidades)[:40],
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A1628")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        el.append(table)
        doc.build(el)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════════════════════
    # Plantillas
    # ══════════════════════════════════════════════════════════════════════════
    async def list_plantillas(self) -> list[PlantillaResponse]:
        plantillas = await self.repo.list_plantillas(solo_activas=True)
        return [await self._plantilla_to_response(p) for p in plantillas]

    async def get_plantilla(self, plantilla_id: int) -> PlantillaResponse:
        plantilla = await self.repo.get_plantilla(plantilla_id)
        if not plantilla or not plantilla.activo:
            raise NotFoundError("Plantilla no encontrada")
        return await self._plantilla_to_response(plantilla)

    async def create_plantilla(
        self, data: PlantillaCreate, current_user: Empleado
    ) -> PlantillaResponse:
        plantilla = Eval360Plantilla(
            nombre=data.nombre,
            descripcion=data.descripcion,
            escala_id=data.escala_id,
            config=data.config.model_dump(mode="json") if data.config else None,
            created_by=current_user.empleado_id,
        )
        self.db.add(plantilla)
        await self.db.flush()
        self._sync_plantilla_hijos(plantilla, data.competencias, data.evaluador_tipos)
        await self.db.flush()
        return await self.get_plantilla(plantilla.id)

    async def update_plantilla(
        self, plantilla_id: int, data: PlantillaUpdate, current_user: Empleado
    ) -> PlantillaResponse:
        plantilla = await self.repo.get_plantilla(plantilla_id)
        if not plantilla or not plantilla.activo:
            raise NotFoundError("Plantilla no encontrada")
        payload = data.model_dump(exclude_unset=True)
        for field in ("nombre", "descripcion", "escala_id", "activo"):
            if field in payload:
                setattr(plantilla, field, payload[field])
        if "config" in payload and data.config is not None:
            plantilla.config = data.config.model_dump(mode="json")
        if data.competencias is not None or data.evaluador_tipos is not None:
            for c in list(plantilla.competencias):
                await self.db.delete(c)
            for t in list(plantilla.evaluador_tipos):
                await self.db.delete(t)
            await self.db.flush()
            self._sync_plantilla_hijos(
                plantilla,
                data.competencias or [],
                data.evaluador_tipos or [],
            )
        plantilla.updated_by = current_user.empleado_id
        await self.db.flush()
        return await self.get_plantilla(plantilla.id)

    async def delete_plantilla(self, plantilla_id: int) -> None:
        plantilla = await self.repo.get_plantilla(plantilla_id)
        if not plantilla or not plantilla.activo:
            raise NotFoundError("Plantilla no encontrada")
        plantilla.activo = False
        await self.db.flush()

    def _sync_plantilla_hijos(
        self, plantilla: Eval360Plantilla, competencias, evaluador_tipos
    ) -> None:
        for c in competencias:
            self.db.add(Eval360PlantillaCompetencia(
                plantilla_id=plantilla.id, competencia_id=c.competencia_id,
                peso=c.peso, num_preguntas=c.num_preguntas,
                nivel_esperado=c.nivel_esperado, obligatoria=c.obligatoria, orden=c.orden,
            ))
        for t in evaluador_tipos:
            self.db.add(Eval360PlantillaEvaluadorTipo(
                plantilla_id=plantilla.id, tipo=t.tipo, peso=t.peso, activo=t.activo,
            ))

    async def _plantilla_to_response(self, plantilla: Eval360Plantilla) -> PlantillaResponse:
        comp_nombres = await self._competencia_nombres(
            [c.competencia_id for c in plantilla.competencias]
        )
        return PlantillaResponse(
            id=plantilla.id,
            nombre=plantilla.nombre,
            descripcion=plantilla.descripcion,
            escala_id=plantilla.escala_id,
            activo=plantilla.activo,
            config=plantilla.config,
            competencias=[
                CampanaCompetenciaResponse(
                    competencia_id=c.competencia_id,
                    competencia_nombre=comp_nombres.get(c.competencia_id),
                    peso=_f(c.peso) or 0.0, num_preguntas=c.num_preguntas,
                    nivel_esperado=c.nivel_esperado, obligatoria=c.obligatoria, orden=c.orden,
                )
                for c in sorted(plantilla.competencias, key=lambda x: (x.orden or 0, x.id))
            ],
            evaluador_tipos=[
                CampanaEvaluadorTipoResponse(tipo=t.tipo, peso=_f(t.peso) or 0.0, activo=t.activo)
                for t in plantilla.evaluador_tipos
            ],
        )

    # ══════════════════════════════════════════════════════════════════════════
    # Recordatorios automáticos (invocado por APScheduler)
    # ══════════════════════════════════════════════════════════════════════════
    async def procesar_recordatorios(self) -> RecordatoriosResultado:
        """Marca evaluaciones vencidas y envía recordatorios según config.dias_antes."""
        config = await self.get_or_create_config()
        dias_antes = [3, 1, 0]
        if isinstance(config.recordatorios, dict):
            valor = config.recordatorios.get("dias_antes")
            if isinstance(valor, list) and valor:
                dias_antes = [int(v) for v in valor]

        hoy = date.today()
        enviados = 0
        vencidas = 0
        pendientes = await self.repo.list_evaluaciones_pendientes_con_limite()
        for ev, campana in pendientes:
            dias = (ev.fecha_limite - hoy).days
            if dias < 0:
                ev.estado = "vencida"
                vencidas += 1
                asunto, cuerpo = self._texto_correo(
                    config, "vencida",
                    asunto_def="Evaluación 360° vencida",
                    cuerpo_def=(
                        f"La evaluación de la campaña '{campana.nombre}' venció el "
                        f"{ev.fecha_limite.isoformat()}."
                    ),
                    campana=campana.nombre,
                )
                if ev.evaluador_empleado_id:
                    await self._enviar_a_empleado(
                        ev.evaluador_empleado_id, asunto, cuerpo,
                        target_url="#/mis-evaluaciones",
                        metadata={"campana_id": campana.id, "evento": "vencida"},
                    )
            elif dias in dias_antes:
                enviados += 1
                asunto, cuerpo = self._texto_correo(
                    config, "recordatorio",
                    asunto_def="Recordatorio: Evaluación 360° pendiente",
                    cuerpo_def=(
                        f"Tienes una evaluación pendiente en '{campana.nombre}'. "
                        f"Fecha límite: {ev.fecha_limite.isoformat()}."
                    ),
                    campana=campana.nombre,
                )
                if ev.evaluador_empleado_id:
                    await self._enviar_a_empleado(
                        ev.evaluador_empleado_id, asunto, cuerpo,
                        target_url="#/mis-evaluaciones",
                        metadata={"campana_id": campana.id, "evento": "recordatorio"},
                    )
        await self.db.flush()
        return RecordatoriosResultado(
            recordatorios_enviados=enviados, vencidas_marcadas=vencidas
        )

    # ══════════════════════════════════════════════════════════════════════════
    # Dashboard
    # ══════════════════════════════════════════════════════════════════════════
    async def get_dashboard(self) -> DashboardResponse:
        activas = await self.repo.count_campanas_por_estado(["activa", "en_progreso"])
        finalizadas = await self.repo.count_campanas_por_estado(["finalizada", "cerrada"])
        pendientes = await self.repo.count_evaluaciones_por_estado(["pendiente", "en_progreso"])
        respondidas = await self.repo.count_evaluaciones_por_estado(["completada"])
        participantes = await self.repo.count_participantes()

        # Promedios por competencia sobre resultados calculados.
        resultados = (
            await self.db.execute(
                select(Eval360Resultado).where(Eval360Resultado.competencia_id.isnot(None))
            )
        ).scalars().all()
        por_competencia: dict[int, list[float]] = {}
        for r in resultados:
            if r.promedio_general is not None:
                por_competencia.setdefault(r.competencia_id, []).append(_f(r.promedio_general))
        comp_nombres = await self._competencia_nombres(list(por_competencia.keys()))
        promedios = {
            cid: round(sum(vals) / len(vals), 2)
            for cid, vals in por_competencia.items() if vals
        }
        ordenados = sorted(promedios.items(), key=lambda kv: kv[1])
        promedio_general = (
            round(sum(promedios.values()) / len(promedios), 2) if promedios else None
        )

        kpis = DashboardKpis(
            campanas_activas=activas,
            campanas_finalizadas=finalizadas,
            evaluaciones_pendientes=pendientes,
            evaluaciones_respondidas=respondidas,
            participantes=participantes,
            promedio_general=promedio_general,
            competencia_menor=comp_nombres.get(ordenados[0][0]) if ordenados else None,
            competencia_menor_promedio=ordenados[0][1] if ordenados else None,
            competencia_mayor=comp_nombres.get(ordenados[-1][0]) if ordenados else None,
            competencia_mayor_promedio=ordenados[-1][1] if ordenados else None,
        )
        mejor = [
            DashboardSeriePunto(label=comp_nombres.get(cid, "?"), valor=val)
            for cid, val in sorted(promedios.items(), key=lambda kv: -kv[1])[:5]
        ]
        oportunidad = [
            DashboardSeriePunto(label=comp_nombres.get(cid, "?"), valor=val)
            for cid, val in ordenados[:5]
        ]
        estado_series = [
            DashboardSeriePunto(label="Pendientes", valor=float(pendientes)),
            DashboardSeriePunto(label="Respondidas", valor=float(respondidas)),
        ]
        avance_campana = await self._avance_por_campana()
        distribucion = await self._distribucion_calificaciones()
        return DashboardResponse(
            kpis=kpis,
            estado_evaluaciones=estado_series,
            competencias_mejor=mejor,
            competencias_oportunidad=oportunidad,
            avance_por_campana=avance_campana,
            distribucion_calificaciones=distribucion,
        )

    # ══════════════════════════════════════════════════════════════════════════
    # Helpers privados
    # ══════════════════════════════════════════════════════════════════════════
    def _validar_fechas(self, inicio, cierre) -> None:
        if inicio and cierre and cierre < inicio:
            raise DomainValidationError(
                "La fecha de cierre no puede ser anterior a la de inicio"
            )

    async def _campana_to_response(self, campana: Eval360Campana) -> CampanaResponse:
        evals = await self.repo.list_evaluaciones_campana(campana.id)
        total = len(evals)
        completadas = sum(1 for e in evals if e.estado == "completada")
        evaluadores = len({
            e.evaluador_empleado_id for e in evals if e.evaluador_empleado_id
        })
        participantes = (
            await self.db.execute(
                select(Eval360Participante.id).where(
                    Eval360Participante.campana_id == campana.id
                )
            )
        ).scalars().all()
        return CampanaResponse(
            id=campana.id,
            nombre=campana.nombre,
            descripcion=campana.descripcion,
            objetivo=campana.objetivo,
            fecha_inicio=campana.fecha_inicio,
            fecha_cierre=campana.fecha_cierre,
            estado=campana.estado,
            tipo=campana.tipo,
            escala_id=campana.escala_id,
            config=campana.config,
            participantes=len(participantes),
            evaluadores=evaluadores,
            evaluaciones_total=total,
            evaluaciones_completadas=completadas,
            avance=round(completadas / total * 100, 1) if total else 0.0,
            created_at=campana.created_at,
            updated_at=campana.updated_at,
        )

    async def _sync_competencias(
        self, campana: Eval360Campana, competencias, replace: bool = False
    ) -> None:
        if replace:
            for c in list(campana.competencias):
                await self.db.delete(c)
            await self.db.flush()
        for item in competencias:
            self.db.add(Eval360CampanaCompetencia(
                campana_id=campana.id,
                competencia_id=item.competencia_id,
                peso=item.peso,
                num_preguntas=item.num_preguntas,
                nivel_esperado=item.nivel_esperado,
                obligatoria=item.obligatoria,
                orden=item.orden,
            ))

    async def _sync_evaluador_tipos(
        self, campana: Eval360Campana, tipos, replace: bool = False
    ) -> None:
        if replace:
            for t in list(campana.evaluador_tipos):
                await self.db.delete(t)
            await self.db.flush()
        for item in tipos:
            self.db.add(Eval360CampanaEvaluadorTipo(
                campana_id=campana.id, tipo=item.tipo, peso=item.peso, activo=item.activo,
            ))

    async def _sync_participantes(
        self, campana: Eval360Campana, empleado_ids, replace: bool = False
    ) -> None:
        if replace:
            existentes = await self.repo.list_participantes(campana.id)
            for p in existentes:
                await self.db.delete(p)
            await self.db.flush()
        empleados = await self._get_empleados_by_ids(empleado_ids)
        # Resolver perfil de puesto/grado del empleado (si existe).
        perfiles = await self._perfiles_por_empleado([e.empleado_id for e in empleados])
        for emp in empleados:
            perfil = perfiles.get(emp.empleado_id)
            self.db.add(Eval360Participante(
                campana_id=campana.id,
                empleado_id=emp.empleado_id,
                puesto_perfil_id=perfil.puesto_perfil_id if perfil else None,
                grado_id=perfil.grado_id if perfil else None,
                estado="pendiente",
            ))

    async def _generar_evaluaciones(self, campana: Eval360Campana) -> list[Eval360Evaluacion]:
        tipos_activos = {t.tipo for t in campana.evaluador_tipos if t.activo}
        participantes = await self.repo.list_participantes(campana.id)
        existentes = {
            (e.participante_id, e.evaluador_empleado_id, e.tipo_evaluador)
            for e in await self.repo.list_evaluaciones_campana(campana.id)
        }
        cfg = campana.config or {}
        es_anonima = bool(cfg.get("anonima", False))
        fecha_limite = cfg.get("fecha_limite") or (
            campana.fecha_cierre.isoformat() if campana.fecha_cierre else None
        )
        generadas: list[Eval360Evaluacion] = []
        for p in participantes:
            if not p.empleado:
                continue
            for tipo, evaluador in await self._resolver_evaluadores(p.empleado, tipos_activos):
                evaluador_id = evaluador.empleado_id if evaluador else None
                key = (p.id, evaluador_id, tipo)
                if key in existentes:
                    continue
                existentes.add(key)
                ev = Eval360Evaluacion(
                    campana_id=campana.id,
                    participante_id=p.id,
                    evaluador_empleado_id=evaluador_id,
                    tipo_evaluador=tipo,
                    estado="pendiente",
                    es_anonima=es_anonima,
                    fecha_limite=self._parse_date(fecha_limite),
                )
                self.db.add(ev)
                generadas.append(ev)
        await self.db.flush()
        return generadas

    async def _resolver_evaluadores(
        self, empleado: Empleado, tipos_activos: set[str]
    ) -> list[tuple[str, Optional[Empleado]]]:
        """Devuelve (tipo, evaluador) sugeridos para un evaluado.

        cliente_interno/externo requieren alta manual (no se auto-sugieren).
        """
        pares: list[tuple[str, Optional[Empleado]]] = []
        if "autoevaluacion" in tipos_activos:
            pares.append(("autoevaluacion", empleado))
        if "jefe" in tipos_activos and empleado.lider_id:
            jefe = await self._get_empleado(empleado.lider_id)
            if jefe:
                pares.append(("jefe", jefe))
        if "subordinado" in tipos_activos:
            for sub in await self._get_subordinados(empleado.empleado_id):
                pares.append(("subordinado", sub))
        if "par" in tipos_activos:
            for par in await self._get_pares(empleado):
                pares.append(("par", par))
        return pares

    async def _notificar_evaluadores(
        self, campana: Eval360Campana, evaluaciones: list[Eval360Evaluacion]
    ) -> None:
        """Invitación (in-app + email) a cada evaluador al activar la campaña."""
        config = await self.get_or_create_config()
        destinatarios = {
            e.evaluador_empleado_id for e in evaluaciones if e.evaluador_empleado_id
        }
        for emp_id in destinatarios:
            asunto, cuerpo = self._texto_correo(
                config, "invitacion",
                asunto_def="Evaluación 360° asignada",
                cuerpo_def=(
                    f"Tienes evaluaciones pendientes en la campaña "
                    f"'{campana.nombre}'. Ingresa a la plataforma para responderlas."
                ),
                campana=campana.nombre,
            )
            await self._enviar_a_empleado(
                emp_id, asunto, cuerpo,
                target_url="#/mis-evaluaciones",
                metadata={"campana_id": campana.id, "evento": "invitacion"},
            )

    # SMTP aún no configurado en este entorno: las notificaciones de correo
    # quedan preparadas (plantillas + _texto_correo) pero se entregan solo
    # in-app. Cuando se habilite SMTP, cambiar EVAL360_EMAIL_HABILITADO a True.
    EVAL360_EMAIL_HABILITADO = False

    async def _enviar_a_empleado(
        self, empleado_id: int, asunto: str, cuerpo: str,
        target_url: str | None = None, metadata: Optional[dict] = None,
    ) -> None:
        """Notifica al empleado (in-app; email opcional cuando SMTP esté activo).

        Nunca propaga errores: una falla de notificación no debe romper el flujo.
        """
        try:
            email = None
            canal = "in_app"
            if self.EVAL360_EMAIL_HABILITADO:
                empleado = await self._get_empleado(empleado_id)
                email = getattr(empleado, "email", None) if empleado else None
                canal = "ambos" if email else "in_app"
            await self.notificaciones.enviar(
                destinatario_id=empleado_id,
                asunto=asunto,
                cuerpo=cuerpo,
                canal=canal,
                email_destino=email,
                target_url=target_url,
                metadata=metadata,
            )
        except Exception:  # pragma: no cover - notificacion no debe romper flujo
            logger.exception("Fallo notificando empleado %s", empleado_id)

    def _texto_correo(
        self, config: Eval360Config, evento: str, *,
        asunto_def: str, cuerpo_def: str, **ctx,
    ) -> tuple[str, str]:
        """Devuelve (asunto, cuerpo) usando overrides de config.plantillas_correo.

        Las plantillas admiten placeholders tipo {campana}, {evaluado}.
        """
        asunto, cuerpo = asunto_def, cuerpo_def
        plantillas = config.plantillas_correo or {}
        tpl = plantillas.get(evento) if isinstance(plantillas, dict) else None
        if isinstance(tpl, dict):
            asunto = tpl.get("asunto") or asunto
            cuerpo = tpl.get("cuerpo") or cuerpo
        try:
            asunto = asunto.format(**ctx)
            cuerpo = cuerpo.format(**ctx)
        except (KeyError, IndexError, ValueError):
            pass
        return asunto, cuerpo

    async def _get_evaluacion_editable(
        self, evaluacion_id: int, current_user: Empleado
    ) -> Eval360Evaluacion:
        evaluacion = await self.repo.get_evaluacion(evaluacion_id)
        if not evaluacion:
            raise NotFoundError("Evaluacion no encontrada")
        if evaluacion.evaluador_empleado_id != current_user.empleado_id:
            raise ForbiddenError("No puedes responder esta evaluacion")
        if evaluacion.estado == "completada":
            raise ConflictError("Esta evaluacion ya fue enviada")
        return evaluacion

    async def _guardar_respuestas(
        self, evaluacion: Eval360Evaluacion, data: EvaluacionRespuestasIn
    ) -> None:
        # Validar valores dentro de la escala.
        campana = await self.repo.get_campana_detalle(evaluacion.campana_id)
        escala = await self._escala_de_campana(campana)
        vmin, vmax = (escala.valor_min, escala.valor_max) if escala else (1, 5)
        preguntas_validas = await self._preguntas_validas_ids(campana)
        pregunta_comp = await self._pregunta_competencia_map(campana)

        existentes = {
            r.pregunta_id: r
            for r in await self.repo.get_respuestas_evaluacion(evaluacion.id)
        }
        for r in data.respuestas:
            if r.pregunta_id not in preguntas_validas:
                continue
            if r.valor < vmin or r.valor > vmax:
                raise DomainValidationError(
                    f"El valor {r.valor} esta fuera de la escala {vmin}-{vmax}"
                )
            if r.pregunta_id in existentes:
                existentes[r.pregunta_id].valor = r.valor
            else:
                self.db.add(Eval360Respuesta(
                    evaluacion_id=evaluacion.id,
                    pregunta_id=r.pregunta_id,
                    competencia_id=pregunta_comp.get(r.pregunta_id),
                    valor=r.valor,
                ))
        # Comentarios: reemplazar el set completo si se envian.
        if data.comentarios:
            await self.repo.delete_comentarios_evaluacion(evaluacion.id)
            for c in data.comentarios:
                self.db.add(Eval360Comentario(
                    evaluacion_id=evaluacion.id,
                    competencia_id=c.competencia_id,
                    texto=c.texto,
                    tipo=c.tipo,
                ))
        await self.db.flush()

    async def _actualizar_estado_participante(self, participante_id: int) -> None:
        participante = await self.repo.get_participante(participante_id)
        if not participante:
            return
        evals = participante.evaluaciones
        if evals and all(e.estado == "completada" for e in evals):
            participante.estado = "completada"
        elif any(e.estado in ("en_progreso", "completada") for e in evals):
            participante.estado = "en_progreso"

    # ── Calculo de resultados ─────────────────────────────────────────────────
    async def _calcular_resultados_campana(self, campana: Eval360Campana) -> None:
        participantes = await self.repo.list_participantes(campana.id)
        for p in participantes:
            await self._calcular_resultados_participante(p, campana)

    async def _calcular_resultados_participante(
        self, participante: Eval360Participante, campana: Eval360Campana
    ) -> None:
        escala = await self._escala_de_campana(campana)
        vmin, vmax = (escala.valor_min, escala.valor_max) if escala else (1, 5)
        pesos = {t.tipo: _f(t.peso) or 0.0 for t in campana.evaluador_tipos if t.activo}
        comp_cfg = {c.competencia_id: c for c in campana.competencias}

        # Reunir respuestas de evaluaciones completadas del participante.
        evaluaciones = [
            e for e in participante.evaluaciones if e.estado == "completada"
        ]
        # tipo -> competencia_id -> [valores]
        datos: dict[str, dict[int, list[float]]] = {}
        for ev in evaluaciones:
            respuestas = await self.repo.get_respuestas_evaluacion(ev.id)
            for r in respuestas:
                datos.setdefault(ev.tipo_evaluador, {}).setdefault(
                    r.competencia_id, []
                ).append(_f(r.valor))

        await self.repo.delete_resultados_participante(participante.id)

        calificaciones_comp: list[tuple[float, float]] = []  # (promedio, peso_comp)
        for comp_id, cc in comp_cfg.items():
            por_tipo: dict[str, float] = {}
            for tipo, comps in datos.items():
                vals = comps.get(comp_id)
                if vals:
                    por_tipo[tipo] = round(sum(vals) / len(vals), 2)
            if not por_tipo:
                continue
            auto = por_tipo.get("autoevaluacion")
            # Promedio general ponderado por pesos configurados (renormalizado).
            promedio_general = self._promedio_ponderado(por_tipo, pesos)
            esperado_likert = self._nivel_a_likert(cc.nivel_esperado, vmin, vmax)
            brecha = round(promedio_general - esperado_likert, 2)
            estado_brecha = (
                "cumple" if brecha >= 0 else ("riesgo" if brecha >= -1 else "brecha")
            )
            self.db.add(Eval360Resultado(
                participante_id=participante.id,
                competencia_id=comp_id,
                promedio_general=promedio_general,
                promedio_por_tipo=por_tipo,
                autoevaluacion=auto,
                nivel_esperado=esperado_likert,
                brecha=brecha,
                estado_brecha=estado_brecha,
            ))
            peso_comp = _f(cc.peso) or 0.0
            calificaciones_comp.append((promedio_general, peso_comp))

        # Fila resumen global (competencia_id NULL).
        if calificaciones_comp:
            total_peso = sum(pc for _, pc in calificaciones_comp)
            if total_peso > 0:
                cal_general = round(
                    sum(v * pc for v, pc in calificaciones_comp) / total_peso, 2
                )
            else:
                cal_general = round(
                    sum(v for v, _ in calificaciones_comp) / len(calificaciones_comp), 2
                )
            self.db.add(Eval360Resultado(
                participante_id=participante.id,
                competencia_id=None,
                calificacion_general=cal_general,
                promedio_general=cal_general,
            ))
            if participante.estado != "completada":
                participante.estado = "en_progreso"

    def _promedio_ponderado(
        self, por_tipo: dict[str, float], pesos: dict[str, float]
    ) -> float:
        tipos_presentes = list(por_tipo.keys())
        peso_total = sum(pesos.get(t, 0.0) for t in tipos_presentes)
        if peso_total > 0:
            return round(
                sum(por_tipo[t] * pesos.get(t, 0.0) for t in tipos_presentes) / peso_total,
                2,
            )
        # Sin pesos: media simple.
        return round(sum(por_tipo.values()) / len(por_tipo), 2)

    def _nivel_a_likert(self, nivel_0_4: int, vmin: int, vmax: int) -> float:
        """Convierte un nivel esperado 0-4 a la escala Likert de la campana."""
        frac = max(0, min(nivel_0_4, 4)) / 4.0
        return round(vmin + frac * (vmax - vmin), 2)

    async def _resultado_participante(
        self, participante_id: int
    ) -> ResultadoParticipanteResponse:
        participante = await self.repo.get_participante(participante_id)
        resultados = await self.repo.list_resultados_participante(participante_id)
        comp_ids = [r.competencia_id for r in resultados if r.competencia_id]
        comp_nombres = await self._competencia_nombres(comp_ids)
        competencias: list[ResultadoCompetencia] = []
        calificacion_general = None
        for r in resultados:
            if r.competencia_id is None:
                calificacion_general = _f(r.calificacion_general)
                continue
            competencias.append(ResultadoCompetencia(
                competencia_id=r.competencia_id,
                competencia_nombre=comp_nombres.get(r.competencia_id),
                promedio_general=_f(r.promedio_general),
                promedio_por_tipo=r.promedio_por_tipo,
                autoevaluacion=_f(r.autoevaluacion),
                nivel_esperado=_f(r.nivel_esperado),
                brecha=_f(r.brecha),
                estado_brecha=r.estado_brecha,
            ))
        ordenadas = sorted(
            competencias, key=lambda c: (c.promedio_general or 0), reverse=True
        )
        fortalezas = [c.competencia_nombre for c in ordenadas[:3] if c.competencia_nombre]
        oportunidades = [
            c.competencia_nombre
            for c in sorted(competencias, key=lambda c: (c.promedio_general or 0))[:3]
            if c.estado_brecha in ("riesgo", "brecha") and c.competencia_nombre
        ]
        return ResultadoParticipanteResponse(
            participante_id=participante_id,
            empleado_id=participante.empleado_id if participante else 0,
            empleado_nombre=(
                participante.empleado.nombre
                if participante and participante.empleado else None
            ),
            puesto=self._puesto_nombre(participante.empleado) if participante else None,
            calificacion_general=calificacion_general,
            competencias=competencias,
            fortalezas=fortalezas,
            oportunidades=oportunidades,
        )

    async def _avance_por_campana(self) -> list[CampanaAvance]:
        campanas = (
            await self.db.execute(
                select(Eval360Campana).where(
                    Eval360Campana.activo.is_(True),
                    Eval360Campana.estado.in_(["activa", "en_progreso", "finalizada", "cerrada"]),
                ).order_by(Eval360Campana.id.desc()).limit(8)
            )
        ).scalars().all()
        out: list[CampanaAvance] = []
        for c in campanas:
            evals = await self.repo.list_evaluaciones_campana(c.id)
            total = len(evals)
            completadas = sum(1 for e in evals if e.estado == "completada")
            out.append(CampanaAvance(
                campana_id=c.id, nombre=c.nombre,
                avance=round(completadas / total * 100, 1) if total else 0.0,
            ))
        return out

    async def _distribucion_calificaciones(self) -> list[DashboardSeriePunto]:
        resultados = (
            await self.db.execute(
                select(Eval360Resultado).where(Eval360Resultado.competencia_id.isnot(None))
            )
        ).scalars().all()
        buckets = {"1-2": 0, "2-3": 0, "3-4": 0, "4-5": 0}
        for r in resultados:
            v = _f(r.promedio_general)
            if v is None:
                continue
            if v < 2:
                buckets["1-2"] += 1
            elif v < 3:
                buckets["2-3"] += 1
            elif v < 4:
                buckets["3-4"] += 1
            else:
                buckets["4-5"] += 1
        return [DashboardSeriePunto(label=k, valor=float(v)) for k, v in buckets.items()]

    # ── Acceso a empleados / competencias ─────────────────────────────────────
    async def _get_empleado(self, empleado_id: int) -> Optional[Empleado]:
        return await self.db.get(Empleado, empleado_id)

    async def _get_empleados_by_ids(self, ids) -> list[Empleado]:
        ids = [i for i in dict.fromkeys(ids)]
        if not ids:
            return []
        result = await self.db.execute(
            select(Empleado).where(Empleado.empleado_id.in_(ids))
        )
        return list(result.scalars().all())

    async def _get_campanas_by_ids(self, ids) -> list[Eval360Campana]:
        if not ids:
            return []
        result = await self.db.execute(
            select(Eval360Campana).where(Eval360Campana.id.in_(list(ids)))
        )
        return list(result.scalars().all())

    async def _get_subordinados(self, empleado_id: int) -> list[Empleado]:
        result = await self.db.execute(
            select(Empleado).where(Empleado.lider_id == empleado_id)
        )
        return list(result.scalars().all())

    async def _get_pares(self, empleado: Empleado) -> list[Empleado]:
        # Pares = mismo subarea (fallback area), excluyendo self y el jefe directo.
        if empleado.subarea_id:
            filtro = Empleado.subarea_id == empleado.subarea_id
        elif empleado.area_id:
            filtro = Empleado.area_id == empleado.area_id
        else:
            return []
        result = await self.db.execute(
            select(Empleado)
            .where(filtro, Empleado.empleado_id != empleado.empleado_id)
            .order_by(Empleado.empleado_id)
            .limit(MAX_PARES_SUGERIDOS)
        )
        excluidos = {empleado.lider_id}
        return [e for e in result.scalars().all() if e.empleado_id not in excluidos]

    async def _perfiles_por_empleado(self, empleado_ids) -> dict[int, PerfilFunciones]:
        if not empleado_ids:
            return {}
        result = await self.db.execute(
            select(PerfilFunciones).where(
                PerfilFunciones.empleado_id.in_(list(empleado_ids))
            )
        )
        perfiles: dict[int, PerfilFunciones] = {}
        for pf in result.scalars().all():
            perfiles.setdefault(pf.empleado_id, pf)
        return perfiles

    async def _competencia_nombres(self, comp_ids) -> dict[int, str]:
        comp_ids = [c for c in dict.fromkeys(comp_ids) if c]
        if not comp_ids:
            return {}
        result = await self.db.execute(
            select(Competencia.id, Competencia.nombre).where(
                Competencia.id.in_(comp_ids)
            )
        )
        return {row[0]: row[1] for row in result.all()}

    async def _escala_de_campana(self, campana: Eval360Campana) -> Optional[Eval360Escala]:
        if campana and campana.escala_id:
            escala = await self.repo.get_escala(campana.escala_id)
            if escala:
                return escala
        config = await self.get_or_create_config()
        if config.escala_id:
            return await self.repo.get_escala(config.escala_id)
        return None

    async def _preguntas_validas_ids(self, campana: Eval360Campana) -> set[int]:
        ids: set[int] = set()
        for cc in campana.competencias:
            preguntas = await self.repo.list_preguntas(
                competencia_id=cc.competencia_id, solo_activas=True
            )
            if cc.num_preguntas:
                preguntas = list(preguntas)[: cc.num_preguntas]
            ids.update(p.id for p in preguntas)
        return ids

    async def _pregunta_competencia_map(self, campana: Eval360Campana) -> dict[int, int]:
        mapping: dict[int, int] = {}
        for cc in campana.competencias:
            preguntas = await self.repo.list_preguntas(
                competencia_id=cc.competencia_id, solo_activas=True
            )
            for p in preguntas:
                mapping[p.id] = cc.competencia_id
        return mapping

    async def _num_preguntas_evaluacion(
        self, evaluacion: Eval360Evaluacion, campana: Optional[Eval360Campana] = None
    ) -> int:
        if campana is None:
            campana = await self.repo.get_campana_detalle(evaluacion.campana_id)
        ids = await self._preguntas_validas_ids(campana)
        return len(ids)

    def _puesto_nombre(self, empleado: Optional[Empleado]) -> Optional[str]:
        if empleado and empleado.puesto:
            return empleado.puesto.descripcion
        return None

    def _area_nombre(self, empleado: Optional[Empleado]) -> Optional[str]:
        if empleado and empleado.area:
            return empleado.area.descripcion
        return None

    def _parse_date(self, value):
        if not value:
            return None
        if isinstance(value, str):
            try:
                return datetime.fromisoformat(value).date()
            except ValueError:
                return None
        return value
