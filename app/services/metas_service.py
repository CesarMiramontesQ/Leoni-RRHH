# app/services/metas_service.py
"""Logica de negocio del modulo Metas (OKR ligero).

Responsabilidades:
  - Ciclo de vida de `MetaCiclo`: borrador -> activo -> cerrado.
  - CRUD de `Meta` (individual/equipo) y sus `MetaResultadoClave`, con las
    validaciones cruzadas del spec (nivel/empleado_id/area_id/lider_id,
    `meta_padre_id`, ciclo activo).
  - Seguimiento: `registrar_checkin` (empleado o ajuste del jefe).
  - Cierre: `cerrar_meta` (calificacion 0-100 por el jefe) y `cerrar_ciclo`
    (congela metas de equipo pendientes y calcula cumplimiento).
  - Formulas de avance/cumplimiento (funciones puras + wrappers de service,
    ver seccion "Calculo" mas abajo).

El commit lo realiza la dependencia `get_db` al cierre del request; aqui solo
se usa flush() (via el repositorio), como el resto de services del proyecto.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Optional

from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import ConflictError, DomainValidationError, NotFoundError
from app.models.metas import Meta, MetaCheckin, MetaCiclo, MetaResultadoClave
from app.repositories.metas_repository import MetasRepository
from app.schemas.metas import (
    CheckinResponse,
    EquipoAvanceMiembro,
    EquipoAvanceResponse,
    MetaCicloCreate,
    MetaCicloResponse,
    MetaCicloUpdate,
    MetaCreate,
    MetaFiltros,
    MetaResponse,
    MetaUpdate,
    RecordatoriosResultado,
    ResultadoClaveCreate,
    ResultadoClaveResponse,
    ResultadoClaveUpdate,
)
from app.services.notificacion_service import NotificacionService

Numero = Decimal | int | float

# Recordatorios (Tarea 5) — ver `MetasService.procesar_recordatorios`.
MIS_METAS_TARGET_URL = "#/talento/mis-metas"
DIAS_CIERRE_PROXIMO_DEFAULT = 3
DIAS_SIN_CHECKIN_DEFAULT = 7
# Cadencia minima (dias) entre dos recordatorios automaticos de la MISMA meta
# (fix post-revision: sin esto, el job diario re-notificaba cada corrida
# mientras la condicion se siguiera cumpliendo — en particular "RC sin
# check-in" generaba spam diario indefinido). Sin cadencia configurable por
# ciclo a proposito (mantenido simple, ver `Meta.ultimo_recordatorio_at`).
# Mismo patron que `EncuestaParticipante.ultimo_recordatorio_at` /
# `Encuesta.recordatorio_cada_dias` en Encuestas RH, simplificado a una
# constante (no hay equivalente de "cadencia por ciclo" en el spec de Metas).
RECORDATORIO_CADENCIA_DIAS = 3


# ══════════════════════════════════════════════════════════════════════════
# Calculo — funciones puras (exportables para test directo, sin BD)
# ══════════════════════════════════════════════════════════════════════════


def _clamp_round_pct(valor: float, ndigits: int = 0) -> float:
    """Clamp a [0, 100] y redondeo half-up (no half-even: 62.5 -> 63, no 62).

    Se usa para toda cifra "tipo porcentaje" del modulo (avance de RC/meta,
    cumplimiento ponderado) para que el resultado sea estable y documentado.
    """
    clamped = max(0.0, min(100.0, valor))
    quantum = Decimal(1).scaleb(-ndigits)
    return float(Decimal(str(clamped)).quantize(quantum, rounding=ROUND_HALF_UP))


def calcular_avance_rc(
    tipo_metrica: str,
    direccion: str,
    valor_inicial: Numero,
    valor_objetivo: Numero,
    valor_actual: Numero,
) -> float:
    """Avance % (0-100, clamp) de un resultado clave. Ver spec
    "Formula de avance y cumplimiento":

      - `booleano`: 100 si `valor_actual == valor_objetivo`, si no 0
        (direccion no aplica: el booleano no tiene "sentido" de avance).
      - `subir`: `(actual - inicial) / (objetivo - inicial)`.
      - `bajar`: `(inicial - actual) / (inicial - objetivo)`.
      - Borde documentado: denominador 0 (objetivo == inicial, tipos no
        booleanos) -> 100 si `actual` ya cumple el objetivo, si no 0.
    """
    ini = float(valor_inicial)
    obj = float(valor_objetivo)
    act = float(valor_actual)

    if tipo_metrica == "booleano":
        raw = 100.0 if act == obj else 0.0
    elif direccion == "subir":
        denom = obj - ini
        if denom == 0:
            raw = 100.0 if act >= obj else 0.0
        else:
            raw = (act - ini) / denom * 100.0
    elif direccion == "bajar":
        denom = ini - obj
        if denom == 0:
            raw = 100.0 if act <= obj else 0.0
        else:
            raw = (ini - act) / denom * 100.0
    else:
        raise ValueError(f"direccion invalida: {direccion!r}")

    return _clamp_round_pct(raw)


def _validar_rc_valores(tipo_metrica: str, valor_inicial: Numero, valor_objetivo: Numero) -> None:
    """RC con `valor_objetivo == valor_inicial` solo se permite si es
    `booleano` (donde ambos valores no importan para el avance) — ver spec
    "Ciclo de vida y validaciones"."""
    if tipo_metrica != "booleano" and Decimal(str(valor_objetivo)) == Decimal(str(valor_inicial)):
        raise DomainValidationError(
            "valor_objetivo debe ser distinto de valor_inicial (salvo tipo_metrica booleano)",
            field="valor_objetivo",
        )


def _dt_utc(value: Optional[datetime]) -> Optional[datetime]:
    """Normaliza un datetime leido de BD a aware UTC (Postgres/asyncpg
    devuelve datetimes aware para timestamptz; SQLite en tests los devuelve
    naive). Mismo patron que `EncuestasRhService._dt_utc`."""
    if value is None:
        return None
    return value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)


def _rc_dias_sin_checkin(
    rc: MetaResultadoClave, meta_creada_at: datetime, ahora: datetime
) -> int:
    """Dias transcurridos desde el ultimo check-in del RC. Si el RC nunca
    tuvo un check-in, se usa la fecha de creacion de la meta (asignacion)
    como referencia — un RC recien asignado y nunca tocado tambien debe
    poder marcarse como "estancado" tras M dias."""
    checkins = list(rc.checkins)
    if checkins:
        referencia = max(_dt_utc(c.created_at) for c in checkins)
    else:
        referencia = _dt_utc(meta_creada_at)
    return (ahora - referencia).days


class MetasService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = MetasRepository(db)
        self.notificaciones = NotificacionService(db)

    # ══════════════════════════════════════════════════════════════════════
    # Calculo — wrappers de service (operan sobre objetos ya cargados)
    # ══════════════════════════════════════════════════════════════════════
    def avance_rc(self, rc: MetaResultadoClave) -> float:
        return calcular_avance_rc(
            rc.tipo_metrica, rc.direccion, rc.valor_inicial, rc.valor_objetivo, rc.valor_actual
        )

    def avance_meta(self, meta: Meta) -> float:
        """Promedio simple de los avances de sus resultados clave.

        Roll-up de meta de equipo (decision documentada en el spec): si la
        meta de equipo tiene resultados clave propios, esos mandan (se
        promedian igual que una meta individual); si no tiene, el avance es
        el promedio del avance de sus `submetas` (metas individuales
        enlazadas via `meta_padre_id`). Sin RC y sin submetas -> 0.
        """
        if meta.nivel == "equipo" and not meta.resultados_clave:
            submetas = list(meta.submetas or [])
            if not submetas:
                return 0.0
            return _clamp_round_pct(
                sum(self.avance_meta(sm) for sm in submetas) / len(submetas)
            )

        if not meta.resultados_clave:
            return 0.0
        return _clamp_round_pct(
            sum(self.avance_rc(rc) for rc in meta.resultados_clave) / len(meta.resultados_clave)
        )

    async def cumplimiento_empleado(self, ciclo_id: int, empleado_id: int) -> float:
        """Cumplimiento ponderado del empleado en el ciclo:
        `Σ(peso_i × calificacion_cierre_i) / Σ(peso_i)` sobre sus metas
        individuales ya cerradas (calificadas). Sin metas cerradas -> 0
        (borde documentado, evita division por cero)."""
        metas = await self.repo.list_metas_cerradas_empleado(ciclo_id, empleado_id)
        total_peso = sum(float(m.peso) for m in metas)
        if total_peso <= 0:
            return 0.0
        total_ponderado = sum(
            float(m.peso) * float(m.calificacion_cierre or 0) for m in metas
        )
        return _clamp_round_pct(total_ponderado / total_peso, ndigits=2)

    # ══════════════════════════════════════════════════════════════════════
    # Tablero de equipo / export (Tarea 4)
    # ══════════════════════════════════════════════════════════════════════
    @staticmethod
    def _avance_global(metas: list[MetaResponse]) -> float:
        """Promedio ponderado por `peso` del `avance` (derivado, no la
        calificacion de cierre) de un conjunto de metas individuales. Sin
        metas o con peso total 0 -> 0.0 (mismo borde que `cumplimiento_empleado`,
        documentado ahi)."""
        total_peso = sum(float(m.peso) for m in metas)
        if total_peso <= 0:
            return 0.0
        total_ponderado = sum(float(m.peso) * m.avance for m in metas)
        return _clamp_round_pct(total_ponderado / total_peso, ndigits=2)

    async def construir_equipo_avance(
        self, ciclo_id: int, metas: list[MetaResponse]
    ) -> EquipoAvanceResponse:
        """Agrupa `metas` (ya resueltas/filtradas por el scope de equipo en
        el router, ver `_list_metas_scoped`) por empleado para el tablero
        `GET /equipo/avance`: por miembro, sus metas individuales, su avance
        global ponderado (`_avance_global`) y el nombre del empleado
        (`MetasRepository.get_nombres_empleados`, lectura Bono). Las metas
        de nivel "equipo" (lider_id, sin empleado_id) van aparte en
        `metas_equipo` — no pertenecen a "un miembro"."""
        individuales = [m for m in metas if m.nivel == "individual"]
        metas_equipo = [m for m in metas if m.nivel == "equipo"]

        orden: list[int] = []
        por_empleado: dict[int, list[MetaResponse]] = {}
        for m in individuales:
            eid = m.empleado_id
            if eid is None:
                continue
            if eid not in por_empleado:
                por_empleado[eid] = []
                orden.append(eid)
            por_empleado[eid].append(m)

        nombres = await self.repo.get_nombres_empleados(orden)
        miembros = [
            EquipoAvanceMiembro(
                empleado_id=eid,
                empleado_nombre=nombres.get(eid),
                metas=por_empleado[eid],
                avance_global=self._avance_global(por_empleado[eid]),
            )
            for eid in orden
        ]
        return EquipoAvanceResponse(
            ciclo_id=ciclo_id, miembros=miembros, metas_equipo=metas_equipo
        )

    async def exportar_ciclo_excel(self, ciclo_id: int, metas: list[MetaResponse]) -> BytesIO:
        """Exporta a un `.xlsx` (una hoja) las `metas` del ciclo (ya
        resueltas/filtradas por el scope de equipo en el router, mismo
        patron que `construir_equipo_avance`): meta, avance derivado y
        cumplimiento ponderado del empleado (solo metas individuales,
        calculado una vez por empleado vía `cumplimiento_empleado` — que a
        su vez solo considera metas YA CERRADAS/calificadas; durante el
        ciclo aparece en 0.0, ver borde documentado ahi). Patron de export
        Excel con openpyxl: `Evaluacion360Service._resultados_campana_excel`."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        ciclo = await self.get_ciclo(ciclo_id)

        empleados_individuales = sorted(
            {m.empleado_id for m in metas if m.nivel == "individual" and m.empleado_id is not None}
        )
        cumplimientos = {
            eid: await self.cumplimiento_empleado(ciclo_id, eid) for eid in empleados_individuales
        }
        nombres = await self.repo.get_nombres_empleados(empleados_individuales)

        wb = Workbook()
        ws = wb.active
        ws.title = "Metas y avance"
        ws.cell(row=1, column=1, value=f"Metas — {ciclo.nombre}").font = Font(bold=True, size=14)
        headers = [
            "Empleado", "Meta", "Nivel", "Estado", "Peso",
            "Avance %", "Calificación cierre", "Cumplimiento ponderado",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h).font = Font(bold=True)

        row = 4
        for m in metas:
            if m.nivel == "individual":
                empleado_label = nombres.get(m.empleado_id, str(m.empleado_id))
                cumplimiento = cumplimientos.get(m.empleado_id)
            else:
                empleado_label = f"Equipo (líder {m.lider_id})"
                cumplimiento = None
            ws.cell(row=row, column=1, value=empleado_label)
            ws.cell(row=row, column=2, value=m.titulo)
            ws.cell(row=row, column=3, value=m.nivel)
            ws.cell(row=row, column=4, value=m.estado)
            ws.cell(row=row, column=5, value=float(m.peso))
            ws.cell(row=row, column=6, value=m.avance)
            ws.cell(
                row=row, column=7,
                value=float(m.calificacion_cierre) if m.calificacion_cierre is not None else None,
            )
            ws.cell(row=row, column=8, value=cumplimiento)
            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════════════════
    # Ciclo
    # ══════════════════════════════════════════════════════════════════════
    async def crear_ciclo(self, data: MetaCicloCreate) -> MetaCicloResponse:
        if data.fecha_fin <= data.fecha_inicio:
            raise DomainValidationError(
                "fecha_fin debe ser posterior a fecha_inicio", field="fecha_fin"
            )
        ciclo = MetaCiclo(
            nombre=data.nombre,
            descripcion=data.descripcion,
            fecha_inicio=data.fecha_inicio,
            fecha_fin=data.fecha_fin,
            estado="borrador",
            creado_por_id=data.creado_por_id,
            metas=[],
        )
        self.db.add(ciclo)
        await self.db.flush()
        await self.db.refresh(ciclo)
        return self._ciclo_to_response(ciclo)

    async def get_ciclo(self, ciclo_id: int) -> MetaCicloResponse:
        ciclo = await self.repo.get_ciclo(ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", ciclo_id)
        return self._ciclo_to_response(ciclo)

    async def actualizar_ciclo(self, ciclo_id: int, data: MetaCicloUpdate) -> MetaCicloResponse:
        """Edita nombre/descripcion/fechas de un ciclo (Tarea 3, fix post-revision).

        Politica de edicion (decision documentada, no estaba en el spec de
        Tarea 2): se permite editar un ciclo en "borrador" o "activo" —
        nombre/descripcion sin restriccion, y fechas con la unica validacion
        de que `fecha_fin >= fecha_inicio` (se usa `>=`, no `>`, a proposito:
        mas laxo que `crear_ciclo`, que exige `>` estricto en la creacion
        inicial; edición admite igualar limites en un ajuste puntual). Un
        ciclo "cerrado" es inmutable -> `ConflictError` (409): ya congelo
        metas de equipo y sirvio de base para `cumplimiento_empleado`,
        cambiar sus fechas retroactivamente no tiene sentido de negocio.
        """
        ciclo = await self.repo.get_ciclo(ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", ciclo_id)
        if ciclo.estado == "cerrado":
            raise ConflictError("No se puede editar un ciclo cerrado")

        payload = data.model_dump(exclude_unset=True)
        nueva_fecha_inicio = payload.get("fecha_inicio", ciclo.fecha_inicio)
        nueva_fecha_fin = payload.get("fecha_fin", ciclo.fecha_fin)
        if nueva_fecha_fin < nueva_fecha_inicio:
            raise DomainValidationError(
                "fecha_fin debe ser mayor o igual a fecha_inicio", field="fecha_fin"
            )

        for key, value in payload.items():
            setattr(ciclo, key, value)
        await self.db.flush()
        await self.db.refresh(ciclo)
        return self._ciclo_to_response(ciclo)

    async def activar_ciclo(self, ciclo_id: int) -> MetaCicloResponse:
        ciclo = await self.repo.get_ciclo(ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", ciclo_id)
        if ciclo.estado != "borrador":
            raise ConflictError("Solo se puede activar un ciclo en borrador")
        ciclo.estado = "activo"
        await self.db.flush()
        await self.db.refresh(ciclo)
        return self._ciclo_to_response(ciclo)

    async def cerrar_ciclo(self, ciclo_id: int) -> MetaCicloResponse:
        """activo -> cerrado: congela el ciclo y calcula cumplimiento.

        Decision documentada (spec: "exige calificacion previa o permite
        calificar al cierre"): se exige calificacion previa para METAS
        INDIVIDUALES — deben cerrarse (via `cerrar_meta`) antes de cerrar el
        ciclo, porque `cumplimiento_empleado` solo tiene sentido con
        calificaciones ya puestas por el jefe (si se auto-cerraran sin
        calificar, quedarian con `calificacion_cierre=None` y distorsionarian
        la ponderacion). Las metas de EQUIPO, en cambio, no alimentan
        `cumplimiento_empleado` directamente, asi que se congelan
        automaticamente (pasan a "cerrada" sin exigir calificacion previa).
        """
        ciclo = await self.repo.get_ciclo(ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", ciclo_id)
        if ciclo.estado != "activo":
            raise ConflictError("Solo se puede cerrar un ciclo activo")

        pendientes = await self.repo.list_metas_individuales_no_cerradas(ciclo_id)
        if pendientes:
            ids = ", ".join(str(m.id) for m in pendientes)
            raise ConflictError(
                f"No se puede cerrar el ciclo: {len(pendientes)} meta(s) individual(es) "
                f"sin calificar (usar cerrar_meta primero): {ids}"
            )

        metas_equipo = await self.repo.list_metas_equipo_no_cerradas(ciclo_id)
        for meta_equipo in metas_equipo:
            meta_equipo.estado = "cerrada"

        ciclo.estado = "cerrado"
        await self.db.flush()
        await self.db.refresh(ciclo)
        return self._ciclo_to_response(ciclo)

    async def list_ciclos(self, estado: Optional[str] = None) -> list[MetaCicloResponse]:
        ciclos = await self.repo.list_ciclos(estado=estado)
        return [self._ciclo_to_response(c) for c in ciclos]

    # ══════════════════════════════════════════════════════════════════════
    # Meta
    # ══════════════════════════════════════════════════════════════════════
    async def crear_meta(self, data: MetaCreate) -> MetaResponse:
        ciclo = await self.repo.get_ciclo(data.ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", data.ciclo_id)
        if ciclo.estado != "activo":
            raise ConflictError("Solo se pueden asignar metas en un ciclo activo")

        if data.nivel == "individual":
            if data.empleado_id is None:
                raise DomainValidationError(
                    "empleado_id es obligatorio para una meta individual",
                    field="empleado_id",
                )
        else:  # equipo
            if data.area_id is None or data.lider_id is None:
                raise DomainValidationError(
                    "area_id y lider_id son obligatorios para una meta de equipo",
                    field="area_id",
                )

        if data.meta_padre_id is not None:
            await self._validar_meta_padre(data.meta_padre_id, data.ciclo_id, data.nivel)

        for rc_data in data.resultados_clave:
            _validar_rc_valores(rc_data.tipo_metrica, rc_data.valor_inicial, rc_data.valor_objetivo)

        meta = Meta(
            ciclo_id=data.ciclo_id,
            nivel=data.nivel,
            empleado_id=data.empleado_id,
            area_id=data.area_id,
            lider_id=data.lider_id,
            titulo=data.titulo,
            descripcion=data.descripcion,
            peso=data.peso,
            estado="asignada",
            meta_padre_id=data.meta_padre_id,
            asignada_por_id=data.asignada_por_id,
            resultados_clave=[],
        )
        self.db.add(meta)
        for orden, rc_data in enumerate(data.resultados_clave, start=1):
            meta.resultados_clave.append(self._nuevo_rc_obj(rc_data, orden_defecto=orden))
        await self.db.flush()
        return await self.get_meta(meta.id)

    async def _validar_meta_padre(
        self, meta_padre_id: int, ciclo_id: int, nivel_hijo: str
    ) -> Meta:
        if nivel_hijo != "individual":
            raise DomainValidationError(
                "meta_padre_id solo aplica a una meta hija de nivel individual "
                "(las submetas enlazadas a un padre de equipo son individuales)",
                field="meta_padre_id",
            )
        meta_padre = await self.repo.get_meta(meta_padre_id)
        if not meta_padre:
            raise NotFoundError("Meta", meta_padre_id)
        if meta_padre.nivel != "equipo":
            raise DomainValidationError(
                "meta_padre_id debe apuntar a una meta de nivel equipo",
                field="meta_padre_id",
            )
        if meta_padre.ciclo_id != ciclo_id:
            raise DomainValidationError(
                "meta_padre_id debe pertenecer al mismo ciclo",
                field="meta_padre_id",
            )
        return meta_padre

    @staticmethod
    def _nuevo_rc_obj(data: ResultadoClaveCreate, orden_defecto: int) -> MetaResultadoClave:
        return MetaResultadoClave(
            orden=data.orden if data.orden is not None else orden_defecto,
            titulo=data.titulo,
            tipo_metrica=data.tipo_metrica,
            unidad=data.unidad,
            direccion=data.direccion,
            valor_inicial=data.valor_inicial,
            valor_objetivo=data.valor_objetivo,
            valor_actual=data.valor_actual if data.valor_actual is not None else data.valor_inicial,
            checkins=[],
        )

    async def get_meta(self, meta_id: int) -> MetaResponse:
        meta = await self.repo.get_meta(meta_id)
        if not meta:
            raise NotFoundError("Meta", meta_id)
        return self._meta_to_response(meta)

    async def list_metas(self, filtros: MetaFiltros) -> list[MetaResponse]:
        metas = await self.repo.list_metas(
            ciclo_id=filtros.ciclo_id, empleado_id=filtros.empleado_id, nivel=filtros.nivel
        )
        return [self._meta_to_response(m) for m in metas]

    async def actualizar_meta(self, meta_id: int, data: MetaUpdate) -> MetaResponse:
        meta = await self._get_meta_ciclo_activo(meta_id)
        if meta.estado == "cerrada":
            raise ConflictError("No se puede editar una meta cerrada")

        payload = data.model_dump(exclude_unset=True)

        if "empleado_id" in payload and meta.nivel != "individual":
            raise DomainValidationError(
                "empleado_id solo aplica a metas de nivel individual", field="empleado_id"
            )
        if ("area_id" in payload or "lider_id" in payload) and meta.nivel != "equipo":
            raise DomainValidationError(
                "area_id/lider_id solo aplican a metas de nivel equipo", field="area_id"
            )
        if "meta_padre_id" in payload and payload["meta_padre_id"] is not None:
            await self._validar_meta_padre(payload["meta_padre_id"], meta.ciclo_id, meta.nivel)

        for key, value in payload.items():
            setattr(meta, key, value)
        await self.db.flush()
        return await self.get_meta(meta.id)

    async def eliminar_meta(self, meta_id: int) -> None:
        meta = await self._get_meta_ciclo_activo(meta_id)
        tiene_checkins = any(rc.checkins for rc in meta.resultados_clave)
        if tiene_checkins:
            raise ConflictError("No se puede eliminar una meta con check-ins registrados")
        await self.db.delete(meta)
        await self.db.flush()

    async def cerrar_meta(
        self,
        meta_id: int,
        calificacion: Numero,
        comentario: Optional[str] = None,
        actor_id: Optional[int] = None,
    ) -> MetaResponse:
        """Califica y cierra una meta (asignada/en_progreso -> cerrada).

        `actor_id` se acepta para simetria con el futuro endpoint (Tarea 3
        valida ahi permisos/scoping de equipo); esta capa no lo audita
        todavia — no hay campo de auditoria en el modelo de Tarea 1.
        """
        meta = await self._get_meta_ciclo_activo(meta_id)
        if meta.estado == "cerrada":
            raise ConflictError("La meta ya esta cerrada")

        calificacion_dec = Decimal(str(calificacion))
        if not (Decimal("0") <= calificacion_dec <= Decimal("100")):
            raise DomainValidationError(
                "calificacion debe estar entre 0 y 100", field="calificacion"
            )

        meta.calificacion_cierre = calificacion_dec
        meta.comentario_cierre = comentario
        meta.estado = "cerrada"
        await self.db.flush()
        return await self.get_meta(meta.id)

    async def _get_meta_ciclo_activo(self, meta_id: int) -> Meta:
        meta = await self.repo.get_meta(meta_id)
        if not meta:
            raise NotFoundError("Meta", meta_id)
        if meta.ciclo.estado != "activo":
            raise ConflictError(
                "Solo se pueden modificar metas de un ciclo activo"
            )
        return meta

    # ── Self-service (empleado) ───────────────────────────────────────────
    async def list_mis_metas(
        self, empleado_id: int, ciclo_id: Optional[int] = None
    ) -> list[MetaResponse]:
        metas = await self.repo.list_metas(
            ciclo_id=ciclo_id, empleado_id=empleado_id, nivel="individual"
        )
        return [self._meta_to_response(m) for m in metas]

    async def get_mi_meta(self, meta_id: int, empleado_id: int) -> MetaResponse:
        meta = await self.repo.get_meta(meta_id)
        if not meta or meta.empleado_id != empleado_id:
            raise NotFoundError("Meta", meta_id)
        return self._meta_to_response(meta)

    # ══════════════════════════════════════════════════════════════════════
    # Resultado clave
    # ══════════════════════════════════════════════════════════════════════
    async def agregar_rc(self, meta_id: int, data: ResultadoClaveCreate) -> ResultadoClaveResponse:
        meta = await self._get_meta_ciclo_activo(meta_id)
        if meta.estado == "cerrada":
            raise ConflictError("No se pueden agregar resultados clave a una meta cerrada")
        _validar_rc_valores(data.tipo_metrica, data.valor_inicial, data.valor_objetivo)

        rc = self._nuevo_rc_obj(data, orden_defecto=len(meta.resultados_clave) + 1)
        rc.meta_id = meta_id
        self.db.add(rc)
        await self.db.flush()
        await self.db.refresh(rc)
        return self._rc_to_response(rc)

    async def actualizar_rc(self, rc_id: int, data: ResultadoClaveUpdate) -> ResultadoClaveResponse:
        rc = await self._get_rc_editable(rc_id)
        payload = data.model_dump(exclude_unset=True)
        if "valor_objetivo" in payload:
            _validar_rc_valores(rc.tipo_metrica, rc.valor_inicial, payload["valor_objetivo"])
        for key, value in payload.items():
            setattr(rc, key, value)
        await self.db.flush()
        return self._rc_to_response(rc)

    async def eliminar_rc(self, rc_id: int) -> None:
        rc = await self._get_rc_editable(rc_id)
        if rc.checkins:
            raise ConflictError(
                "No se puede eliminar un resultado clave con check-ins registrados"
            )
        await self.db.delete(rc)
        await self.db.flush()

    async def _get_rc_editable(self, rc_id: int) -> MetaResultadoClave:
        rc = await self.repo.get_rc(rc_id)
        if not rc:
            raise NotFoundError("MetaResultadoClave", rc_id)
        if rc.meta.ciclo.estado != "activo":
            raise ConflictError(
                "Solo se pueden modificar resultados clave de un ciclo activo"
            )
        if rc.meta.estado == "cerrada":
            raise ConflictError(
                "No se pueden modificar resultados clave de una meta cerrada"
            )
        return rc

    async def get_rc_meta(self, rc_id: int) -> MetaResponse:
        """Resuelve (ya serializada) la `Meta` dueña de un resultado clave a
        partir de su `rc_id` suelto — usado por el router (Tarea 3) para el
        scoping de equipo/ownership antes de un check-in (ajuste del jefe o
        self-service), casos en los que el cliente solo envia `rc_id`, no
        `meta_id`. Reemplaza el uso directo de `MetasRepository.get_rc` que
        hacia el router (ver concern de Tarea 3 / fix post-revision)."""
        rc = await self.repo.get_rc(rc_id)
        if not rc:
            raise NotFoundError("Resultado clave", rc_id)
        return await self.get_meta(rc.meta_id)

    # ── Check-in ───────────────────────────────────────────────────────────
    async def registrar_checkin(
        self,
        rc_id: int,
        autor_id: int,
        valor: Numero,
        nota: Optional[str] = None,
        es_ajuste_jefe: bool = False,
    ) -> CheckinResponse:
        """Registra un check-in inmutable y actualiza `valor_actual` del RC.

        Primer check-in: si la meta esta "asignada" pasa a "en_progreso"
        (solo se evalua el estado actual, no un conteo de check-ins previos:
        una meta solo esta en "asignada" antes de su primer check-in por
        diseno del ciclo de vida, ver spec)."""
        rc = await self._get_rc_editable(rc_id)
        meta = rc.meta

        checkin = MetaCheckin(
            autor_id=autor_id,
            valor_registrado=Decimal(str(valor)),
            nota=nota,
            es_ajuste_jefe=es_ajuste_jefe,
        )
        # Se agrega via la relacion (no seteando resultado_clave_id a mano):
        # asi la coleccion `rc.checkins` -ya cargada en memoria por el
        # selectinload de `_get_rc_editable`- queda sincronizada. Si solo se
        # asignara la FK, `rc.checkins` seguiria "vacia" en memoria para el
        # resto de la sesion (p. ej. el guard de `eliminar_meta`/`eliminar_rc`
        # que verifica `rc.checkins` no veria este check-in).
        rc.checkins.append(checkin)
        rc.valor_actual = Decimal(str(valor))
        if meta.estado == "asignada":
            meta.estado = "en_progreso"

        await self.db.flush()
        await self.db.refresh(checkin)
        return CheckinResponse(
            id=checkin.id,
            resultado_clave_id=checkin.resultado_clave_id,
            autor_id=checkin.autor_id,
            valor_registrado=checkin.valor_registrado,
            nota=checkin.nota,
            es_ajuste_jefe=checkin.es_ajuste_jefe,
            created_at=checkin.created_at,
            avance_resultante=self.avance_rc(rc),
        )

    # ══════════════════════════════════════════════════════════════════════
    # Recordatorios (Tarea 5)
    # ══════════════════════════════════════════════════════════════════════
    async def _notificar_recordatorio_meta(
        self,
        meta: Meta,
        ciclo: MetaCiclo,
        *,
        proximo_a_cerrar: bool,
        estancada: bool,
    ) -> None:
        motivos = []
        if proximo_a_cerrar:
            motivos.append(
                f"el ciclo '{ciclo.nombre}' cierra el {ciclo.fecha_fin.isoformat()}"
            )
        if estancada:
            motivos.append(
                f"la meta '{meta.titulo}' no tiene check-ins recientes"
            )
        cuerpo = "Recordatorio de metas: " + "; ".join(motivos) + "."
        await self.notificaciones.enviar(
            destinatario_id=meta.empleado_id,
            asunto="Recordatorio de metas pendientes",
            cuerpo=cuerpo,
            canal="in_app",
            target_url=MIS_METAS_TARGET_URL,
            metadata={"ciclo_id": ciclo.id, "meta_id": meta.id},
        )
        meta.ultimo_recordatorio_at = datetime.now(timezone.utc)

    async def procesar_recordatorios(
        self,
        dias_cierre: int = DIAS_CIERRE_PROXIMO_DEFAULT,
        dias_sin_checkin: int = DIAS_SIN_CHECKIN_DEFAULT,
    ) -> RecordatoriosResultado:
        """Job diario (`app.main._metas_recordatorios_job`): notifica a los
        empleados con metas INDIVIDUALES no cerradas de un ciclo ACTIVO
        cuando aplica al menos uno de estos motivos:

          - El ciclo esta proximo a cerrar: `0 <= (fecha_fin - hoy) <=
            dias_cierre`.
          - Alguno de los resultados clave de la meta lleva `>= dias_sin_checkin`
            dias sin un check-in nuevo (o sin ninguno, desde que se asigno la
            meta) — ver `_rc_dias_sin_checkin`.

        Metas cerradas (calificadas) o de ciclos no activos (borrador/cerrado)
        nunca se consideran (se apoya en
        `MetasRepository.list_metas_individuales_no_cerradas`, que ya filtra
        por ciclo + nivel individual + no cerrada).

        Dedupe temporal (fix post-revision, ver `RECORDATORIO_CADENCIA_DIAS`):
        una meta con `ultimo_recordatorio_at` dentro de los ultimos
        `RECORDATORIO_CADENCIA_DIAS` dias se EXCLUYE aunque su condicion siga
        cumpliendose (evita spam diario mientras, p. ej., el RC sigue sin
        check-in); `ultimo_recordatorio_at is None` (nunca notificada) si
        entra. Mismo patron que `EncuestasRhService.procesar_recordatorios`
        con `EncuestaParticipante.ultimo_recordatorio_at`.

        Devuelve `{notificados, ciclos_por_cerrar}`: `notificados` cuenta
        EMPLEADOS distintos notificados en esta corrida (no notificaciones
        individuales); `ciclos_por_cerrar` cuenta ciclos activos dentro de la
        ventana de `dias_cierre`, independientemente de si tenian metas
        pendientes que notificar."""
        hoy = date.today()
        ahora = datetime.now(timezone.utc)
        ciclos_por_cerrar = 0
        notificados: set[int] = set()

        for ciclo in await self.repo.list_ciclos(estado="activo"):
            dias_restantes = (ciclo.fecha_fin - hoy).days
            proximo_a_cerrar = 0 <= dias_restantes <= dias_cierre
            if proximo_a_cerrar:
                ciclos_por_cerrar += 1

            for meta in await self.repo.list_metas_individuales_no_cerradas(ciclo.id):
                if meta.empleado_id is None:
                    continue
                ultimo = _dt_utc(meta.ultimo_recordatorio_at)
                if ultimo is not None and (ahora - ultimo) < timedelta(
                    days=RECORDATORIO_CADENCIA_DIAS
                ):
                    continue
                estancada = any(
                    _rc_dias_sin_checkin(rc, meta.created_at, ahora) >= dias_sin_checkin
                    for rc in meta.resultados_clave
                )
                if not proximo_a_cerrar and not estancada:
                    continue
                await self._notificar_recordatorio_meta(
                    meta, ciclo, proximo_a_cerrar=proximo_a_cerrar, estancada=estancada
                )
                notificados.add(meta.empleado_id)

        await self.db.flush()
        return RecordatoriosResultado(
            notificados=len(notificados), ciclos_por_cerrar=ciclos_por_cerrar
        )

    async def forzar_recordatorios_ciclo(self, ciclo_id: int) -> int:
        """Endpoint manual de gestion (`POST /ciclos/{id}/recordatorios`):
        fuerza un recordatorio a TODOS los empleados con metas individuales
        pendientes (no cerradas) del ciclo, sin evaluar `dias_cierre`/
        `dias_sin_checkin` NI la cadencia `RECORDATORIO_CADENCIA_DIAS` (a
        diferencia de `procesar_recordatorios`) — un jefe/RH que fuerza el
        envio manualmente quiere notificar YA, sin importar cuando fue el
        ultimo automatico. Si actualiza `ultimo_recordatorio_at` (vía
        `_notificar_recordatorio_meta`), asi el job automatico siguiente
        respeta la cadencia desde este envio forzado.
        Devuelve el numero de empleados distintos notificados."""
        ciclo = await self.repo.get_ciclo(ciclo_id)
        if not ciclo:
            raise NotFoundError("MetaCiclo", ciclo_id)

        notificados: set[int] = set()
        for meta in await self.repo.list_metas_individuales_no_cerradas(ciclo_id):
            if meta.empleado_id is None:
                continue
            await self._notificar_recordatorio_meta(
                meta, ciclo, proximo_a_cerrar=True, estancada=False
            )
            notificados.add(meta.empleado_id)

        await self.db.flush()
        return len(notificados)

    # ══════════════════════════════════════════════════════════════════════
    # Serializacion
    # ══════════════════════════════════════════════════════════════════════
    def _ciclo_to_response(self, ciclo: MetaCiclo) -> MetaCicloResponse:
        return MetaCicloResponse.model_validate(ciclo)

    def _rc_to_response(self, rc: MetaResultadoClave) -> ResultadoClaveResponse:
        data = ResultadoClaveResponse.model_validate(rc)
        data.avance = self.avance_rc(rc)
        return data

    def _meta_to_response(self, meta: Meta) -> MetaResponse:
        data = MetaResponse.model_validate(meta)
        data.avance = self.avance_meta(meta)
        data.resultados_clave = [self._rc_to_response(rc) for rc in meta.resultados_clave]
        return data
