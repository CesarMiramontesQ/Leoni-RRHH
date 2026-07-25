"""Analitica de cobertura y polivalencia por area (solo lectura).

Reutiliza `CompetenciaService.obtener_multihabilidades` (grilla empleado x
competencia por puesto, con manejo de grado) y agrega por area. Scope por rol
via `empleado_ids_scope_por_modulo`. Sin tabla nueva.
"""
from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from sqlalchemy import exists, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.data_scope import empleado_ids_scope_por_modulo
from app.core.exceptions import ForbiddenError, NotFoundError
from app.models.catalogos import Area
from app.models.talento import CompetenciaRequisito, PuestoPerfil
from app.repositories.empleado_repository import EmpleadoRepository
from app.services.competencia_service import CompetenciaService
from app.services.operaciones import calculo
from app.services.operaciones.types import (
    CandidatoCrossTrain,
    CoberturaCompetencia,
    CompetenciaMeta,
    EmpleadoCompetencias,
)

MODULE_KEY = "operaciones"


@dataclass
class PuestoArea:
    puesto_perfil_id: int
    puesto_nombre: str
    area_id: int
    area_nombre: str


@dataclass
class AreaResumen:
    area_id: int
    area_nombre: str
    pol_area_pct: float
    resiliencia_pct: float
    n_criticas: int
    n_empleados: int


@dataclass
class PolivalenciaEmpleado:
    """Indice de polivalencia de UN empleado dentro de un area.

    Existe para el Dashboard de Talento: `indice_polivalencia_empleado` ya
    calculaba esto, pero hasta ahora solo se consumia promediado por area.
    """

    empleado_id: int
    no_empleado: int | str
    nombre: str
    puesto_nombre: str
    pol_pct: float | None


@dataclass
class PuestoCobertura:
    puesto_perfil_id: int
    puesto_nombre: str
    competencias: list[CoberturaCompetencia]


@dataclass
class Critica:
    competencia_id: int
    competencia_nombre: str
    severidad: str
    candidatos: list[CandidatoCrossTrain]


@dataclass
class CoberturaArea:
    resumen: AreaResumen
    competencias: list[CoberturaCompetencia]
    puestos: list[PuestoCobertura]
    criticas: list[Critica]


class OperacionesService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.comp_svc = CompetenciaService(db)
        self.empleado_repo = EmpleadoRepository(db)

    async def _puestos_con_area(self) -> list[PuestoArea]:
        """Puestos activos con >= 1 competencia requisito, con su area."""
        result = await self.db.execute(
            select(PuestoPerfil.id, PuestoPerfil.nombre, PuestoPerfil.area_id, Area.descripcion)
            .join(Area, Area.area_id == PuestoPerfil.area_id)
            .where(
                PuestoPerfil.activo.is_(True),
                exists().where(CompetenciaRequisito.puesto_perfil_id == PuestoPerfil.id),
            )
        )
        return [
            PuestoArea(puesto_perfil_id=r[0], puesto_nombre=r[1], area_id=r[2], area_nombre=r[3])
            for r in result.all()
        ]

    async def _empleados_de_puesto(
        self, puesto_perfil_id: int, puesto_nombre: str
    ) -> tuple[list[EmpleadoCompetencias], dict[int, CompetenciaMeta]]:
        resp = await self.comp_svc.obtener_multihabilidades(puesto_perfil_id)
        comp_meta = {
            c.competencia_id: CompetenciaMeta(c.competencia_id, c.competencia_nombre, c.tipo_nombre)
            for c in resp.competencias
        }
        empleados: list[EmpleadoCompetencias] = []
        for emp in resp.empleados:
            comps: dict[int, tuple[int, int]] = {}
            for comp_id, requerido in emp.requisitos.items():
                if requerido < 1:
                    continue
                comps[comp_id] = (emp.niveles.get(comp_id, 0), requerido)
            empleados.append(
                EmpleadoCompetencias(
                    empleado_id=emp.empleado_id,
                    no_empleado=emp.no_empleado,
                    nombre=emp.nombre,
                    puesto_perfil_id=resp.puesto_perfil_id,
                    puesto_nombre=puesto_nombre,
                    competencias=comps,
                )
            )
        return empleados, comp_meta

    async def _cargar_area(
        self, area_id: int, scope: list[int] | None
    ) -> tuple[str, list[PuestoCobertura], list[EmpleadoCompetencias], dict[int, CompetenciaMeta]]:
        """Devuelve (area_nombre, puestos_cobertura, empleados_scope, comp_meta)."""
        puestos = [p for p in await self._puestos_con_area() if p.area_id == area_id]
        area_nombre = puestos[0].area_nombre if puestos else ""
        puestos_cob: list[PuestoCobertura] = []
        empleados_area: list[EmpleadoCompetencias] = []
        comp_meta: dict[int, CompetenciaMeta] = {}
        for p in puestos:
            emps, meta = await self._empleados_de_puesto(p.puesto_perfil_id, p.puesto_nombre)
            comp_meta.update(meta)
            if scope is not None:
                emps = [e for e in emps if e.empleado_id in scope]
            empleados_area.extend(emps)
            puestos_cob.append(
                PuestoCobertura(
                    puesto_perfil_id=p.puesto_perfil_id,
                    puesto_nombre=p.puesto_nombre,
                    competencias=calculo.cobertura_por_competencia(emps, meta),
                )
            )
        return area_nombre, puestos_cob, empleados_area, comp_meta

    def _resumen(self, area_id, area_nombre, empleados, coberturas) -> AreaResumen:
        n_criticas = sum(1 for c in coberturas if c.severidad != "ok")
        return AreaResumen(
            area_id=area_id,
            area_nombre=area_nombre,
            pol_area_pct=calculo.indice_polivalencia_area(empleados),
            resiliencia_pct=calculo.resiliencia_area(coberturas),
            n_criticas=n_criticas,
            n_empleados=len({e.empleado_id for e in empleados}),
        )

    async def listar_areas(self, current_user, rh_ui_mode) -> list[AreaResumen]:
        scope = await empleado_ids_scope_por_modulo(
            self.empleado_repo, current_user, MODULE_KEY, rh_ui_mode
        )
        return await self.listar_areas_con_scope(scope)

    async def listar_areas_con_scope(self, scope: list[int] | None) -> list[AreaResumen]:
        """Resumen por area con el scope YA resuelto (`None` = universo).

        Separado de `listar_areas` para que el Dashboard de Talento pueda pasar
        el scope que el resolvio con SU module_key, en vez de recalcularlo con
        el de Operaciones -- si cada bloque resolviera el suyo, dos columnas de
        la misma fila saldrian sobre poblaciones distintas."""
        puestos = await self._puestos_con_area()
        # Agrupa puestos por area en una sola pasada.
        por_area: dict[int, tuple[str, list[EmpleadoCompetencias], dict[int, CompetenciaMeta]]] = {}
        for p in puestos:
            emps, meta = await self._empleados_de_puesto(p.puesto_perfil_id, p.puesto_nombre)
            if scope is not None:
                emps = [e for e in emps if e.empleado_id in scope]
            nombre, acc_emps, acc_meta = por_area.setdefault(p.area_id, (p.area_nombre, [], {}))
            acc_emps.extend(emps)
            acc_meta.update(meta)
        resumenes: list[AreaResumen] = []
        for area_id, (nombre, emps, meta) in por_area.items():
            if not emps:
                continue  # area sin personal en scope
            coberturas = calculo.cobertura_por_competencia(emps, meta)
            resumenes.append(self._resumen(area_id, nombre, emps, coberturas))
        resumenes.sort(key=lambda a: (-a.n_criticas, a.area_nombre))
        return resumenes

    async def polivalencia_empleados_area(
        self, area_id: int, scope: list[int] | None
    ) -> list[PolivalenciaEmpleado]:
        """Indice de polivalencia por empleado del area (scope ya resuelto).

        Dedup por empleado_id: si esta asignado a varios puestos del area, se
        queda el indice mas alto -- mismo criterio de `candidatos_crosstrain`.

        Visibilidad: mismo criterio que `cobertura_area` (mismas comprobaciones,
        mismo orden, mismos mensajes) -- un area inexistente da `NotFoundError`
        y un area fuera del scope da `ForbiddenError`, en vez de devolver un
        resultado vacio con status 200."""
        _nombre, puestos_cob, empleados, _meta = await self._cargar_area(area_id, scope)
        if not puestos_cob:
            # Area inexistente o sin puestos con competencias requisito: nada que mostrar.
            raise NotFoundError(entidad="Area", id=area_id)
        if not empleados:
            # Existe pero no hay personal visible: fuera de scope (jefe) o sin datos (RH).
            if scope is not None:
                raise ForbiddenError(detail="Area fuera de tu alcance")
            raise NotFoundError(entidad="Area", id=area_id)
        por_empleado: dict[int, PolivalenciaEmpleado] = {}
        for e in empleados:
            pct = calculo.indice_polivalencia_empleado(e)
            prev = por_empleado.get(e.empleado_id)
            if prev is not None and (prev.pol_pct or 0.0) >= (pct or 0.0):
                continue
            por_empleado[e.empleado_id] = PolivalenciaEmpleado(
                empleado_id=e.empleado_id,
                no_empleado=e.no_empleado,
                nombre=e.nombre,
                puesto_nombre=e.puesto_nombre,
                pol_pct=pct,
            )
        return sorted(por_empleado.values(), key=lambda p: p.nombre)

    async def cobertura_area(self, current_user, area_id: int, rh_ui_mode) -> CoberturaArea:
        scope = await empleado_ids_scope_por_modulo(
            self.empleado_repo, current_user, MODULE_KEY, rh_ui_mode
        )
        area_nombre, puestos_cob, empleados, comp_meta = await self._cargar_area(area_id, scope)
        if not puestos_cob:
            # Area inexistente o sin puestos con competencias requisito: nada que mostrar.
            raise NotFoundError(entidad="Area", id=area_id)
        if not empleados:
            # Existe pero no hay personal visible: fuera de scope (jefe) o sin datos (RH).
            if scope is not None:
                raise ForbiddenError(detail="Area fuera de tu alcance")
            raise NotFoundError(entidad="Area", id=area_id)
        coberturas = calculo.cobertura_por_competencia(empleados, comp_meta)
        criticas = [
            Critica(
                competencia_id=c.competencia_id,
                competencia_nombre=c.competencia_nombre,
                severidad=c.severidad,
                candidatos=calculo.candidatos_crosstrain(c.competencia_id, empleados),
            )
            for c in coberturas
            if c.severidad != "ok"
        ]
        return CoberturaArea(
            resumen=self._resumen(area_id, area_nombre, empleados, coberturas),
            competencias=coberturas,
            puestos=puestos_cob,
            criticas=criticas,
        )

    async def exportar_area_excel(self, current_user, area_id: int, rh_ui_mode) -> BytesIO:
        """xlsx con 3 hojas: Resumen, Cobertura por competencia, Cross-training."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        cob = await self.cobertura_area(current_user, area_id, rh_ui_mode)
        r = cob.resumen
        wb = Workbook()

        ws = wb.active
        ws.title = "Resumen"
        ws.cell(row=1, column=1, value=f"Cobertura — {r.area_nombre}").font = Font(bold=True, size=14)
        filas = [
            ("Indice de polivalencia (%)", r.pol_area_pct),
            ("Resiliencia (% sin punto unico) ", r.resiliencia_pct),
            ("Competencias criticas", r.n_criticas),
            ("Empleados", r.n_empleados),
        ]
        for i, (etq, val) in enumerate(filas, start=3):
            ws.cell(row=i, column=1, value=etq).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)

        ws2 = wb.create_sheet("Cobertura por competencia")
        headers = ["Competencia", "Tipo", "Requieren", "Cubren", "En entrenamiento", "Cobertura %", "Semaforo", "Severidad"]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row, c in enumerate(cob.competencias, start=2):
            for col, val in enumerate(
                [c.competencia_nombre, c.tipo_nombre, c.requieren, c.cubren,
                 c.en_entrenamiento, c.cobertura_pct, c.semaforo, c.severidad], 1
            ):
                ws2.cell(row=row, column=col, value=val)

        ws3 = wb.create_sheet("Cross-training")
        h3 = ["Competencia", "Severidad", "Candidato", "No. empleado", "Nivel actual", "Nivel requerido"]
        for col, h in enumerate(h3, 1):
            ws3.cell(row=1, column=col, value=h).font = Font(bold=True)
        row = 2
        for crit in cob.criticas:
            for cand in crit.candidatos:
                for col, val in enumerate(
                    [crit.competencia_nombre, crit.severidad, cand.nombre,
                     cand.no_empleado, cand.nivel_actual, cand.nivel_requerido], 1
                ):
                    ws3.cell(row=row, column=col, value=val)
                row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
