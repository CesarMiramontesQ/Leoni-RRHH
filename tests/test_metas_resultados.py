# tests/test_metas_resultados.py
"""Tests HTTP del tablero de equipo, cumplimiento y export Excel (Tarea 4).

La logica de dominio (formulas de avance/cumplimiento, roll-up de meta de
equipo) ya esta cubierta a nivel service en tests/test_metas_service.py
(Tarea 2) y el scoping de equipo generico en tests/test_metas_api.py
(Tarea 3); aqui se ejercita especificamente:
  - `GET /equipo/avance`: agrupado por miembro (metas + avance global
    ponderado) + metas de equipo (con roll-up) aparte, con scoping.
  - `GET /empleados/{id}/cumplimiento`: 0 antes del cierre, ponderado despues.
  - `GET /ciclos/{id}/export/excel`: workbook valido (abierto con openpyxl),
    con scoping de equipo tambien aplicado al export.
"""

from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

from tests.conftest import auth_headers, make_empleado

pytestmark = pytest.mark.asyncio

BASE = "/api/v1/metas"


def _fechas():
    hoy = date.today()
    return hoy.isoformat(), (hoy + timedelta(days=90)).isoformat()


async def _rh(db, **kw):
    return await make_empleado(db, rol="rh", modulos_rh={"metas": True}, **kw)


async def _jefe(db, **kw):
    return await make_empleado(db, rol="supervisor", **kw)


async def _empleado_de(db, jefe, **kw):
    return await make_empleado(db, rol="empleado", lider_id=jefe.empleado_id, **kw)


async def _crear_ciclo_activo(client, headers_rh, nombre="Ciclo resultados"):
    inicio, fin = _fechas()
    resp = await client.post(
        f"{BASE}/ciclos",
        json={"nombre": nombre, "fecha_inicio": inicio, "fecha_fin": fin},
        headers=headers_rh,
    )
    assert resp.status_code == 201, resp.text
    ciclo = resp.json()
    resp_act = await client.post(f"{BASE}/ciclos/{ciclo['id']}/activar", headers=headers_rh)
    assert resp_act.status_code == 200, resp_act.text
    return resp_act.json()


async def _crear_meta_individual(
    client, headers, ciclo_id, empleado_id, asignada_por_id, titulo, peso, rcs=None
):
    resp = await client.post(
        f"{BASE}/metas",
        json={
            "ciclo_id": ciclo_id,
            "nivel": "individual",
            "empleado_id": empleado_id,
            "titulo": titulo,
            "peso": peso,
            "asignada_por_id": asignada_por_id,
            "resultados_clave": rcs or [],
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


# ══════════════════════════════════════════════════════════════════════════
# GET /equipo/avance — agrupado por miembro + avance global ponderado
# ══════════════════════════════════════════════════════════════════════════
async def test_equipo_avance_agrupa_por_miembro_con_avance_global_ponderado(client, db):
    rh = await _rh(db, email="metasresrh1@leoni.test")
    jefe = await _jefe(db, email="metasresjefe1@leoni.test")
    emp = await _empleado_de(db, jefe, email="metasresemp1@leoni.test", nombre="Ana Torres")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo tablero 1")

    # Dos metas del mismo empleado con distinto peso: avance global debe ser
    # el promedio PONDERADO del avance de cada una (no el simple).
    meta1 = await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp.empleado_id, jefe.empleado_id,
        "Meta A", peso=75,
        rcs=[{
            "orden": 1, "titulo": "RC A", "tipo_metrica": "numero",
            "direccion": "subir", "valor_inicial": 0, "valor_objetivo": 100,
        }],
    )
    meta2 = await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp.empleado_id, jefe.empleado_id,
        "Meta B", peso=25,
        rcs=[{
            "orden": 1, "titulo": "RC B", "tipo_metrica": "numero",
            "direccion": "subir", "valor_inicial": 0, "valor_objetivo": 100,
        }],
    )
    rc1_id = meta1["resultados_clave"][0]["id"]
    rc2_id = meta2["resultados_clave"][0]["id"]

    # Meta A al 80% de avance, Meta B al 0% (sin check-in).
    resp_checkin = await client.post(
        f"{BASE}/resultados/{rc1_id}/checkin",
        json={"valor": 80, "nota": "ajuste jefe"},
        headers=headers_jefe,
    )
    assert resp_checkin.status_code == 201, resp_checkin.text
    assert resp_checkin.json()["avance_resultante"] == 80.0

    resp_tablero = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_jefe
    )
    assert resp_tablero.status_code == 200, resp_tablero.text
    tablero = resp_tablero.json()

    assert tablero["ciclo_id"] == ciclo["id"]
    assert tablero["metas_equipo"] == []
    assert len(tablero["miembros"]) == 1
    miembro = tablero["miembros"][0]
    assert miembro["empleado_id"] == emp.empleado_id
    assert miembro["empleado_nombre"] == "Ana Torres"
    ids_metas = {m["id"] for m in miembro["metas"]}
    assert ids_metas == {meta1["id"], meta2["id"]}
    # Ponderado: (75*80 + 25*0) / 100 = 60.0 (no el simple (80+0)/2 = 40).
    assert miembro["avance_global"] == 60.0
    assert rc2_id  # sanity: el segundo RC existe y sigue en 0.


async def test_equipo_avance_metas_equipo_aparte_con_rollup_de_submetas(client, db):
    rh = await _rh(db, email="metasresrh2@leoni.test")
    jefe = await _jefe(db, email="metasresjefe2@leoni.test")
    emp1 = await _empleado_de(db, jefe, email="metasresemp2a@leoni.test")
    emp2 = await _empleado_de(db, jefe, email="metasresemp2b@leoni.test")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo tablero rollup")

    # Meta de equipo (sin RC propios) del jefe.
    resp_meta_equipo = await client.post(
        f"{BASE}/metas",
        json={
            "ciclo_id": ciclo["id"],
            "nivel": "equipo",
            "area_id": 1,
            "lider_id": jefe.empleado_id,
            "titulo": "Meta de equipo",
            "peso": 100,
            "asignada_por_id": jefe.empleado_id,
            "resultados_clave": [],
        },
        headers=headers_jefe,
    )
    assert resp_meta_equipo.status_code == 201, resp_meta_equipo.text
    meta_equipo_id = resp_meta_equipo.json()["id"]

    # Dos submetas individuales enlazadas via meta_padre_id.
    for emp, rc_objetivo, valor_actual in ((emp1, 100, 100), (emp2, 100, 0)):
        resp_sub = await client.post(
            f"{BASE}/metas",
            json={
                "ciclo_id": ciclo["id"],
                "nivel": "individual",
                "empleado_id": emp.empleado_id,
                "meta_padre_id": meta_equipo_id,
                "titulo": f"Submeta {emp.empleado_id}",
                "peso": 100,
                "asignada_por_id": jefe.empleado_id,
                "resultados_clave": [
                    {
                        "orden": 1, "titulo": "RC", "tipo_metrica": "numero",
                        "direccion": "subir", "valor_inicial": 0,
                        "valor_objetivo": rc_objetivo,
                    }
                ],
            },
            headers=headers_jefe,
        )
        assert resp_sub.status_code == 201, resp_sub.text
        rc_id = resp_sub.json()["resultados_clave"][0]["id"]
        if valor_actual:
            resp_ci = await client.post(
                f"{BASE}/resultados/{rc_id}/checkin",
                json={"valor": valor_actual},
                headers=headers_jefe,
            )
            assert resp_ci.status_code == 201, resp_ci.text

    resp_tablero = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_jefe
    )
    assert resp_tablero.status_code == 200, resp_tablero.text
    tablero = resp_tablero.json()

    # La meta de equipo va en `metas_equipo`, no en `miembros`.
    assert len(tablero["metas_equipo"]) == 1
    meta_equipo = tablero["metas_equipo"][0]
    assert meta_equipo["id"] == meta_equipo_id
    # Roll-up: promedio del avance de sus 2 submetas (100 y 0) = 50.0.
    assert meta_equipo["avance"] == 50.0
    # Las 2 submetas individuales SI aparecen agrupadas por miembro.
    assert len(tablero["miembros"]) == 2


async def test_equipo_avance_scoping_jefe_solo_ve_su_equipo(client, db):
    rh = await _rh(db, email="metasresrh3@leoni.test")
    jefe = await _jefe(db, email="metasresjefe3@leoni.test")
    otro_jefe = await _jefe(db, email="metasresotrojefe3@leoni.test")
    emp = await _empleado_de(db, jefe, email="metasresemp3@leoni.test")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)
    headers_otro_jefe = await auth_headers(client, otro_jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo tablero scope")

    await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp.empleado_id, jefe.empleado_id,
        "Meta del equipo de jefe", peso=100,
    )

    # El jefe dueño ve a su unico miembro.
    resp_jefe = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_jefe
    )
    assert resp_jefe.status_code == 200, resp_jefe.text
    assert len(resp_jefe.json()["miembros"]) == 1

    # Otro jefe (sin ese reporte directo) no ve nada de ese equipo.
    resp_otro = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_otro_jefe
    )
    assert resp_otro.status_code == 200, resp_otro.text
    assert resp_otro.json()["miembros"] == []

    # RH con modulo ve el ciclo completo (incluye al menos ese miembro).
    resp_rh = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_rh
    )
    assert resp_rh.status_code == 200, resp_rh.text
    ids_rh = {m["empleado_id"] for m in resp_rh.json()["miembros"]}
    assert emp.empleado_id in ids_rh


async def test_equipo_avance_rh_global_devuelve_todos_los_equipos_y_metas(client, db):
    """Hueco senalado en revision: T4 no probo explicitamente que RH (sin
    scope de equipo) vea TODOS los equipos/metas del ciclo, no solo uno.
    Se arman 2 jefes con equipos distintos + una meta de equipo de cada uno
    y se verifica que RH ve a los miembros de ambos y las 2 metas de equipo,
    mientras que cada jefe solo ve lo propio."""
    rh = await _rh(db, email="metasresrhglobal@leoni.test")
    jefe_a = await _jefe(db, email="metasresjefeglobala@leoni.test")
    jefe_b = await _jefe(db, email="metasresjefeglobalb@leoni.test")
    emp_a = await _empleado_de(db, jefe_a, email="metasresempglobala@leoni.test")
    emp_b = await _empleado_de(db, jefe_b, email="metasresempglobalb@leoni.test")

    headers_rh = await auth_headers(client, rh)
    headers_jefe_a = await auth_headers(client, jefe_a)
    headers_jefe_b = await auth_headers(client, jefe_b)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo tablero RH global")

    await _crear_meta_individual(
        client, headers_jefe_a, ciclo["id"], emp_a.empleado_id, jefe_a.empleado_id,
        "Meta individual equipo A", peso=100,
    )
    await _crear_meta_individual(
        client, headers_jefe_b, ciclo["id"], emp_b.empleado_id, jefe_b.empleado_id,
        "Meta individual equipo B", peso=100,
    )

    for headers, jefe in ((headers_jefe_a, jefe_a), (headers_jefe_b, jefe_b)):
        resp_meta_equipo = await client.post(
            f"{BASE}/metas",
            json={
                "ciclo_id": ciclo["id"],
                "nivel": "equipo",
                "area_id": jefe.empleado_id,
                "lider_id": jefe.empleado_id,
                "titulo": f"Meta de equipo de {jefe.empleado_id}",
                "peso": 100,
                "asignada_por_id": jefe.empleado_id,
                "resultados_clave": [],
            },
            headers=headers,
        )
        assert resp_meta_equipo.status_code == 201, resp_meta_equipo.text

    # RH (sin scope) ve a los miembros de AMBOS equipos y las 2 metas de equipo.
    resp_rh = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_rh
    )
    assert resp_rh.status_code == 200, resp_rh.text
    tablero_rh = resp_rh.json()
    ids_miembros_rh = {m["empleado_id"] for m in tablero_rh["miembros"]}
    assert {emp_a.empleado_id, emp_b.empleado_id} <= ids_miembros_rh
    assert len(tablero_rh["metas_equipo"]) == 2
    lideres_rh = {m["lider_id"] for m in tablero_rh["metas_equipo"]}
    assert lideres_rh == {jefe_a.empleado_id, jefe_b.empleado_id}

    # Cada jefe, en cambio, solo ve su propio equipo (scoping intacto).
    resp_jefe_a = await client.get(
        f"{BASE}/equipo/avance", params={"ciclo_id": ciclo["id"]}, headers=headers_jefe_a
    )
    assert resp_jefe_a.status_code == 200, resp_jefe_a.text
    ids_miembros_a = {m["empleado_id"] for m in resp_jefe_a.json()["miembros"]}
    assert ids_miembros_a == {emp_a.empleado_id}
    assert len(resp_jefe_a.json()["metas_equipo"]) == 1


# ══════════════════════════════════════════════════════════════════════════
# GET /empleados/{id}/cumplimiento — solo tras cierre
# ══════════════════════════════════════════════════════════════════════════
async def test_cumplimiento_es_cero_antes_del_cierre_y_ponderado_despues(client, db):
    rh = await _rh(db, email="metasrescumplrh@leoni.test")
    jefe = await _jefe(db, email="metasrescumpljefe@leoni.test")
    emp = await _empleado_de(db, jefe, email="metasrescumplemp@leoni.test")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo cumplimiento")

    meta = await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp.empleado_id, jefe.empleado_id,
        "Meta a calificar", peso=100,
    )

    # Antes de cerrar la meta: cumplimiento 0 (borde documentado, sin metas
    # cerradas todavia — la meta esta "asignada", no "cerrada").
    resp_antes = await client.get(
        f"{BASE}/empleados/{emp.empleado_id}/cumplimiento",
        params={"ciclo_id": ciclo["id"]},
        headers=headers_jefe,
    )
    assert resp_antes.status_code == 200, resp_antes.text
    assert resp_antes.json()["cumplimiento"] == 0.0
    assert resp_antes.json()["metas_consideradas"] == 0

    # Jefe cierra y califica la meta.
    resp_cerrar = await client.post(
        f"{BASE}/metas/{meta['id']}/cerrar",
        json={"calificacion": 85},
        headers=headers_jefe,
    )
    assert resp_cerrar.status_code == 200, resp_cerrar.text

    resp_despues = await client.get(
        f"{BASE}/empleados/{emp.empleado_id}/cumplimiento",
        params={"ciclo_id": ciclo["id"]},
        headers=headers_jefe,
    )
    assert resp_despues.status_code == 200, resp_despues.text
    assert resp_despues.json()["cumplimiento"] == 85.0
    assert resp_despues.json()["metas_consideradas"] == 1


# ══════════════════════════════════════════════════════════════════════════
# GET /ciclos/{id}/export/excel — workbook valido + scoping
# ══════════════════════════════════════════════════════════════════════════
async def test_export_excel_devuelve_workbook_valido_con_metas_avance_cumplimiento(client, db):
    rh = await _rh(db, email="metasresexprh@leoni.test")
    jefe = await _jefe(db, email="metasresexpjefe@leoni.test")
    emp = await _empleado_de(db, jefe, email="metasresexpemp@leoni.test", nombre="Beto Ruiz")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo export")

    meta = await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp.empleado_id, jefe.empleado_id,
        "Meta exportable", peso=100,
    )
    resp_cerrar = await client.post(
        f"{BASE}/metas/{meta['id']}/cerrar",
        json={"calificacion": 70},
        headers=headers_jefe,
    )
    assert resp_cerrar.status_code == 200, resp_cerrar.text

    resp_export = await client.get(
        f"{BASE}/ciclos/{ciclo['id']}/export/excel", headers=headers_rh
    )
    assert resp_export.status_code == 200, resp_export.text
    assert "spreadsheetml" in resp_export.headers["content-type"]
    assert "metas_ciclo_" in resp_export.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp_export.content))
    ws = wb["Metas y avance"]
    headers_fila = [ws.cell(row=3, column=c).value for c in range(1, 9)]
    assert headers_fila == [
        "Empleado", "Meta", "Nivel", "Estado", "Peso",
        "Avance %", "Calificación cierre", "Cumplimiento ponderado",
    ]
    fila = [ws.cell(row=4, column=c).value for c in range(1, 9)]
    assert fila[0] == "Beto Ruiz"
    assert fila[1] == "Meta exportable"
    assert fila[2] == "individual"
    assert fila[3] == "cerrada"
    assert fila[4] == 100.0
    # Columna F "Avance %": meta sin RC (avance 0.0), saltada en el assert
    # original (hueco senalado en revision).
    assert fila[5] == 0.0
    assert fila[6] == 70.0
    assert fila[7] == 70.0


async def test_export_excel_ciclo_inexistente_404(client, db):
    rh = await _rh(db, email="metasresexp404@leoni.test")
    headers_rh = await auth_headers(client, rh)

    resp = await client.get(f"{BASE}/ciclos/999999/export/excel", headers=headers_rh)
    assert resp.status_code == 404, resp.text


async def test_export_excel_scoping_jefe_solo_incluye_su_equipo(client, db):
    rh = await _rh(db, email="metasresexpscoperh@leoni.test")
    jefe = await _jefe(db, email="metasresexpscopejefe@leoni.test")
    otro_jefe = await _jefe(db, email="metasresexpscopeotrojefe@leoni.test")
    emp_jefe = await _empleado_de(db, jefe, email="metasresexpscopeemp1@leoni.test")
    emp_otro = await _empleado_de(db, otro_jefe, email="metasresexpscopeemp2@leoni.test")

    headers_rh = await auth_headers(client, rh)
    headers_jefe = await auth_headers(client, jefe)
    headers_otro_jefe = await auth_headers(client, otro_jefe)

    ciclo = await _crear_ciclo_activo(client, headers_rh, nombre="Ciclo export scope")

    await _crear_meta_individual(
        client, headers_jefe, ciclo["id"], emp_jefe.empleado_id, jefe.empleado_id,
        "Meta de jefe", peso=100,
    )
    await _crear_meta_individual(
        client, headers_otro_jefe, ciclo["id"], emp_otro.empleado_id, otro_jefe.empleado_id,
        "Meta de otro jefe", peso=100,
    )

    def _filas(resp) -> int:
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Metas y avance"]
        count = 0
        row = 4
        while ws.cell(row=row, column=2).value is not None:
            count += 1
            row += 1
        return count

    resp_rh = await client.get(f"{BASE}/ciclos/{ciclo['id']}/export/excel", headers=headers_rh)
    assert resp_rh.status_code == 200, resp_rh.text
    assert _filas(resp_rh) == 2

    resp_jefe = await client.get(
        f"{BASE}/ciclos/{ciclo['id']}/export/excel", headers=headers_jefe
    )
    assert resp_jefe.status_code == 200, resp_jefe.text
    assert _filas(resp_jefe) == 1
