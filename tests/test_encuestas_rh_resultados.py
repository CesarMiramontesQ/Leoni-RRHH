# tests/test_encuestas_rh_resultados.py
"""Tests de resultados/analitica del modulo Encuestas RH (Tarea 4).

Cubre: agregaciones (promedio/distribucion likert, conteo de opciones),
regla min-N global (solo anonimas) y por segmento (anonimas y nominales),
textos barajados, "Sin dato" para segmento NULL, export Excel y el caso
borrador -> sin resultados.
"""

from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from app.core.exceptions import ConflictError, DomainValidationError, NotFoundError
from app.schemas.encuestas_rh import (
    AudienciaFiltros,
    OpcionCreate,
    PreguntaCreate,
    PublicarRequest,
    ResponderItem,
    ResponderRequest,
)
from app.services.encuestas_rh_service import EncuestasRhService
from tests.conftest import make_empleado
from tests.test_encuestas_rh_service import (
    _crear_encuesta_basica,
    _make_area,
    _make_clasificacion,
    _make_turno,
    _pregunta_likert,
    _pregunta_opcion_multiple,
    _pregunta_texto,
)

pytestmark = pytest.mark.asyncio

# Filtro de audiencia que excluye al creador (rol="rh") de la audiencia,
# para que el conteo de participantes/n quede limpio en los tests.
_SOLO_EMPLEADOS = AudienciaFiltros(roles=["empleado"])


async def _publicar_con_participantes(
    db,
    n_empleados: int,
    *,
    es_anonima: bool,
    umbral: int = 3,
    preguntas=None,
    areas: list[str] | None = None,
):
    """Crea una encuesta, la publica (audiencia filtrada a rol=empleado) y
    devuelve (service, encuesta_publicada, lista_de_empleados).

    `areas` (opcional): nombre de area por empleado (misma longitud que
    n_empleados); None = sin area (segmento "Sin dato").
    """
    creador = await make_empleado(db, rol="rh")
    empleados = []
    for i in range(n_empleados):
        empleado = await make_empleado(db, rol="empleado")
        if areas and areas[i] is not None:
            area = await _make_area(db, 7000 + i, areas[i])
            empleado.area_id = area.area_id
            await db.flush()
        empleados.append(empleado)

    service = EncuestasRhService(db)
    resp = await _crear_encuesta_basica(
        service, creador, preguntas=preguntas, es_anonima=es_anonima
    )
    # umbral se setea directo via update-like flujo: crear_encuesta no expone
    # umbral en _crear_encuesta_basica, asi que se ajusta aqui antes de publicar.
    from app.schemas.encuestas_rh import EncuestaUpdate

    await service.actualizar_encuesta(resp.id, EncuestaUpdate(umbral_minimo_respuestas=umbral))

    publicada = await service.publicar_encuesta(
        resp.id,
        PublicarRequest(
            filtros=_SOLO_EMPLEADOS, fecha_cierre_programada=date.today() + timedelta(days=7)
        ),
    )
    return service, publicada, empleados


async def _responder(service, encuesta, empleado, *, likert=None, opcion_ids=None, texto=None):
    pregunta_likert = next((p for p in encuesta.preguntas if p.tipo == "likert"), None)
    pregunta_opcion = next((p for p in encuesta.preguntas if p.tipo == "opcion_multiple"), None)
    pregunta_texto = next((p for p in encuesta.preguntas if p.tipo == "texto"), None)

    respuestas = []
    if pregunta_likert and likert is not None:
        respuestas.append(ResponderItem(pregunta_id=pregunta_likert.id, valor_likert=likert))
    if pregunta_opcion and opcion_ids is not None:
        respuestas.append(ResponderItem(pregunta_id=pregunta_opcion.id, opcion_ids=opcion_ids))
    if pregunta_texto is not None:
        respuestas.append(ResponderItem(pregunta_id=pregunta_texto.id, texto=texto))

    await service.responder(
        encuesta.id, empleado.empleado_id, ResponderRequest(respuestas=respuestas)
    )


def _preguntas_mixtas():
    return [
        _pregunta_likert(orden=1, requerida=True),
        _pregunta_opcion_multiple(orden=2, requerida=True, n_opciones=2),
        _pregunta_texto(orden=3, requerida=False),
    ]


# ── Agregaciones correctas (calculadas a mano) ──────────────────────────────


async def test_agregaciones_likert_opciones_y_texto_calculadas_a_mano(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 4, es_anonima=False, umbral=3, preguntas=_preguntas_mixtas()
    )
    pregunta_opcion = next(p for p in encuesta.preguntas if p.tipo == "opcion_multiple")
    opcion_a, opcion_b = pregunta_opcion.opciones[0], pregunta_opcion.opciones[1]

    # Likert: 2, 3, 4, 5 -> promedio 3.5, distribucion {2:1,3:1,4:1,5:1}
    # Opcion: A,A,A,B -> A=3, B=1
    # Texto: 3 no vacios, 1 vacio (None)
    await _responder(service, encuesta, empleados[0], likert=2, opcion_ids=[opcion_a.id], texto="Comentario 1")
    await _responder(service, encuesta, empleados[1], likert=3, opcion_ids=[opcion_a.id], texto="Comentario 2")
    await _responder(service, encuesta, empleados[2], likert=4, opcion_ids=[opcion_a.id], texto="Comentario 3")
    await _responder(service, encuesta, empleados[3], likert=5, opcion_ids=[opcion_b.id], texto=None)

    resultados = await service.obtener_resultados_globales(encuesta.id)
    assert resultados.oculto_global is False
    assert resultados.n == 4
    assert resultados.total_participantes == 4
    assert resultados.tasa_respuesta == 100.0

    pregunta_likert_res = next(p for p in resultados.preguntas if p.tipo == "likert")
    assert pregunta_likert_res.n == 4
    assert pregunta_likert_res.promedio == 3.5
    dist = {d.valor: d.conteo for d in pregunta_likert_res.distribucion}
    assert dist == {1: 0, 2: 1, 3: 1, 4: 1, 5: 1}

    pregunta_opcion_res = next(p for p in resultados.preguntas if p.tipo == "opcion_multiple")
    assert pregunta_opcion_res.n == 4
    conteos = {o.opcion_id: o.conteo for o in pregunta_opcion_res.opciones}
    assert conteos == {opcion_a.id: 3, opcion_b.id: 1}

    pregunta_texto_res = next(p for p in resultados.preguntas if p.tipo == "texto")
    assert pregunta_texto_res.n == 3


# ── Min-N global (solo anonimas) ────────────────────────────────────────────


async def test_min_n_global_anonima_oculta_bajo_umbral(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 5, es_anonima=True, umbral=5, preguntas=[_pregunta_likert()]
    )
    for empleado in empleados[:3]:
        await _responder(service, encuesta, empleado, likert=4)

    resultados = await service.obtener_resultados_globales(encuesta.id)
    assert resultados.oculto_global is True
    assert resultados.n == 3
    assert resultados.umbral_minimo_respuestas == 5
    assert resultados.preguntas == []


async def test_min_n_global_anonima_visible_con_suficientes_respuestas(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 5, es_anonima=True, umbral=5, preguntas=[_pregunta_likert()]
    )
    for empleado in empleados:
        await _responder(service, encuesta, empleado, likert=4)

    resultados = await service.obtener_resultados_globales(encuesta.id)
    assert resultados.oculto_global is False
    assert resultados.n == 5
    assert len(resultados.preguntas) == 1


async def test_nominal_global_visible_pese_a_bajo_n(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 5, es_anonima=False, umbral=5, preguntas=[_pregunta_likert()]
    )
    for empleado in empleados[:2]:
        await _responder(service, encuesta, empleado, likert=3)

    resultados = await service.obtener_resultados_globales(encuesta.id)
    assert resultados.oculto_global is False
    assert resultados.n == 2
    assert len(resultados.preguntas) == 1
    assert resultados.preguntas[0].n == 2


# ── Min-N por segmento (anonimas y nominales) ───────────────────────────────


async def test_min_n_por_segmento_oculta_celda_bajo_umbral_y_muestra_la_otra(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db,
        6,
        es_anonima=True,
        umbral=3,
        preguntas=[_pregunta_likert()],
        areas=["Area Grande"] * 4 + ["Area Chica"] * 2,
    )
    # Area Grande: 4 respuestas (>= umbral 3) -> visible
    for empleado in empleados[:4]:
        await _responder(service, encuesta, empleado, likert=5)
    # Area Chica: 2 respuestas (< umbral 3) -> oculta
    for empleado in empleados[4:]:
        await _responder(service, encuesta, empleado, likert=1)

    resultados = await service.obtener_resultados_segmentos(encuesta.id, "area")
    celdas_por_nombre = {c.segmento: c for c in resultados.celdas}

    grande = celdas_por_nombre["Area Grande"]
    assert grande.oculto is False
    assert grande.n == 4
    assert grande.preguntas[0].promedio == 5.0

    chica = celdas_por_nombre["Area Chica"]
    assert chica.oculto is True
    assert chica.n == 2
    assert chica.preguntas == []


async def test_distribucion_likert_visible_en_celda_de_segmento(db):
    """El brief de Tarea 7 solo verificaba `promedio` en resultados por
    segmento; aqui se cubre tambien la `distribucion` likert completa en una
    celda visible."""
    service, encuesta, empleados = await _publicar_con_participantes(
        db,
        4,
        es_anonima=False,
        umbral=2,
        preguntas=[_pregunta_likert()],
        areas=["Area Unica"] * 4,
    )
    valores = [3, 4, 4, 5]
    for empleado, valor in zip(empleados, valores):
        await _responder(service, encuesta, empleado, likert=valor)

    resultados = await service.obtener_resultados_segmentos(encuesta.id, "area")
    celda = next(c for c in resultados.celdas if c.segmento == "Area Unica")
    assert celda.oculto is False

    pregunta = celda.preguntas[0]
    assert pregunta.promedio == 4.0
    dist = {d.valor: d.conteo for d in pregunta.distribucion}
    assert dist == {1: 0, 2: 0, 3: 1, 4: 2, 5: 1}


async def test_nominal_segmentos_ocultan_pese_a_global_visible(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db,
        4,
        es_anonima=False,
        umbral=3,
        preguntas=[_pregunta_likert()],
        areas=["Area Unica"] * 2 + ["Area Otra"] * 2,
    )
    for empleado in empleados[:2]:
        await _responder(service, encuesta, empleado, likert=4)
    for empleado in empleados[2:]:
        await _responder(service, encuesta, empleado, likert=2)

    global_ = await service.obtener_resultados_globales(encuesta.id)
    assert global_.oculto_global is False  # nominal: global nunca se oculta

    segmentos = await service.obtener_resultados_segmentos(encuesta.id, "area")
    for celda in segmentos.celdas:
        assert celda.oculto is True  # cada celda tiene n=2 < umbral=3
        assert celda.preguntas == []


# ── Segmento NULL agrupa como "Sin dato" ────────────────────────────────────


async def test_segmento_null_agrupa_como_sin_dato(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 3, es_anonima=False, umbral=2, preguntas=[_pregunta_likert()], areas=[None, None, None]
    )
    for empleado in empleados:
        await _responder(service, encuesta, empleado, likert=3)

    resultados = await service.obtener_resultados_segmentos(encuesta.id, "area")
    assert len(resultados.celdas) == 1
    assert resultados.celdas[0].segmento == "Sin dato"
    assert resultados.celdas[0].n == 3


# ── Textos: barajados y sujetos a la regla global ───────────────────────────


async def test_textos_barajados_como_set_y_regla_min_n(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 5, es_anonima=True, umbral=5, preguntas=[_pregunta_texto()]
    )
    pregunta_id = encuesta.preguntas[0].id

    # Solo 3 respuestas (< umbral 5) -> oculto
    for i, empleado in enumerate(empleados[:3]):
        await _responder(service, encuesta, empleado, texto=f"Texto {i}")

    textos_ocultos = await service.obtener_textos(encuesta.id, pregunta_id)
    assert textos_ocultos.oculto is True
    assert textos_ocultos.textos == []
    assert textos_ocultos.n == 3

    # Completar a 5 respuestas -> visible, barajado (comparar como set)
    for i, empleado in enumerate(empleados[3:], start=3):
        await _responder(service, encuesta, empleado, texto=f"Texto {i}")

    textos_visibles = await service.obtener_textos(encuesta.id, pregunta_id)
    assert textos_visibles.oculto is False
    assert set(textos_visibles.textos) == {f"Texto {i}" for i in range(5)}
    assert textos_visibles.n == 5


async def test_textos_nominal_siempre_visibles(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 5, es_anonima=False, umbral=5, preguntas=[_pregunta_texto()]
    )
    pregunta_id = encuesta.preguntas[0].id
    for i, empleado in enumerate(empleados[:2]):
        await _responder(service, encuesta, empleado, texto=f"Nominal {i}")

    textos = await service.obtener_textos(encuesta.id, pregunta_id)
    assert textos.oculto is False
    assert set(textos.textos) == {"Nominal 0", "Nominal 1"}


async def test_obtener_textos_de_pregunta_no_texto_falla(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db, 2, es_anonima=False, umbral=1, preguntas=[_pregunta_likert()]
    )
    pregunta_id = encuesta.preguntas[0].id
    with pytest.raises(DomainValidationError):
        await service.obtener_textos(encuesta.id, pregunta_id)


# ── Export Excel ─────────────────────────────────────────────────────────


async def test_export_excel_hojas_y_valores_sin_fuga_de_segmento_oculto(db):
    service, encuesta, empleados = await _publicar_con_participantes(
        db,
        6,
        es_anonima=True,
        umbral=3,
        preguntas=[_pregunta_likert()],
        areas=["Area Grande"] * 4 + ["Area Chica"] * 2,
    )
    for empleado in empleados[:4]:
        await _responder(service, encuesta, empleado, likert=5)
    for empleado in empleados[4:]:
        await _responder(service, encuesta, empleado, likert=1)

    output, filename = await service.exportar_resultados_excel(encuesta.id)
    assert filename.startswith("resultados_encuesta_")
    assert filename.endswith(".xlsx")

    wb = load_workbook(output)
    assert wb.sheetnames == ["Resumen", "Preguntas", "Segmentos", "Textos"]

    ws_resumen = wb["Resumen"]
    valores_resumen = {ws_resumen.cell(row=r, column=1).value: ws_resumen.cell(row=r, column=2).value for r in range(3, 11)}
    assert valores_resumen["Total de respuestas (n)"] == 6

    ws_seg = wb["Segmentos"]
    filas = [
        [ws_seg.cell(row=r, column=c).value for c in range(1, 7)]
        for r in range(2, ws_seg.max_row + 1)
    ]
    area_rows = [f for f in filas if f[0] == "area"]
    grande_row = next(f for f in area_rows if f[1] == "Area Grande")
    chica_row = next(f for f in area_rows if f[1] == "Area Chica")

    assert grande_row[3] == "No"  # no oculto
    assert grande_row[5] is not None and "promedio=5.0" in grande_row[5]

    assert chica_row[3] == "Si"  # oculto: SIN metricas
    assert chica_row[4] is not None and "Oculto" in chica_row[4]  # placeholder, no pregunta real
    assert chica_row[5] is None  # columna "Metrica" vacia (sin fuga de datos)


# ── Borrador: sin resultados ─────────────────────────────────────────────


async def test_borrador_sin_resultados_conflicto(db):
    creador = await make_empleado(db, rol="rh")
    service = EncuestasRhService(db)
    resp = await _crear_encuesta_basica(service, creador, preguntas=[_pregunta_likert()])

    with pytest.raises(ConflictError):
        await service.obtener_resultados_globales(resp.id)
    with pytest.raises(ConflictError):
        await service.obtener_resultados_segmentos(resp.id, "area")
    with pytest.raises(ConflictError):
        await service.obtener_textos(resp.id, resp.preguntas[0].id)


async def test_encuesta_inexistente_404(db):
    service = EncuestasRhService(db)
    with pytest.raises(NotFoundError):
        await service.obtener_resultados_globales(999999)


async def test_dimension_invalida_422(db):
    service, encuesta, _ = await _publicar_con_participantes(
        db, 2, es_anonima=False, umbral=1, preguntas=[_pregunta_likert()]
    )
    with pytest.raises(DomainValidationError):
        await service.obtener_resultados_segmentos(encuesta.id, "invalida")
