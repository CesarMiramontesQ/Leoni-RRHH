"""El export nunca falla por culpa del bloque externo: si DATOS_ANALISIS no
responde, la columna del indice objetivo queda vacia con una nota."""
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from app.services.talento_service import (
    AreaCapacitacion,
    AreaPdi,
    AreaPolivalencia,
    BloqueCapacitacion,
    BloqueDesempeno,
    BloqueObjetivo,
    BloquePdi,
    BloquePolivalencia,
    EmpleadoFoco,
    TalentoService,
)
from tests.conftest import make_empleado


def _bloques_stub():
    return {
        "pol": BloquePolivalencia(
            disponible=True, org=None,
            areas=[AreaPolivalencia(1, "Arneses A", 40, 70.0, 60.0, 2, "ambar")],
        ),
        "cap": BloqueCapacitacion(
            disponible=True, org=None,
            areas=[AreaCapacitacion(1, "Arneses A", 10, 9, 90.0, 1, "verde")],
        ),
        "pdi": BloquePdi(
            disponible=True, org=None,
            areas=[AreaPdi(1, "Arneses A", 6, 4, 0, 66.7, 1, 1, "ambar")],
        ),
        "des": BloqueDesempeno(disponible=False, motivo="sin_ciclo"),
    }


@pytest.mark.asyncio
async def test_export_tiene_dos_hojas_y_datos(db):
    rh = await make_empleado(
        db, rol="rh", email="tal_exp@leoni.test",
        modulos_rh={"dashboard-talento": True}, inscrito_modulos_rh=True,
    )
    svc = TalentoService(db)
    b = _bloques_stub()
    foco = [EmpleadoFoco(5, 500, "Ana", "Crimpado", ["desempeno_bajo", "pdi_vencido"])]
    bloque_pol_mock = AsyncMock(return_value=b["pol"])
    foco_mock = AsyncMock(return_value=foco)
    with patch.object(TalentoService, "bloque_polivalencia", bloque_pol_mock), \
         patch.object(TalentoService, "bloque_capacitacion", AsyncMock(return_value=b["cap"])), \
         patch.object(TalentoService, "bloque_pdi", AsyncMock(return_value=b["pdi"])), \
         patch.object(TalentoService, "bloque_desempeno", AsyncMock(return_value=b["des"])), \
         patch.object(TalentoService, "bloque_objetivo", AsyncMock(return_value=BloqueObjetivo(disponible=True, areas=[]))), \
         patch.object(TalentoService, "_empleados_foco_area", foco_mock):
        output = await svc.exportar_excel(rh, None, None)

    wb = load_workbook(output)
    assert wb.sheetnames == ["Resumen por area", "Empleados en foco"]
    resumen = wb["Resumen por area"]
    assert resumen.cell(row=2, column=1).value == "Arneses A"
    hoja_foco = wb["Empleados en foco"]
    assert hoja_foco.cell(row=2, column=3).value == "Ana"
    # El fix del loop: 1 sola area -> 1 sola llamada al helper de foco (y a
    # bloque_polivalencia, que solo se pide UNA vez, no una por area). Antes
    # el loop llamaba a detalle_area por area, que recomputaba los tres
    # bloques org-wide en cada vuelta.
    assert foco_mock.call_count == 1
    assert bloque_pol_mock.call_count == 1


@pytest.mark.asyncio
async def test_export_sobrevive_a_datos_analisis_caido(db):
    rh = await make_empleado(
        db, rol="rh", email="tal_exp2@leoni.test",
        modulos_rh={"dashboard-talento": True}, inscrito_modulos_rh=True,
    )
    svc = TalentoService(db)
    b = _bloques_stub()
    with patch.object(TalentoService, "bloque_polivalencia", AsyncMock(return_value=b["pol"])), \
         patch.object(TalentoService, "bloque_capacitacion", AsyncMock(return_value=b["cap"])), \
         patch.object(TalentoService, "bloque_pdi", AsyncMock(return_value=b["pdi"])), \
         patch.object(TalentoService, "bloque_desempeno", AsyncMock(return_value=b["des"])), \
         patch.object(TalentoService, "bloque_objetivo", AsyncMock(side_effect=RuntimeError("caido"))), \
         patch.object(TalentoService, "_empleados_foco_area", AsyncMock(return_value=[])):
        output = await svc.exportar_excel(rh, None, None)

    wb = load_workbook(output)
    resumen = wb["Resumen por area"]
    # Columna del indice objetivo (6a) con la nota, no una excepcion.
    assert resumen.cell(row=2, column=6).value == "no disponible"
