"""Tests del service de cobertura/polivalencia por area (Task 2).

Los seams `_puestos_con_area` y `CompetenciaService.obtener_multihabilidades`
se mockean para probar agregacion y scope sin sembrar el grafo de
competencias (puestos, requisitos, grados).

Los tests de scope explicito/polivalencia por empleado (`_area_con_dos_
empleados`) si siembran el grafo real via los factories de talento: ahi lo
que se prueba es el filtrado por scope y el dedup por empleado, no la
agregacion de `obtener_multihabilidades` (ya cubierta arriba con mocks).
"""
from unittest.mock import AsyncMock, patch

import pytest

from app.services.operaciones_service import OperacionesService, PuestoArea


async def _area_con_dos_empleados(db) -> dict:
    """Area con 1 puesto, 1 competencia requisito y 2 empleados asignados.

    Escenario base (datos reales, sin mocks) para los tests de
    `listar_areas_con_scope` y `polivalencia_empleados_area`."""
    from tests.conftest import make_empleado
    from tests.conftest_talento import (
        make_area,
        make_competencia,
        make_competencia_requisito,
        make_perfil_funciones,
        make_puesto_perfil,
    )

    area = await make_area(db, descripcion="Ensamble Test")
    puesto = await make_puesto_perfil(db, nombre="Crimpado", area_id=area.area_id)
    competencia = await make_competencia(
        db, nombre="Crimpado manual", categoria="tecnica"
    )
    await make_competencia_requisito(
        db, competencia_id=competencia.id, puesto_perfil_id=puesto.id, nivel_requerido=3
    )
    empleado_a = await make_empleado(db, nombre="Ana")
    empleado_b = await make_empleado(db, nombre="Beto")
    await make_perfil_funciones(
        db, puesto_perfil_id=puesto.id, empleado_id=empleado_a.empleado_id
    )
    await make_perfil_funciones(
        db, puesto_perfil_id=puesto.id, empleado_id=empleado_b.empleado_id
    )
    return {
        "area": area,
        "puesto": puesto,
        "empleado_a": empleado_a,
        "empleado_b": empleado_b,
    }


def _multihab(puesto_id, puesto_nombre, competencias, empleados):
    """Construye un objeto tipo MultihabilidadesResponse (los reales de
    app.schemas.talento). empleados: list[(eid, nombre, niveles, requisitos)]."""
    from app.schemas.talento import (
        MultihabilidadesResponse,
        MultihabilidadesCompetenciaItem,
        MultihabilidadesEmpleadoItem,
    )
    return MultihabilidadesResponse(
        puesto_perfil_id=puesto_id,
        puesto_nombre=puesto_nombre,
        competencias=[
            MultihabilidadesCompetenciaItem(
                competencia_id=c[0], competencia_nombre=c[1],
                tipo_competencia_id=1, tipo_nombre=c[2], nivel_requerido=0,
            ) for c in competencias
        ],
        empleados=[
            MultihabilidadesEmpleadoItem(
                empleado_id=e[0], nombre=e[1], no_empleado=e[0],
                grado_id=1, grado_nombre="", niveles=e[2], requisitos=e[3],
            ) for e in empleados
        ],
        metodos_calificacion=[],
    )


@pytest.mark.asyncio
async def test_cobertura_area_agrega_y_detecta_criticas(db):
    svc = OperacionesService(db)
    # area 5 con 1 puesto; comp 10 cubierta por 1 (punto_unico), comp 20 hueco
    puestos = [PuestoArea(puesto_perfil_id=1, puesto_nombre="Crimpado", area_id=5, area_nombre="Ensamble")]
    multihab = _multihab(
        1, "Crimpado",
        competencias=[(10, "Crimpado manual", "Operacion"), (20, "LOTO", "Seguridad")],
        empleados=[
            (100, "Ana", {10: 3, 20: 1}, {10: 3, 20: 3}),
            (101, "Beto", {10: 1, 20: 0}, {10: 3, 20: 3}),
        ],
    )
    with patch.object(svc, "_puestos_con_area", AsyncMock(return_value=puestos)), \
         patch("app.services.competencia_service.CompetenciaService.obtener_multihabilidades",
               AsyncMock(return_value=multihab)), \
         patch("app.services.operaciones_service.empleado_ids_scope_por_modulo",
               AsyncMock(return_value=None)):
        res = await svc.cobertura_area(current_user=object(), area_id=5, rh_ui_mode=None)
    cobs = {c.competencia_id: c for c in res.competencias}
    assert cobs[10].cubren == 1 and cobs[10].severidad == "punto_unico"
    assert cobs[20].cubren == 0 and cobs[20].severidad == "hueco"
    # criticas: ambas; candidatos de comp 20 = Ana(1) y Beto(0) ordenados por nivel desc
    crit = {c.competencia_id: c for c in res.criticas}
    assert crit[20].candidatos[0].empleado_id == 100  # Ana nivel 1 primero


@pytest.mark.asyncio
async def test_scope_supervisor_filtra_personal(db):
    svc = OperacionesService(db)
    puestos = [PuestoArea(1, "Crimpado", 5, "Ensamble")]
    multihab = _multihab(
        1, "Crimpado", [(10, "Crimpado manual", "Op")],
        empleados=[
            (100, "Ana", {10: 3}, {10: 3}),
            (101, "Beto", {10: 0}, {10: 3}),
        ],
    )
    with patch.object(svc, "_puestos_con_area", AsyncMock(return_value=puestos)), \
         patch("app.services.competencia_service.CompetenciaService.obtener_multihabilidades",
               AsyncMock(return_value=multihab)), \
         patch("app.services.operaciones_service.empleado_ids_scope_por_modulo",
               AsyncMock(return_value=[100])):  # supervisor solo ve a Ana
        res = await svc.cobertura_area(current_user=object(), area_id=5, rh_ui_mode=None)
    cob = res.competencias[0]
    assert cob.requieren == 1 and cob.cubren == 1  # solo Ana entra al calculo


@pytest.mark.asyncio
async def test_area_fuera_de_scope_403(db):
    from app.core.exceptions import ForbiddenError
    svc = OperacionesService(db)
    puestos = [PuestoArea(1, "Crimpado", 5, "Ensamble")]
    multihab = _multihab(1, "Crimpado", [(10, "X", "Op")],
                         empleados=[(100, "Ana", {10: 3}, {10: 3})])
    with patch.object(svc, "_puestos_con_area", AsyncMock(return_value=puestos)), \
         patch("app.services.competencia_service.CompetenciaService.obtener_multihabilidades",
               AsyncMock(return_value=multihab)), \
         patch("app.services.operaciones_service.empleado_ids_scope_por_modulo",
               AsyncMock(return_value=[999])):  # scope sin nadie del area
        with pytest.raises(ForbiddenError):
            await svc.cobertura_area(current_user=object(), area_id=5, rh_ui_mode=None)


@pytest.mark.asyncio
async def test_listar_areas_ordena_por_criticas(db):
    svc = OperacionesService(db)
    puestos = [
        PuestoArea(1, "P1", 5, "Area A"),
        PuestoArea(2, "P2", 6, "Area B"),
    ]

    def _obtener(puesto_id, nombre_filtro=None):
        if puesto_id == 1:  # area A: 1 hueco
            return _multihab(1, "P1", [(10, "X", "Op")],
                             empleados=[(100, "Ana", {10: 0}, {10: 3})])
        return _multihab(2, "P2", [(20, "Y", "Op")],  # area B: cubierta
                         empleados=[(200, "Cid", {20: 3}, {20: 3}), (201, "Dan", {20: 3}, {20: 3})])

    with patch.object(svc, "_puestos_con_area", AsyncMock(return_value=puestos)), \
         patch("app.services.competencia_service.CompetenciaService.obtener_multihabilidades",
               AsyncMock(side_effect=_obtener)), \
         patch("app.services.operaciones_service.empleado_ids_scope_por_modulo",
               AsyncMock(return_value=None)):
        areas = await svc.listar_areas(current_user=object(), rh_ui_mode=None)
    assert [a.area_id for a in areas] == [5, 6]  # A (1 critica) antes que B (0)
    assert areas[0].n_criticas == 1 and areas[1].n_criticas == 0


@pytest.mark.asyncio
async def test_export_area_genera_xlsx(db, monkeypatch):
    from io import BytesIO
    from openpyxl import load_workbook
    from app.services.operaciones_service import (
        AreaResumen, CoberturaArea, Critica, OperacionesService, PuestoCobertura,
    )
    from app.services.operaciones.types import CandidatoCrossTrain, CoberturaCompetencia

    svc = OperacionesService(db)
    candidato = CandidatoCrossTrain(
        empleado_id=100, no_empleado=100, nombre="Ana Garcia",
        nivel_actual=1, nivel_requerido=3
    )
    critica = Critica(
        competencia_id=20, competencia_nombre="LOTO",
        severidad="punto_unico", candidatos=[candidato]
    )
    fake = CoberturaArea(
        resumen=AreaResumen(5, "Ensamble", 75.0, 50.0, 1, 2),
        competencias=[CoberturaCompetencia(10, "Crimpado", "Op", 2, 1, 1, 50.0, "ambar", "punto_unico")],
        puestos=[PuestoCobertura(1, "Crimpado", [])],
        criticas=[critica],
    )
    monkeypatch.setattr(svc, "cobertura_area", AsyncMock(return_value=fake))
    out = await svc.exportar_area_excel(current_user=object(), area_id=5, rh_ui_mode=None)
    data = out.getvalue()
    assert data[:2] == b"PK" and len(data) > 100  # xlsx = zip, no vacio

    # Verificar que el candidato aparece en la hoja Cross-training
    wb = load_workbook(BytesIO(data))
    ws = wb["Cross-training"]
    valores = [c.value for fila in ws.iter_rows() for c in fila]
    assert "Ana Garcia" in valores


@pytest.mark.asyncio
async def test_listar_areas_con_scope_filtra_sin_tocar_current_user(db):
    """`listar_areas_con_scope` recibe los ids ya resueltos: no consulta rol ni
    modulo. Es el punto de entrada que usa el Dashboard de Talento."""
    from app.services.operaciones_service import OperacionesService

    datos = await _area_con_dos_empleados(db)  # helper existente del archivo
    svc = OperacionesService(db)

    todos = await svc.listar_areas_con_scope(None)
    assert todos and todos[0].n_empleados == 2

    uno = await svc.listar_areas_con_scope([datos["empleado_a"].empleado_id])
    assert uno[0].n_empleados == 1


@pytest.mark.asyncio
async def test_polivalencia_empleados_area_devuelve_indice_por_persona(db):
    from app.services.operaciones_service import OperacionesService

    datos = await _area_con_dos_empleados(db)
    svc = OperacionesService(db)

    filas = await svc.polivalencia_empleados_area(datos["area"].area_id, None)
    por_id = {f.empleado_id: f for f in filas}
    assert set(por_id) == {
        datos["empleado_a"].empleado_id,
        datos["empleado_b"].empleado_id,
    }
    assert all(f.pol_pct is None or 0.0 <= f.pol_pct <= 100.0 for f in filas)
    assert all(f.nombre for f in filas)


@pytest.mark.asyncio
async def test_polivalencia_empleados_area_respeta_scope(db):
    from app.services.operaciones_service import OperacionesService

    datos = await _area_con_dos_empleados(db)
    svc = OperacionesService(db)

    filas = await svc.polivalencia_empleados_area(
        datos["area"].area_id, [datos["empleado_a"].empleado_id]
    )
    assert [f.empleado_id for f in filas] == [datos["empleado_a"].empleado_id]
