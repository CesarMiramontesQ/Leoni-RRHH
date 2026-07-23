# tests/test_historial_objetivo_api.py
"""Tests HTTP del modulo Historial Objetivo (Tarea 5): router + registro de
modulo + scoping por equipo + self-service + export.

La logica de dominio (formula, agregacion de fuentes, scoping) ya esta
cubierta a nivel service en tests/test_historial_objetivo_service.py
(Tarea 4); aqui se ejercita el router end-to-end via `client`: permisos
(RH por modulo 'historial-objetivo', jefe/director con scoping resuelto
por el service, self-service), default de rango de fechas (ultimos 12
meses), serializacion y mapeo de errores.

Mockea el engine/repos de bono (NUNCA toca la BD externa real) con el mismo
patron que `tests/test_historial_objetivo_service.py`.
"""

from contextlib import contextmanager
from datetime import date, timedelta
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.core.rh_module_registry import is_rh_self_service_api_path, is_valid_module_key
from app.models.actas import ActaAdministrativa
from app.services.incidencia_fuentes.constants import TIPO_INCIDENCIA_CALIDAD
from tests.conftest import auth_headers, make_empleado

pytestmark = pytest.mark.asyncio

BASE = "/api/v1/historial-objetivo"


async def _crear_acta(db, *, empleado_id: int, generado_por: int, estado: str = "signed"):
    acta = ActaAdministrativa(
        empleado_id=empleado_id,
        generado_por=generado_por,
        estado=estado,
        descripcion_hechos="Hechos de prueba",
        tipo_falta="Falta de prueba",
    )
    db.add(acta)
    await db.flush()
    await db.refresh(acta)
    return acta


@contextmanager
def _mock_bono_repos(
    *,
    incidencias_raw: list | None = None,
    faltas_raw: list | None = None,
    engine_configurado: bool = True,
):
    """Patchea `create_read_engine` + ambos repos de bono en el namespace del
    service (mismo patron que `tests/test_historial_objetivo_service.py`) --
    nunca toca la BD externa real."""
    mock_engine = MagicMock()
    mock_engine.dispose = AsyncMock()

    inc_mock = AsyncMock()
    inc_mock.aggregate_empleados_top_por_tipo = AsyncMock(
        return_value=incidencias_raw if incidencias_raw is not None else []
    )

    falt_mock = AsyncMock()
    falt_mock.aggregate_empleados_top_por_tipo = AsyncMock(
        return_value=faltas_raw if faltas_raw is not None else []
    )

    with (
        patch(
            "app.services.historial_objetivo_service.BonoProductividadReadClient.create_read_engine",
            return_value=(mock_engine if engine_configurado else None),
        ),
        patch(
            "app.services.historial_objetivo_service.BonoHistoricoIncidenciasRepository",
            return_value=inc_mock,
        ),
        patch(
            "app.services.historial_objetivo_service.BonoFaltasRetardosRepository",
            return_value=falt_mock,
        ),
    ):
        yield mock_engine, inc_mock, falt_mock


# ══════════════════════════════════════════════════════════════════════════
# GET /empleados/{empleado_id}
# ══════════════════════════════════════════════════════════════════════════
async def test_indice_empleado_en_scope_200_con_indice_y_desglose(client, db):
    rh = await make_empleado(db, rol="rh", email="ho_api_rh1@leoni.test")
    emp = await make_empleado(db, rol="empleado", email="ho_api_emp1@leoni.test")
    await _crear_acta(db, empleado_id=emp.empleado_id, generado_por=rh.empleado_id, estado="signed")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/empleados/{emp.empleado_id}", headers=headers)

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["empleado_id"] == emp.empleado_id
    assert body["bono_disponible"] is True
    assert body["resultado"]["indice"] == 85.0
    assert body["resultado"]["semaforo"] == "verde"
    fuentes = {d["fuente"] for d in body["resultado"]["desglose"]}
    assert fuentes == {"actas", "faltas", "incidencias", "progresivo"}


async def test_indice_empleado_fuera_de_scope_403(client, db):
    jefe_a = await make_empleado(db, rol="supervisor", email="ho_api_jefeA@leoni.test")
    jefe_b = await make_empleado(db, rol="supervisor", email="ho_api_jefeB@leoni.test")
    emp_de_b = await make_empleado(
        db, rol="empleado", email="ho_api_empB@leoni.test", lider_id=jefe_b.empleado_id
    )
    headers = await auth_headers(client, jefe_a)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/empleados/{emp_de_b.empleado_id}", headers=headers)

    assert res.status_code == 403, res.text


async def test_indice_empleado_inexistente_404(client, db):
    rh = await make_empleado(db, rol="rh", email="ho_api_rh2@leoni.test")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/empleados/999999999", headers=headers)

    assert res.status_code == 404, res.text


async def test_rango_fechas_invalido_422(client, db):
    rh = await make_empleado(db, rol="rh", email="ho_api_rh3@leoni.test")
    emp = await make_empleado(db, rol="empleado", email="ho_api_emp3@leoni.test")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos():
        res = await client.get(
            f"{BASE}/empleados/{emp.empleado_id}",
            params={"fecha_inicio": "2026-06-01", "fecha_fin": "2026-01-01"},
            headers=headers,
        )

    assert res.status_code == 422, res.text


async def test_bono_no_configurado_degrada_a_solo_actas(client, db):
    rh = await make_empleado(db, rol="rh", email="ho_api_rh4@leoni.test")
    emp = await make_empleado(db, rol="empleado", email="ho_api_emp4@leoni.test")
    await _crear_acta(db, empleado_id=emp.empleado_id, generado_por=rh.empleado_id, estado="signed")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos(engine_configurado=False):
        res = await client.get(f"{BASE}/empleados/{emp.empleado_id}", headers=headers)

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["bono_disponible"] is False
    assert body["resultado"]["indice"] == 85.0


# ══════════════════════════════════════════════════════════════════════════
# GET /equipo
# ══════════════════════════════════════════════════════════════════════════
async def test_equipo_jefe_ve_solo_su_equipo(client, db):
    jefe = await make_empleado(db, rol="supervisor", email="ho_api_jefe1@leoni.test")
    otro_jefe = await make_empleado(db, rol="supervisor", email="ho_api_jefe2@leoni.test")
    emp_propio = await make_empleado(
        db, rol="empleado", email="ho_api_propio@leoni.test", lider_id=jefe.empleado_id
    )
    emp_ajeno = await make_empleado(
        db, rol="empleado", email="ho_api_ajeno@leoni.test", lider_id=otro_jefe.empleado_id
    )
    headers = await auth_headers(client, jefe)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo", headers=headers)

    assert res.status_code == 200, res.text
    ids = {item["empleado_id"] for item in res.json()["items"]}
    assert jefe.empleado_id in ids
    assert emp_propio.empleado_id in ids
    assert emp_ajeno.empleado_id not in ids


async def test_equipo_sin_fechas_usa_default_ultimos_12_meses_no_da_422(client, db):
    """RH sin equipo delimitado (scope universo) exige rango de fechas a
    nivel service -- este router debe completar el default (ultimos 12
    meses) para que la llamada sin querystring nunca truene con 422."""
    rh = await make_empleado(db, rol="rh", email="ho_api_rh5@leoni.test")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo", headers=headers)

    assert res.status_code == 200, res.text


async def test_no_rh_con_modulo_otorgado_ve_universo_en_equipo(client, db):
    """No-RH con el modulo 'historial-objetivo' otorgado -> scope elevado a
    'rh' (universo, vista "top offenders" -- ver `HistorialObjetivoService.
    indice_equipo`, solo incluye empleados con algun evento en el rango),
    no solo su equipo -- verifica que el registro del modulo en RH_MODULES
    quedo conectado end-to-end."""
    grantee = await make_empleado(
        db,
        rol="supervisor",
        email="ho_api_grant@leoni.test",
        modulos_rh={"historial-objetivo": True},
        inscrito_modulos_rh=True,
    )
    ajeno = await make_empleado(db, rol="empleado", email="ho_api_grant_ajeno@leoni.test")
    # Sin lider_id/reporte hacia `grantee`: solo aparece en su ranking si el
    # scope se elevo a universo (con lider_id normal, un supervisor sin
    # reportes directos vería equipo vacío).
    await _crear_acta(db, empleado_id=ajeno.empleado_id, generado_por=grantee.empleado_id, estado="signed")
    headers = await auth_headers(client, grantee)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo", headers=headers)

    assert res.status_code == 200, res.text
    ids = {item["empleado_id"] for item in res.json()["items"]}
    assert ajeno.empleado_id in ids


async def test_no_rh_con_modulo_otorgado_ve_universo_en_equipo_con_rango_de_fechas_explicito(
    client, db
):
    """Hueco T8 #2: variante de `test_no_rh_con_modulo_otorgado_ve_universo_
    en_equipo` con `fecha_inicio`/`fecha_fin` explicitos en vez del default
    de 12 meses -- confirma que la elevacion de scope por permiso RH tambien
    funciona end-to-end cuando el cliente manda su propio rango."""
    grantee = await make_empleado(
        db,
        rol="supervisor",
        email="ho_api_grant_rango@leoni.test",
        modulos_rh={"historial-objetivo": True},
        inscrito_modulos_rh=True,
    )
    ajeno = await make_empleado(db, rol="empleado", email="ho_api_grant_rango_ajeno@leoni.test")
    await _crear_acta(db, empleado_id=ajeno.empleado_id, generado_por=grantee.empleado_id, estado="signed")
    headers = await auth_headers(client, grantee)

    with _mock_bono_repos():
        res = await client.get(
            f"{BASE}/equipo",
            params={"fecha_inicio": "2026-01-01", "fecha_fin": "2026-12-31"},
            headers=headers,
        )

    assert res.status_code == 200, res.text
    ids = {item["empleado_id"] for item in res.json()["items"]}
    assert ajeno.empleado_id in ids


async def test_equipo_degrada_a_solo_actas_si_bono_no_configurado(client, db):
    """Hueco T8 #3: la degradacion de bono no solo aplica en
    `GET /empleados/{id}` (`test_bono_no_configurado_degrada_a_solo_actas`) --
    tambien debe funcionar en `GET /equipo`: con `create_read_engine` en
    None, el ranking se calcula solo con actas sin crashear y el flag
    `bono_disponible` (a nivel de la respuesta, no por item) viaja en False."""
    jefe = await make_empleado(db, rol="supervisor", email="ho_api_equipo_sin_bono_jefe@leoni.test")
    emp = await make_empleado(
        db, rol="empleado", email="ho_api_equipo_sin_bono_emp@leoni.test", lider_id=jefe.empleado_id
    )
    await _crear_acta(db, empleado_id=emp.empleado_id, generado_por=jefe.empleado_id, estado="signed")
    headers = await auth_headers(client, jefe)

    with _mock_bono_repos(engine_configurado=False):
        res = await client.get(f"{BASE}/equipo", headers=headers)

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["bono_disponible"] is False
    por_id = {item["empleado_id"]: item for item in body["items"]}
    assert por_id[emp.empleado_id]["resultado"]["indice"] == 85.0
    assert por_id[jefe.empleado_id]["resultado"]["indice"] == 100.0


# ══════════════════════════════════════════════════════════════════════════
# GET /equipo/export
# ══════════════════════════════════════════════════════════════════════════
async def test_equipo_export_excel(client, db):
    rh = await make_empleado(db, rol="rh", email="ho_api_rh6@leoni.test")
    emp = await make_empleado(db, rol="empleado", email="ho_api_emp6@leoni.test")
    headers = await auth_headers(client, rh)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo/export", headers=headers)

    assert res.status_code == 200, res.text
    assert res.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=historial_objetivo_equipo_" in res.headers["content-disposition"]


async def test_equipo_export_excel_contenido_hoja_y_filas(client, db):
    """Hueco T8 #1: hasta ahora solo se validaba content-type/disposition --
    aqui se abre el .xlsx real con openpyxl y se verifica la hoja + al menos
    una fila con empleado/indice/semaforo (mismo formato que arma el router
    en `export_equipo_excel`: empleado_id, no_empleado, nombre, indice,
    semaforo, penalizacion_total)."""
    from io import BytesIO

    from openpyxl import load_workbook

    jefe = await make_empleado(db, rol="supervisor", email="ho_api_export_jefe@leoni.test")
    emp = await make_empleado(
        db, rol="empleado", email="ho_api_export_emp@leoni.test", lider_id=jefe.empleado_id
    )
    await _crear_acta(db, empleado_id=emp.empleado_id, generado_por=jefe.empleado_id, estado="signed")
    headers = await auth_headers(client, jefe)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo/export", headers=headers)

    assert res.status_code == 200, res.text

    wb = load_workbook(BytesIO(res.content))
    ws = wb["Historial Objetivo"]

    header = [c.value for c in ws[1]]
    assert header == [
        "empleado_id",
        "no_empleado",
        "nombre",
        "indice",
        "semaforo",
        "penalizacion_total",
    ]

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    por_empleado_id = {fila[0]: fila for fila in filas}
    assert jefe.empleado_id in por_empleado_id
    assert emp.empleado_id in por_empleado_id

    # emp tiene un acta signed -> indice 85.0 / verde (solo actas, sin bono
    # en este mock, ver `_mock_bono_repos` default engine_configurado=True
    # pero sin incidencias/faltas -> solo penaliza el acta).
    fila_emp = por_empleado_id[emp.empleado_id]
    assert fila_emp[2] == emp.nombre
    assert fila_emp[3] == 85.0
    assert fila_emp[4] == "verde"


# ══════════════════════════════════════════════════════════════════════════
# GET /mi-historial (self-service)
# ══════════════════════════════════════════════════════════════════════════
async def test_mi_historial_usa_el_token_ignora_query(client, db):
    emp = await make_empleado(db, rol="empleado", email="ho_api_mi1@leoni.test")
    otro = await make_empleado(db, rol="empleado", email="ho_api_mi2@leoni.test")
    headers = await auth_headers(client, emp)

    with _mock_bono_repos():
        res = await client.get(
            f"{BASE}/mi-historial", params={"empleado_id": otro.empleado_id}, headers=headers
        )

    assert res.status_code == 200, res.text
    assert res.json()["empleado_id"] == emp.empleado_id


async def test_mi_historial_es_self_service_registrado():
    assert is_rh_self_service_api_path(f"{BASE}/mi-historial")


async def test_rh_restringido_sin_modulo_403_pero_mi_historial_ok(client, db):
    """RH con `modulos_rh` explicito que NO incluye 'historial-objetivo' es
    bloqueado por el middleware de permisos en gestion, pero conserva su
    self-service."""
    rh_restringido = await make_empleado(
        db,
        rol="rh",
        email="ho_api_rh_restr@leoni.test",
        modulos_rh={"incidencias": True},
    )
    headers = await auth_headers(client, rh_restringido)

    with _mock_bono_repos():
        res_equipo = await client.get(f"{BASE}/equipo", headers=headers)
        res_mi = await client.get(f"{BASE}/mi-historial", headers=headers)

    assert res_equipo.status_code == 403, res_equipo.text
    assert res_mi.status_code == 200, res_mi.text


async def test_rh_con_modulo_historial_objetivo_200_en_equipo(client, db):
    rh_con_modulo = await make_empleado(
        db,
        rol="rh",
        email="ho_api_rh_con_modulo@leoni.test",
        modulos_rh={"historial-objetivo": True},
    )
    headers = await auth_headers(client, rh_con_modulo)

    with _mock_bono_repos():
        res = await client.get(f"{BASE}/equipo", headers=headers)

    assert res.status_code == 200, res.text


# ══════════════════════════════════════════════════════════════════════════
# Registro del modulo
# ══════════════════════════════════════════════════════════════════════════
async def test_modulo_historial_objetivo_registrado():
    assert is_valid_module_key("historial-objetivo")
